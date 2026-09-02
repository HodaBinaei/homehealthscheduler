import pickle
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
os.makedirs('./output', exist_ok=True)

with open('./roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

# Each carer's own "I worked this weekday" calendar -- the denominator for judging
# consistency, so weeks they simply weren't on duty (leave, part-time pattern, etc.)
# never count against a relationship.
carer_weekday_active_weeks = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))

def classify_relative(carer, wd, occurrences):
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    dates = sorted(set(o[0].date() for o in occurrences))
    span_start, span_end = dates[0], dates[-1]

    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    span_start_wk, span_end_wk = isoweek(span_start), isoweek(span_end)
    active_weeks_in_span = sorted(w for w in active_weeks_all if span_start_wk <= w <= span_end_wk)

    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_span)
    n_hits = len(set(active_weeks_in_span) & hit_weeks)

    if n_active < 3:
        pattern, ratio = 'Insufficient history', (n_hits / n_active if n_active else None)
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_span)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_span if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'

    if pattern == 'Weekly':
        week_label = 'Both'
    elif pattern == 'Fortnightly':
        parities = [d.isocalendar()[1] % 2 for d in dates]
        maj_parity, _ = Counter(parities).most_common(1)[0]
        week_label = 'A' if maj_parity == 0 else 'B'
    else:
        week_label = None  # Occasional / Insufficient history: not split into A/B

    start_times = sorted(set(d.strftime('%H:%M') for d in starts))
    time_note = f"varies {start_times[0]}-{start_times[-1]}" if len(start_times) > 1 else ""

    return {
        'pattern': pattern, 'week': week_label, 'ratio': ratio,
        'start': median_time(starts), 'end': median_time(ends) if ends else '',
        'time_note': time_note, 'count': len(dates),
        'active_weeks_in_span': n_active,
        'first': dates[0], 'last': dates[-1],
    }

records = defaultdict(lambda: defaultdict(list))
pattern_counts = Counter()

for carer, wd_map in roster.items():
    raw = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                raw[(wd, v['client'])].append((v['start_dt'], v['end_dt']))
    for (wd, client), occ in raw.items():
        for cluster in cluster_by_time(occ):
            info = classify_relative(carer, wd, cluster)
            pattern_counts[info['pattern']] += 1
            records[carer][wd].append({'client': client, **info})

print("Pattern counts (carer-relative):", dict(pattern_counts))

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='D9E1F2')
day_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
pattern_fills = {
    'Weekly': PatternFill('solid', fgColor='E2EFDA'),
    'Fortnightly': PatternFill('solid', fgColor='FCE4D6'),
    'Occasional': PatternFill('solid', fgColor='FFFFFF'),
    'Insufficient history': PatternFill('solid', fgColor='F2F2F2'),
}

def sheet_name(name, used):
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Carer-relative visit pattern classification",
    "",
    "Each carer's own working calendar is built first: for every (carer, weekday) pair, the set "
    "of ISO calendar weeks in which that carer did AT LEAST ONE Personal Care visit (to any "
    "client) on that weekday. This is their own 'I was on duty this weekday' timeline.",
    "",
    "For each (carer, weekday, client) relationship (visits to the same client at a similar time "
    "of day, clustered with a 90-minute gap threshold so a morning and an evening visit to the "
    "same client stay separate), the relationship's own span is its first visit to its last visit. "
    "Consistency is then judged ONLY against the carer's own active weeks within that span -- so a "
    "week the carer was on leave, or simply didn't work that weekday, is excluded from the "
    "denominator entirely rather than counting as a 'missed' visit.",
    "",
    "Classification:",
    "- Weekly: visited in >=75% of the carer's own active weeks in the span.",
    "- Fortnightly: visited in 35-75% of active weeks, AND the hits alternate cleanly "
    "(mostly every-other active week) rather than being scattered.",
    "- Occasional: visited in <35% of active weeks, or a 35-75% hit rate that doesn't alternate cleanly.",
    "- Insufficient history: fewer than 3 of the carer's own active weeks fall within the "
    "relationship's span -- too little data to judge a pattern either way.",
    "",
    "This replaces an earlier, simpler version that judged every carer against the same flat rule "
    "(e.g. calling 2 visits 7 days apart 'Weekly' regardless of how long the carer has worked that "
    "weekday). That version produced false positives -- e.g. a client seen only twice in 13 months "
    "was being labelled 'Weekly'. This version checks each carer's own behaviour on its own terms "
    "before comparing across carers.",
    "",
    "Week A / Week B (for Fortnightly rows only) = ISO calendar week number parity (even = A, "
    "odd = B) -- a fixed reference, not necessarily your system's own 'Week 1' label.",
    "",
    f"Slot pattern totals across all carers: {dict(pattern_counts)}",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Days Worked', 'Weekly', 'Fortnightly', 'Occasional', 'Insufficient History']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for carer in sorted(records.keys()):
    days_present = [wd for wd in WEEKDAYS if wd in records[carer]]
    all_entries = [e for wd in records[carer] for e in records[carer][wd]]
    counts = Counter(e['pattern'] for e in all_entries)
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_present)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=counts.get('Weekly', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=counts.get('Fortnightly', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=counts.get('Occasional', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=6, value=counts.get('Insufficient history', 0)).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 55
for col in 'CDEF':
    summary_ws.column_dimensions[col].width = 18
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
headers = ['Weekday', 'Start Time', 'End Time', 'Client', 'Pattern', 'Week',
           'Consistency', 'Visits', 'Active Weeks', 'Date Range', 'Notes']

for carer in sorted(records.keys()):
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:K1')
    r = 3
    for wd in WEEKDAYS:
        entries = records[carer].get(wd, [])
        if not entries:
            continue
        ws.cell(row=r, column=1, value=wd).font = day_font
        for col in range(1, 12):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        entries = sorted(entries, key=lambda e: e['start'])
        for e in entries:
            date_range = (e['first'].strftime('%d/%m/%Y') if e['first'] == e['last']
                          else f"{e['first'].strftime('%d/%m/%Y')} - {e['last'].strftime('%d/%m/%Y')}")
            ratio_display = f"{e['ratio']:.0%}" if e['ratio'] is not None else ''
            row_vals = [wd, e['start'], e['end'], e['client'], e['pattern'],
                        e['week'] or '', ratio_display, e['count'],
                        e['active_weeks_in_span'], date_range, e['time_note']]
            fill = pattern_fills.get(e['pattern'])
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 11 else normal_font
                if fill:
                    cell.fill = fill
            r += 1
        r += 1

    widths = [12, 11, 11, 32, 18, 6, 11, 8, 12, 24, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

out_path = './output/Carer_Weekday_Roster_v2.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
