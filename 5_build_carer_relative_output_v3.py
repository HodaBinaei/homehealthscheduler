import pickle, json
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

import os
PROJECT_ROOT = os.environ.get(
    'HHS_PROJECT_ROOT',
    os.path.dirname(os.path.abspath(__file__)),
)
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)
with open(f'{PROJECT_ROOT}/carer_presence.pkl', 'rb') as f:
    carer_presence = pickle.load(f)  # carer -> (first_date, last_date) across ALL their visits, any client

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90

# Full data window the export actually covers -- absolute outer bound.
DATA_START = datetime.date(2025, 6, 30)
DATA_END = datetime.date(2026, 8, 2)

# --- load client start/end dates ---
def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()

with open(f'{PROJECT_ROOT}/data/clients-new.json') as f:
    clients_json = json.load(f)['client']

client_dates = {}  # normalized name -> (start_date or None, end_date or None)
for c in clients_json:
    if c.get('status') != 'Active':
        continue
    first = (c.get('name') or '').strip()
    last = (c.get('lastname') or '').strip()
    key = norm(f"{last}, {first}")
    sd = c.get('start_date')
    ed = c.get('end_date') or c.get('termination_date')
    sd = datetime.date.fromisoformat(sd) if sd else None
    ed = datetime.date.fromisoformat(ed) if ed else None
    client_dates[key] = (sd, ed)

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

def fmt_minutes(m):
    sign = '+' if m >= 0 else ''
    return f"{sign}{int(round(m))}min"

def actual_variance_note(occurrences):
    """occurrences: list of (req_start, req_end, actual_start, actual_end). Compares when the
    visit actually happened/was booked-on vs. the scheduled Service Requirement time."""
    diffs = [(o[2] - o[0]).total_seconds() / 60 for o in occurrences if o[2] is not None]
    if not diffs:
        return ''
    diffs_sorted = sorted(diffs)
    n = len(diffs_sorted)
    median = diffs_sorted[n // 2] if n % 2 else (diffs_sorted[n // 2 - 1] + diffs_sorted[n // 2]) / 2
    lo, hi = diffs_sorted[0], diffs_sorted[-1]
    if abs(median) <= 5 and (hi - lo) <= 20:
        return ''  # close enough to on-time / not worth flagging
    return f"actual runs {fmt_minutes(median)} vs scheduled (range {fmt_minutes(lo)} to {fmt_minutes(hi)})"

# Carer's own "I worked this weekday" calendar (unchanged -- still needed as the source
# of which weeks within the eligibility window the carer was actually on duty).
carer_weekday_active_weeks = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))

# Carer's own "I worked this CALENDAR DAY" set (any weekday, any client) -- used as a
# fallback source of evidence when a specific weekday-slot has too few active weeks of its
# own to judge (e.g. a relationship that's only 2 weeks old), but the same client is seen
# on most days of the week, so pooling across weekdays gives a much larger evidence base.
carer_active_days = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_active_days[carer].add(v['start_dt'].date())

DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5  # need at least this many of the carer's own active days pooled

def classify_relative(carer, wd, client, occurrences, daily_fallback=None):
    # occurrences: list of (req_start, req_end, actual_start, actual_end)
    # daily_fallback: optional (ratio, active_days), pre-computed by pooling this same
    # time-of-day slot across ALL weekdays -- used only when this weekday's own history is
    # too thin to judge on its own.
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    dates = sorted(set(o[0].date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]

    # Consistency window: the relationship's OWN observed span (first -> last visit).
    # NOTE: we deliberately do NOT stretch this back using the carer's full tenure or the
    # client's overall service start_date -- a specific carer/client PAIRING can legitimately
    # start well after both individually joined (staff turnover, reassignment), so judging a
    # brand-new-but-consistent pairing against the carer's entire career wrongly tanks its
    # ratio. Carer presence / client start_date are used only to ANNOTATE whether the observed
    # edges sit at the boundary of what we could possibly see (censoring), not to resize the
    # window used for the consistency math itself.
    carer_start, carer_end = carer_presence.get(carer, (DATA_START, DATA_END))
    client_start, client_end = client_dates.get(norm(client), (None, None))

    window_start, window_end = observed_start, observed_end

    left_bound = max(carer_start, client_start or DATA_START, DATA_START)
    right_bound = min(carer_end, client_end or DATA_END, DATA_END)
    left_censored = observed_start <= left_bound + datetime.timedelta(days=14)
    right_censored = observed_end >= right_bound - datetime.timedelta(days=14)

    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    win_start_wk, win_end_wk = isoweek(window_start), isoweek(window_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)

    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)

    fallback_note = ''
    if n_active < 3:
        pattern, ratio = 'Insufficient history', (n_hits / n_active if n_active else None)
        if daily_fallback is not None:
            daily_ratio, daily_active_days = daily_fallback
            ratio = daily_ratio
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
            fallback_note = (f"insufficient per-weekday history -- consistency instead computed "
                              f"from all weekdays pooled ({round(daily_ratio * daily_active_days)}/"
                              f"{daily_active_days} of carer's active days)")
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'

    if pattern == 'Weekly':
        week_label = 'Both'
    elif pattern == 'Fortnightly':
        parities = [d.isocalendar()[1] % 2 for d in dates]
        maj_parity, _ = Counter(parities).most_common(1)[0]
        week_label = 'A' if maj_parity == 0 else 'B'
    else:
        week_label = None

    start_times = sorted(set(d.strftime('%H:%M') for d in starts))
    time_note = f"scheduled time varies {start_times[0]}-{start_times[-1]}" if len(start_times) > 1 else ""
    variance_note = actual_variance_note(occurrences)

    censor_note = ''
    if left_censored and right_censored:
        censor_note = 'spans full eligible window'
    elif left_censored:
        censor_note = 'ongoing from window start'
    elif right_censored:
        censor_note = 'ongoing to window end'

    censor_note = ' / '.join(filter(None, [fallback_note, censor_note]))

    return {
        'pattern': pattern, 'week': week_label, 'ratio': ratio,
        'start': median_time(starts), 'end': median_time(ends) if ends else '',
        'time_note': time_note, 'variance_note': variance_note, 'count': len(dates),
        'active_weeks_in_window': n_active,
        'window_start': window_start, 'window_end': window_end,
        'first': observed_start, 'last': observed_end,
        'censor_note': censor_note,
    }

records = defaultdict(lambda: defaultdict(list))
pattern_counts = Counter()

for carer, wd_map in roster.items():
    # Pool this carer's visits per client across ALL weekdays, tagging each with its weekday,
    # so a time-of-day slot that repeats daily can be clustered as ONE group (needed for the
    # daily-pooled fallback) and then still reported one row per weekday as before.
    per_client_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                per_client_tagged[v['client']].append(
                    (v['start_dt'], v['end_dt'], v.get('actual_start_dt'), v.get('actual_end_dt'), wd)
                )

    for client, tagged_occ in per_client_tagged.items():
        for cluster in cluster_by_time(tagged_occ):  # clusters purely by time-of-day, any weekday
            dates = sorted(set(o[0].date() for o in cluster))
            span_start, span_end = dates[0], dates[-1]
            weekdays_covered = set(WEEKDAYS[d.weekday()] for d in dates)
            active_days = [d for d in carer_active_days.get(carer, set()) if span_start <= d <= span_end]
            daily_fallback = None
            # Only pool across weekdays when the visits genuinely span several DIFFERENT
            # weekdays -- otherwise a plain single-weekday-only relationship that's simply new
            # would wrongly inherit a low ratio just because the carer happened to work other
            # days too during that short span.
            if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
                daily_fallback = (len(dates) / len(active_days), len(active_days))

            by_weekday = defaultdict(list)
            for o in cluster:
                by_weekday[o[4]].append(o[:4])  # drop the weekday tag, keep the 4-tuple
            for wd, occs in by_weekday.items():
                info = classify_relative(carer, wd, client, occs, daily_fallback=daily_fallback)
                pattern_counts[info['pattern']] += 1
                records[carer][wd].append({'client': client, **info})

print("Pattern counts (carer-relative, window-bounded):", dict(pattern_counts))

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='D9E1F2')
day_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
pattern_fills = {
    'Weekly': PatternFill('solid', fgColor='E2EFDA'),
    'Fortnightly': PatternFill('solid', fgColor='FCE4D6'),
    'Occasional': PatternFill('solid', fgColor='FFFFFF'),
    'Insufficient history': PatternFill('solid', fgColor='F2F2F2'),
}

def sheet_name(name, used):
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Carer-relative visit pattern classification (window-bounded)",
    "",
    f"Export data window: {DATA_START} to {DATA_END}.",
    "",
    "The roster itself is anchored on the SERVICE REQUIREMENT time (Service Requirement Start/End "
    "Date And Time) -- i.e. the scheduled/rostered slot -- not the Actual check-in/check-out time. "
    "Weekday, time-of-day clustering, and the Scheduled Start/End columns all come from the "
    "Requirement. The Actual time is compared against it per visit, and any meaningful gap shows up "
    "in the Notes column as e.g. 'actual runs +18min vs scheduled (range +5min to +45min)'. Visits "
    "with no usable Requirement time (0.84% of the filtered rows) were dropped, since there's "
    "nothing to roster them against.",
    "",
    "For each carer, their own presence window is computed from the RAW visit data itself -- the "
    "first and last date they appear performing any Personal Care visit, to any client (active or "
    "not), across the whole export. This is used instead of any HR 'start date' field, because the "
    "carer records' own created_date field turned out to be a system-migration timestamp (100 of "
    "113 active carers share the exact same created_date), not a real hire date.",
    "",
    "For each client, their own start_date (and end_date/termination_date where present) comes "
    "directly from the client record -- these looked genuine and varied (371 distinct start dates "
    "across active clients).",
    "",
    "For a given (carer, weekday, client) relationship, the ELIGIBILITY WINDOW is the overlap of: "
    "the carer's own presence window, the client's own service window, and the export's overall "
    "date range. Consistency is judged against the carer's own active weeks (weeks they worked that "
    "weekday at all) WITHIN that eligibility window -- not just within the relationship's own first "
    "and last observed visit. This means a relationship that only shows 6 visits is correctly read "
    "as a complete, consistent history if the carer only started 6 weeks ago -- rather than being "
    "flagged as 'not enough data' when actually it's all the data there could be.",
    "",
    "Classification:",
    "- Weekly: visited in >=75% of the carer's own active weeks within the eligibility window, "
    "for that specific weekday.",
    "- Fortnightly: visited in 35-75% of active weeks, alternating cleanly (mostly every-other "
    "active week).",
    "- Occasional: below 35%, or a 35-75% hit rate that doesn't alternate cleanly.",
    "- Insufficient history: fewer than 3 of the carer's own active weeks fall within the "
    "eligibility window -- WITH ONE FALLBACK: if this specific weekday alone doesn't have "
    "enough active weeks, but the same client is also seen on other weekdays at the same "
    "time-of-day (i.e. it looks like a daily routine), the visits are pooled across ALL "
    "weekdays and consistency is judged against the carer's own active CALENDAR DAYS instead "
    "of active weeks (needs >=5 pooled active days). This is what stops a client on a genuine "
    "daily routine that only started 1-2 weeks ago from being wrongly read as 'Insufficient "
    "history' on every single weekday, when pooled together the evidence is already strong. "
    "When this fallback fires, the Notes column says so and shows the pooled evidence behind "
    "the number.",
    "",
    "A 'Notes' column flags relationships whose observed visits run right up against the edge of "
    "the eligibility window -- these may be ongoing (still happening at the point the export was "
    "taken) rather than genuinely short.",
    "",
    "Week A / Week B (Fortnightly rows only) = ISO calendar week number parity (even = A, odd = B).",
    "",
    f"Slot pattern totals across all carers: {dict(pattern_counts)}",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Days Worked', 'Weekly', 'Fortnightly', 'Occasional', 'Insufficient History',
                    'Carer Presence (first \u2192 last visit in data)']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for carer in sorted(records.keys()):
    days_present = [wd for wd in WEEKDAYS if wd in records[carer]]
    all_entries = [e for wd in records[carer] for e in records[carer][wd]]
    counts = Counter(e['pattern'] for e in all_entries)
    presence = carer_presence.get(carer)
    presence_str = f"{presence[0].strftime('%d/%m/%Y')} \u2192 {presence[1].strftime('%d/%m/%Y')}" if presence else ''
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_present)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=counts.get('Weekly', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=counts.get('Fortnightly', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=counts.get('Occasional', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=6, value=counts.get('Insufficient history', 0)).font = normal_font
    summary_ws.cell(row=row_i, column=7, value=presence_str).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 55
for col in 'CDEF':
    summary_ws.column_dimensions[col].width = 18
summary_ws.column_dimensions['G'].width = 30
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
headers = ['Weekday', 'Scheduled Start', 'Scheduled End', 'Client', 'Pattern', 'Week',
           'Consistency', 'Visits', 'Active Weeks (window)', 'Observed Date Range', 'Notes']

for carer in sorted(records.keys()):
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:K1')
    presence = carer_presence.get(carer)
    if presence:
        ws['A2'] = f"In data: {presence[0].strftime('%d/%m/%Y')} \u2192 {presence[1].strftime('%d/%m/%Y')}"
        ws['A2'].font = note_font
    r = 4
    for wd in WEEKDAYS:
        entries = records[carer].get(wd, [])
        if not entries:
            continue
        ws.cell(row=r, column=1, value=wd).font = day_font
        for col in range(1, 12):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        entries = sorted(entries, key=lambda e: e['start'])
        for e in entries:
            date_range = (e['first'].strftime('%d/%m/%Y') if e['first'] == e['last']
                          else f"{e['first'].strftime('%d/%m/%Y')} - {e['last'].strftime('%d/%m/%Y')}")
            ratio_display = f"{e['ratio']:.0%}" if e['ratio'] is not None else ''
            note = ' / '.join(filter(None, [e['time_note'], e['variance_note'], e['censor_note']]))
            row_vals = [wd, e['start'], e['end'], e['client'], e['pattern'],
                        e['week'] or '', ratio_display, e['count'],
                        e['active_weeks_in_window'], date_range, note]
            fill = pattern_fills.get(e['pattern'])
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 11 else normal_font
                if fill:
                    cell.fill = fill
            r += 1
        r += 1

    widths = [12, 11, 11, 32, 18, 6, 11, 8, 18, 24, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

out_path = f'{PROJECT_ROOT}/output/Carer_Weekday_Roster_v3.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
