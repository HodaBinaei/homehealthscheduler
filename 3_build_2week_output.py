import pickle
from collections import defaultdict, Counter
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
os.makedirs('./output', exist_ok=True)

with open('./roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

# Group at (carer, weekday, client) level -- ignores small minute-level time drift,
# which otherwise fragments one recurring visit into many "occasional" slots.
raw_groups = defaultdict(list)  # (carer, weekday, client) -> list of (start_dt, end_dt)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if not v['start_dt']:
                continue
            key = (carer, wd, v['client'])
            raw_groups[key].append((v['start_dt'], v['end_dt']))

# Split each (carer, weekday, client) group into time-of-day clusters -- a client visited
# both morning and evening on the same weekday is two distinct slots, not one with "drift".
# Gap threshold: a new cluster starts when consecutive sorted start-times are >90 min apart.
TIME_GAP_MINUTES = 90

def cluster_by_time(occurrences):
    occ_sorted = sorted(occurrences, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters = []
    current = [occ_sorted[0]]
    for prev, cur in zip(occ_sorted, occ_sorted[1:]):
        prev_min = prev[0].hour * 60 + prev[0].minute
        cur_min = cur[0].hour * 60 + cur[0].minute
        if cur_min - prev_min > TIME_GAP_MINUTES:
            clusters.append(current)
            current = [cur]
        else:
            current.append(cur)
    clusters.append(current)
    return clusters

groups = {}  # (carer, weekday, client, cluster_idx) -> list of (start_dt, end_dt)
for key, occ in raw_groups.items():
    for idx, cluster in enumerate(cluster_by_time(occ)):
        groups[key + (idx,)] = cluster

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

def classify(occurrences):
    """occurrences: list of (start_dt, end_dt). Returns dict with pattern/week/stats."""
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    dates = sorted(set(d.date() for d in starts))
    n_dates = len(dates)

    if n_dates == 1:
        pattern = 'Occasional'
    else:
        diffs = [(dates[i+1] - dates[i]).days for i in range(n_dates - 1)]
        weekly_like = sum(1 for d in diffs if 6 <= d <= 8)
        biweekly_like = sum(1 for d in diffs if 13 <= d <= 15)
        n = len(diffs)
        if weekly_like / n > 0.5:
            pattern = 'Weekly'
        elif biweekly_like / n > 0.5:
            pattern = 'Fortnightly'
        else:
            pattern = 'Occasional'

    if pattern == 'Weekly':
        week_label = 'Both'
    else:
        parities = [d.isocalendar()[1] % 2 for d in dates]
        maj_parity, maj_count = Counter(parities).most_common(1)[0]
        week_label = 'A' if maj_parity == 0 else 'B'

    start_times = sorted(set(d.strftime('%H:%M') for d in starts))
    time_display = median_time(starts)
    if len(start_times) > 1:
        time_note = f"varies {start_times[0]}-{start_times[-1]}"
    else:
        time_note = ""
    end_display = median_time(ends) if ends else ""

    return {
        'pattern': pattern, 'week': week_label,
        'start': time_display, 'end': end_display, 'time_note': time_note,
        'count': n_dates, 'first': dates[0], 'last': dates[-1],
    }

records = defaultdict(lambda: defaultdict(list))
pattern_counts = Counter()
for (carer, wd, client, cluster_idx), occ in groups.items():
    info = classify(occ)
    pattern_counts[info['pattern']] += 1
    records[carer][wd].append({'client': client, **info})

print("Pattern counts (carer, weekday, client level):", dict(pattern_counts))

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='D9E1F2')
day_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
weekA_fill = PatternFill('solid', fgColor='FCE4D6')
weekB_fill = PatternFill('solid', fgColor='E2EFDA')
weekA_font = Font(name=FONT, bold=True, size=13, color='C55A11')
weekB_font = Font(name=FONT, bold=True, size=13, color='548235')

def sheet_name(name, used):
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Read Me sheet ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Two-week (fortnightly) carer rotation roster",
    "",
    "Patterns are classified per (carer, weekday, client) -- i.e. treating all visits for a given "
    "client on a given weekday as one recurring slot, using its actual visit dates across the full "
    "export date range (minor start/end-time drift between weeks is folded together; the displayed "
    "time is the median, with a 'varies' note if it moved by more than a few minutes):",
    "- Weekly: recurs at ~7-day intervals -> shown in BOTH Week A and Week B.",
    "- Fortnightly: recurs at ~14-day intervals -> shown ONLY in Week A or ONLY in Week B, "
    "whichever the actual dates fall on.",
    "- Occasional: single or irregular occurrences (ad hoc / one-off cover, hospital appointments, "
    "holidays, etc.) -> shown under whichever week most of its occurrences fell on, flagged Occasional.",
    "",
    "Week A / Week B are defined by ISO calendar week number parity (even week number = Week A, "
    "odd = Week B). This is a fixed, absolute reference across the whole dataset, but may not match "
    "whatever internal 'Week 1 / Week 2' label the roster system itself uses -- tell me which actual "
    "calendar week your system calls 'Week 1' and I can relabel A/B to match it exactly.",
    "",
    f"Slot pattern totals across all carers: {dict(pattern_counts)}",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary sheet ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Days Worked', 'Weekly Clients', 'Fortnightly Clients', 'Occasional Clients']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for carer in sorted(records.keys()):
    days_present = [wd for wd in WEEKDAYS if wd in records[carer]]
    all_entries = [e for wd in records[carer] for e in records[carer][wd]]
    weekly_n = sum(1 for e in all_entries if e['pattern'] == 'Weekly')
    fortnightly_n = sum(1 for e in all_entries if e['pattern'] == 'Fortnightly')
    occasional_n = sum(1 for e in all_entries if e['pattern'] == 'Occasional')
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_present)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=weekly_n).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=fortnightly_n).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=occasional_n).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 55
for col in 'CDE':
    summary_ws.column_dimensions[col].width = 18
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
headers = ['Weekday', 'Start Time', 'End Time', 'Client', 'Pattern', 'Visits', 'Date Range', 'Notes']

def write_week_block(ws, r, carer, week_label, fill, font):
    ws.cell(row=r, column=1, value=f"WEEK {week_label}").font = font
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = fill
    r += 2
    for wd in WEEKDAYS:
        entries = [e for e in records[carer].get(wd, [])
                   if e['week'] == week_label or e['week'] == 'Both']
        if not entries:
            continue
        ws.cell(row=r, column=1, value=wd).font = day_font
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        entries.sort(key=lambda e: e['start'])
        for e in entries:
            date_range = (e['first'].strftime('%d/%m/%Y') if e['first'] == e['last']
                          else f"{e['first'].strftime('%d/%m/%Y')} - {e['last'].strftime('%d/%m/%Y')}")
            row_vals = [wd, e['start'], e['end'], e['client'], e['pattern'], e['count'], date_range, e['time_note']]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 8 else normal_font
            r += 1
        r += 1
    return r + 1

for carer in sorted(records.keys()):
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:H1')
    r = 3
    r = write_week_block(ws, r, carer, 'A', weekA_fill, weekA_font)
    r = write_week_block(ws, r, carer, 'B', weekB_fill, weekB_font)

    widths = [12, 11, 11, 32, 12, 8, 24, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

out_path = './output/Carer_2Week_Rotation_Roster.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
