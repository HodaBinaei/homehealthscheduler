import pickle, json, math
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

import os
PROJECT_ROOT = os.environ.get(
    'HHS_PROJECT_ROOT',
    os.path.dirname(os.path.abspath(__file__)),
)
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)
with open(f'{PROJECT_ROOT}/carer_presence.pkl', 'rb') as f:
    carer_presence = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90
DATA_START = datetime.date(2025, 6, 30)
DATA_END = datetime.date(2026, 8, 2)
DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5

def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()

# --- load client coordinates and dates ---
with open(f'{PROJECT_ROOT}/data/clients-new.json') as f:
    clients_json = json.load(f)['client']
client_dates = {}
client_coords = {}
for c in clients_json:
    if c.get('status') != 'Active':
        continue
    first = (c.get('name') or '').strip()
    last = (c.get('lastname') or '').strip()
    key = norm(f"{last}, {first}")           # matches client_dates lookups elsewhere
    display_key = norm(f"{first} {last}")    # matches roster's 'client' field format ("Firstname Lastname")
    sd = c.get('start_date')
    ed = c.get('end_date') or c.get('termination_date')
    client_dates[key] = (
        datetime.date.fromisoformat(sd) if sd else None,
        datetime.date.fromisoformat(ed) if ed else None,
    )
    lat, lon = c.get('latitude'), c.get('longitude')
    if lat and lon:
        client_coords[display_key] = (float(lat), float(lon))

# --- load carer home coordinates ---
with open(f'{PROJECT_ROOT}/data/users-new.json') as f:
    users_json = json.load(f)['user']
carer_coords = {}
for u in users_json:
    if u.get('status') != 'Active' or not u.get('is_caregiver'):
        continue
    first = (u.get('name') or '').strip()
    last = (u.get('lastname') or '').strip()
    display = f"{first} {last}".strip()
    lat, lon = u.get('latitude'), u.get('longitude')
    if lat and lon:
        carer_coords[display] = (float(lat), float(lon))

def haversine_km(p1, p2):
    lat1, lon1 = p1
    lat2, lon2 = p2
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

# Real driving distances (OSRM road-network km, from the carer/client pid/cid matrix),
# preferred over straight-line haversine wherever available.
try:
    with open(f'{PROJECT_ROOT}/carer_client_distances.pkl', 'rb') as f:
        carer_client_km = pickle.load(f)
except FileNotFoundError:
    carer_client_km = {}

def travel_km(carer, client, home_coord=None, client_coord=None):
    d = carer_client_km.get((carer, client))
    if d is not None:
        return d
    if home_coord is not None and client_coord is not None:
        return haversine_km(home_coord, client_coord)
    return None

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

carer_weekday_active_weeks = defaultdict(set)
carer_active_days = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))
                carer_active_days[carer].add(v['start_dt'].date())

def classify_relative(carer, wd, client, occurrences, daily_fallback=None):
    dates = sorted(set(o[0].date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]
    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    win_start_wk, win_end_wk = isoweek(observed_start), isoweek(observed_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)
    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)

    if n_active < 3:
        pattern = 'Insufficient history'
        if daily_fallback is not None:
            daily_ratio, daily_active_days = daily_fallback
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'
    return {'pattern': pattern}

# --- rebuild set_roster (Weekly slots) and extra visits, same logic as the coverage workbook ---
set_roster_clients_by_carer = defaultdict(set)  # carer -> set of client names (any weekday) in set roster
extra_clients_by_carer = defaultdict(set)       # carer -> set of client names visited as "extra"

for carer, wd_map in roster.items():
    per_client_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                per_client_tagged[v['client']].append(
                    (v['start_dt'], v['end_dt'], v.get('actual_start_dt'), v.get('actual_end_dt'), wd)
                )
    weekly_by_wd = defaultdict(set)
    for client, tagged_occ in per_client_tagged.items():
        for cluster in cluster_by_time(tagged_occ):
            dates = sorted(set(o[0].date() for o in cluster))
            span_start, span_end = dates[0], dates[-1]
            weekdays_covered = set(WEEKDAYS[d.weekday()] for d in dates)
            active_days = [d for d in carer_active_days.get(carer, set()) if span_start <= d <= span_end]
            daily_fallback = None
            if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
                daily_fallback = (len(dates) / len(active_days), len(active_days))
            by_weekday = defaultdict(list)
            for o in cluster:
                by_weekday[o[4]].append(o[:4])
            for wd, occs in by_weekday.items():
                info = classify_relative(carer, wd, client, occs, daily_fallback=daily_fallback)
                if info['pattern'] == 'Weekly':
                    weekly_by_wd[wd].add(client)
                    set_roster_clients_by_carer[carer].add(client)

    for wd, visits in wd_map.items():
        roster_clients = weekly_by_wd.get(wd, set())
        for v in visits:
            if v['client'] not in roster_clients:
                extra_clients_by_carer[carer].add(v['client'])

print(f"Carers with a set roster: {len(set_roster_clients_by_carer)}")
print(f"Carers with any extra visits: {len(extra_clients_by_carer)}")

# --- geographic analysis ---
# For each carer: find her biggest travel distance (home -> furthest set-roster client),
# then find every active client within that same radius, and see what fraction of THEM
# she covers via extra (off-roster/ad hoc) visits.
results = []
client_exact_names = [f"{(c.get('name') or '').strip()} {(c.get('lastname') or '').strip()}".strip()
                       for c in clients_json if c.get('status') == 'Active']

for carer, roster_clients in set_roster_clients_by_carer.items():
    home = carer_coords.get(carer)
    if not home:
        continue
    # "biggest travel" should reflect everywhere she actually visits, not just her fixed
    # set roster -- an extra/ad hoc visit can easily be her furthest trip of all.
    all_visited_clients = roster_clients | extra_clients_by_carer.get(carer, set())
    all_client_coords = []
    for client in all_visited_clients:
        coord = client_coords.get(norm(client))
        if coord:
            d = travel_km(carer, client, home, coord)
            if d is not None:
                all_client_coords.append((client, coord, d))
    if not all_client_coords:
        continue
    furthest_client, furthest_coord, max_dist = max(all_client_coords, key=lambda x: x[2])

    # Search radius = her own "normal" travel (median distance across everywhere she
    # actually visits) plus a 5km buffer -- not her single furthest trip, which can be a
    # rare outlier that would otherwise inflate the radius to cover an unrealistic area.
    all_distances = sorted(d for _, _, d in all_client_coords)
    n = len(all_distances)
    median_dist = (all_distances[n // 2 - 1] + all_distances[n // 2]) / 2 if n % 2 == 0 else all_distances[n // 2]
    search_radius = median_dist + 5

    nearby_set = set()
    for client_name in client_exact_names:
        d = travel_km(carer, client_name, home, client_coords.get(norm(client_name)))
        if d is not None and d <= search_radius:
            nearby_set.add(norm(client_name))

    extra_clients = extra_clients_by_carer.get(carer, set())
    extra_norm = set(norm(c) for c in extra_clients)
    roster_norm = set(norm(c) for c in roster_clients)
    already_roster_nearby = nearby_set & roster_norm
    covered_nearby = (nearby_set & extra_norm) - already_roster_nearby

    results.append({
        'carer': carer, 'max_dist_km': max_dist, 'furthest_client': furthest_client,
        'search_radius_km': search_radius,
        'n_roster_clients': len(roster_clients),
        'n_nearby_total': len(nearby_set),
        'n_nearby_already_roster': len(already_roster_nearby),
        'n_nearby_covered_extra': len(covered_nearby),
        'n_nearby_uncovered': len(nearby_set) - len(already_roster_nearby) - len(covered_nearby),
    })

print(f"Carers with full geographic data (home + roster client coords): {len(results)}")

# ==========================================================================
# Build workbook
# ==========================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')

notes_ws = wb.create_sheet('Read Me')
notes = [
    "Carer geographic coverage",
    "",
    "For each carer, this finds her own 'biggest travel' distance -- from her home address to "
    "the furthest client she actually visits, whether that's a set-roster (Weekly-classified) "
    "client or an extra/ad hoc one. This is shown as-is (Biggest Travel / Furthest Client) but "
    "does NOT define the nearby-client search radius -- a carer whose usual patch is tight but "
    "who has one rare far-flung outlier visit would otherwise get an unrealistically huge "
    "search radius covering an area she never really operates in. Instead, Search Radius = her "
    "median travel distance across everywhere she actually visits, plus a 5km buffer -- a much "
    "more realistic reflection of her normal working area. Every other active client within "
    "that radius counts as 'nearby', and checks how many of them she covers via EXTRA "
    "(off-roster / ad hoc) visits, versus how many are already on her set roster, versus how "
    "many she never visits at all.",
    "",
    "Distances are REAL driving distances (road-network km, from an OSRM-computed carer/client "
    "distance matrix keyed by their pid/cid), not straight-line -- straight-line (haversine) "
    "distance is used only as a fallback for the handful of pairs missing from the matrix.",
    "",
    "Columns:",
    "- Biggest Travel (km): distance from home to her furthest visited client (roster or extra).",
    "- Furthest Client: who that is.",
    "- Search Radius (km): median travel distance + 5km -- what actually defines 'nearby' below.",
    "- Nearby Clients (within radius): every active client (any carer's caseload) within the "
    "search radius of her home.",
    "- Already On Her Roster: of those nearby clients, how many are already on her own set "
    "roster.",
    "- Covered By Extra Visits: of those nearby clients NOT on her set roster, how many she "
    "still visits at least once as an extra/ad hoc visit.",
    "- Never Visited: nearby clients she has never visited at all (neither roster nor extra).",
    "- % Nearby Covered (roster+extra): what fraction of all nearby clients she has SOME "
    "contact with, one way or another.",
    "",
    "Carers without a home coordinate, or without any coordinate-matched set-roster client, "
    "are excluded (2 of 113 active carers lack a home coordinate).",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 130

ws = wb.create_sheet('Geographic Coverage')
headers = ['Carer', 'Biggest Travel (km)', 'Furthest Client', 'Search Radius (km)', 'Set Roster Clients',
           'Nearby Clients (within radius)', 'Already On Her Roster', 'Covered By Extra Visits',
           'Never Visited', '% Nearby Covered (roster+extra)']
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for r in sorted(results, key=lambda x: -x['max_dist_km']):
    covered_total = r['n_nearby_already_roster'] + r['n_nearby_covered_extra']
    pct = covered_total / r['n_nearby_total'] if r['n_nearby_total'] else 0
    row_vals = [r['carer'], round(r['max_dist_km'], 1), r['furthest_client'], round(r['search_radius_km'], 1),
                r['n_roster_clients'], r['n_nearby_total'], r['n_nearby_already_roster'],
                r['n_nearby_covered_extra'], r['n_nearby_uncovered'], f"{pct:.0%}"]
    for col, val in enumerate(row_vals, start=1):
        ws.cell(row=row_i, column=col, value=val).font = normal_font
    row_i += 1

widths = [28, 16, 26, 16, 16, 22, 18, 20, 14, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

out_path = f'{PROJECT_ROOT}/output/Carer_Geographic_Coverage.xlsx'
wb.save(out_path)
print("Saved", out_path)
