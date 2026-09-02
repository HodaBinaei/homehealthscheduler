import pickle, json
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

PROJECT_ROOT = '/home/hoda/Desktop/homehealthscheduler_local_pipeline'
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)
with open(f'{PROJECT_ROOT}/carer_presence.pkl', 'rb') as f:
    carer_presence = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90
DATA_START = datetime.date(2025, 6, 30)
DATA_END = datetime.date(2026, 8, 2)

def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()

with open(f'{PROJECT_ROOT}/data/clients-new.json') as f:
    clients_json = json.load(f)['client']

client_dates = {}
for c in clients_json:
    if c.get('status') != 'Active':
        continue
    first = (c.get('name') or '').strip()
    last = (c.get('lastname') or '').strip()
    key = norm(f"{last}, {first}")
    sd = c.get('start_date')
    ed = c.get('end_date') or c.get('termination_date')
    client_dates[key] = (
        datetime.date.fromisoformat(sd) if sd else None,
        datetime.date.fromisoformat(ed) if ed else None,
    )

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

def fmt_minutes(m):
    sign = '+' if m >= 0 else ''
    return f"{sign}{int(round(m))}min"

def actual_variance_note(occurrences):
    """occurrences: list of (req_start, req_end, actual_start, actual_end)."""
    diffs = [(o[2] - o[0]).total_seconds() / 60 for o in occurrences if o[2] is not None]
    if not diffs:
        return ''
    diffs_sorted = sorted(diffs)
    n = len(diffs_sorted)
    median = diffs_sorted[n // 2] if n % 2 else (diffs_sorted[n // 2 - 1] + diffs_sorted[n // 2]) / 2
    lo, hi = diffs_sorted[0], diffs_sorted[-1]
    if abs(median) <= 5 and (hi - lo) <= 20:
        return ''
    return f"actual runs {fmt_minutes(median)} vs scheduled (range {fmt_minutes(lo)} to {fmt_minutes(hi)})"

carer_weekday_active_weeks = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))

carer_active_days = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_active_days[carer].add(v['start_dt'].date())

DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5

def classify_relative(carer, wd, client, occurrences, daily_fallback=None):
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    dates = sorted(set(o[0].date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]

    carer_start, carer_end = carer_presence.get(carer, (DATA_START, DATA_END))
    client_start, client_end = client_dates.get(norm(client), (None, None))
    window_start, window_end = observed_start, observed_end
    left_bound = max(carer_start, client_start or DATA_START, DATA_START)
    right_bound = min(carer_end, client_end or DATA_END, DATA_END)
    left_censored = observed_start <= left_bound + datetime.timedelta(days=14)
    right_censored = observed_end >= right_bound - datetime.timedelta(days=14)

    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    win_start_wk, win_end_wk = isoweek(window_start), isoweek(window_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)

    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)

    fallback_note = ''
    if n_active < 3:
        pattern, ratio = 'Insufficient history', (n_hits / n_active if n_active else None)
        if daily_fallback is not None:
            daily_ratio, daily_active_days = daily_fallback
            ratio = daily_ratio
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
            fallback_note = (f"insufficient per-weekday history -- consistency instead computed "
                              f"from all weekdays pooled ({round(daily_ratio * daily_active_days)}/"
                              f"{daily_active_days} of carer's active days)")
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'

    start_times = sorted(set(d.strftime('%H:%M') for d in starts))
    time_note = f"scheduled time varies {start_times[0]}-{start_times[-1]}" if len(start_times) > 1 else ""
    variance_note = actual_variance_note(occurrences)
    censor_note = ''
    if left_censored and right_censored:
        censor_note = 'spans full eligible window'
    elif left_censored:
        censor_note = 'ongoing from window start'
    elif right_censored:
        censor_note = 'ongoing to window end'
    censor_note = ' / '.join(filter(None, [fallback_note, censor_note]))

    return {
        'pattern': pattern, 'ratio': ratio,
        'start': median_time(starts), 'end': median_time(ends) if ends else '',
        'time_note': time_note, 'variance_note': variance_note, 'count': len(dates), 'active_weeks': n_active,
        'first': observed_start, 'last': observed_end, 'censor_note': censor_note,
    }

# --- classify everything, keep only Weekly ("repeating") relationships, restricted to
# carers with at least 2 months (60 days) of presence in the data ---

MIN_TENURE_DAYS = 60
qualifying_carers = {
    c for c, (first, last) in carer_presence.items() if (last - first).days >= MIN_TENURE_DAYS
}
print(f"{len(qualifying_carers)} of {len(carer_presence)} carers have >= {MIN_TENURE_DAYS} days of history")

permanent = defaultdict(list)  # carer -> list of repeating (weekly) slot records
all_slot_counts = Counter()    # carer -> total classified slots (any pattern) -- for stability %
weekly_slot_counts = Counter()  # carer -> count of Weekly-pattern slots

for carer, wd_map in roster.items():
    if carer not in qualifying_carers:
        continue
    per_client_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                per_client_tagged[v['client']].append(
                    (v['start_dt'], v['end_dt'], v.get('actual_start_dt'), v.get('actual_end_dt'), wd)
                )

    for client, tagged_occ in per_client_tagged.items():
        for cluster in cluster_by_time(tagged_occ):
            dates = sorted(set(o[0].date() for o in cluster))
            span_start, span_end = dates[0], dates[-1]
            weekdays_covered = set(WEEKDAYS[d.weekday()] for d in dates)
            active_days = [d for d in carer_active_days.get(carer, set()) if span_start <= d <= span_end]
            daily_fallback = None
            if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
                daily_fallback = (len(dates) / len(active_days), len(active_days))

            by_weekday = defaultdict(list)
            for o in cluster:
                by_weekday[o[4]].append(o[:4])
            for wd, occs in by_weekday.items():
                info = classify_relative(carer, wd, client, occs, daily_fallback=daily_fallback)
                all_slot_counts[carer] += 1
                if info['pattern'] == 'Weekly':
                    weekly_slot_counts[carer] += 1
                    permanent[carer].append({'weekday': wd, 'client': client, **info})

print(f"Carers (>= {MIN_TENURE_DAYS} days tenure) with at least one repeating client: {len(permanent)}")
total_relationships = sum(len(v) for v in permanent.values())
print(f"Total repeating carer-client-weekday relationships: {total_relationships}")

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
title_font = Font(name=FONT, bold=True, size=14)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
gold_fill = PatternFill('solid', fgColor='FFF2CC')  # 100% consistency highlight

def sheet_name(name, used):
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Repeating weekday roster per carer (established carers only)",
    "",
    f"Restricted to carers with at least {MIN_TENURE_DAYS} days (~2 months) of their own presence "
    "in the data -- their own first-to-last visit date across all clients, computed from the raw "
    "visit history itself (carer records don't have a reliable hire date; see the main roster "
    "workbook's Read Me for why). This excludes recent hires who simply haven't been here long "
    "enough yet to show a settled pattern one way or the other.",
    "",
    "'Repeating' here = a (carer, weekday, client) relationship classified as WEEKLY: the carer "
    "visited that client in at least 75% of their own active weeks (weeks they actually worked "
    "that weekday) across the relationship's observed span. This includes one fallback: if this "
    "specific weekday alone has too little history (fewer than 3 active weeks), but the same "
    "client is seen on 4+ other weekdays too at the same time-of-day (i.e. it looks like a daily "
    "routine), consistency is instead computed by pooling across all weekdays and comparing "
    "against the carer's own active calendar days.",
    "",
    "Scheduled Start/End come from the Service Requirement time (the rostered slot), not the "
    "Actual check-in/out time. Where Actual ran meaningfully different from what was scheduled, "
    "that shows up in the Notes column, e.g. 'actual runs +18min vs scheduled (range ...)'.",
    "",
    "A carer can have more than one repeating client (different clients on different weekdays, or "
    "even the same weekday at different times of day). Each row is one such relationship.",
    "",
    "The Summary sheet's Roster Stability % is (repeating/Weekly slots) \u00f7 (all slots this carer "
    "was ever classified for, any pattern) -- a rough measure of how much of a carer's week is a "
    "fixed, repeating roster versus flexible/ad hoc cover. Sorted highest-first, so the carers with "
    "the most settled weekly roster are at the top.",
    "",
    "The Consistency column is the exact ratio behind the Weekly label -- rows highlighted in "
    "gold are 100% consistency, useful if you want a stricter cut than the 75% threshold.",
    "",
    "The Notes column flags relationships whose visits run right up to the edge of what the data "
    "can show ('ongoing from/to window start/end') -- these are very likely still continuing "
    "beyond the export date.",
    "",
    f"Total: {len(permanent)} of {len(qualifying_carers)} qualifying carers have at least one "
    f"repeating client; {total_relationships} repeating carer-weekday-client relationships in total.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 130

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Roster Stability %', 'Repeating Clients (count)', 'Client Names',
                    'Carer Presence (first \u2192 last visit in data)']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
carers_by_stability = sorted(
    permanent.keys(),
    key=lambda c: weekly_slot_counts[c] / all_slot_counts[c] if all_slot_counts[c] else 0,
    reverse=True,
)
for carer in carers_by_stability:
    entries = permanent[carer]
    client_names = sorted(set(e['client'] for e in entries))
    stability = weekly_slot_counts[carer] / all_slot_counts[carer] if all_slot_counts[carer] else 0
    presence = carer_presence.get(carer)
    presence_str = f"{presence[0].strftime('%d/%m/%Y')} \u2192 {presence[1].strftime('%d/%m/%Y')}" if presence else ''
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=f"{stability:.0%}").font = normal_font
    summary_ws.cell(row=row_i, column=3, value=len(client_names)).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=', '.join(client_names)).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=presence_str).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 18
summary_ws.column_dimensions['C'].width = 22
summary_ws.column_dimensions['D'].width = 90
summary_ws.column_dimensions['E'].width = 30
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
headers = ['Weekday', 'Scheduled Start', 'Scheduled End', 'Client', 'Consistency', 'Visits',
           'Active Weeks', 'Date Range', 'Notes']

for carer in sorted(permanent.keys()):
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    entries = sorted(permanent[carer], key=lambda e: (WEEKDAYS.index(e['weekday']), e['start']))
    r = 4
    for e in entries:
        date_range = (e['first'].strftime('%d/%m/%Y') if e['first'] == e['last']
                      else f"{e['first'].strftime('%d/%m/%Y')} - {e['last'].strftime('%d/%m/%Y')}")
        note = ' / '.join(filter(None, [e['time_note'], e['variance_note'], e['censor_note']]))
        row_vals = [e['weekday'], e['start'], e['end'], e['client'], f"{e['ratio']:.0%}",
                    e['count'], e['active_weeks'], date_range, note]
        is_perfect = e['ratio'] >= 0.999
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = note_font if col == 9 else normal_font
            if is_perfect:
                cell.fill = gold_fill
        r += 1

    widths = [12, 13, 13, 32, 12, 8, 13, 24, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

out_path = f'{PROJECT_ROOT}/output/Carer_Repeating_Roster_2mo.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
