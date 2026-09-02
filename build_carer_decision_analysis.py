"""
Builds the per-carer DECISION analysis -- one sheet per carer working today, showing the
actual reasoning behind every number that ends up in caregivers.json and
crid_prid_feasible.json: her set roster, her extra visits and why, her extend_feasibility
classifier verdict and the evidence behind it, her data-driven distance/time bounds and the
observations behind those, and -- for every patient she's a feasible candidate for today --
the full weight breakdown (consistency / freq_factor / recency_decay / status /
concentration_factor) that produced that specific number.

Reuses the EXACT SAME foundation as build_full_recompute_real_day.py (same historical
analysis, same classifier, same weight formula) -- this is not a separate analysis, it's
the visible version of the same one that produces the JSON files.

Run this AFTER build_full_recompute_real_day.py.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json, math

import os
PROJECT_ROOT = os.environ.get(
    'HHS_PROJECT_ROOT',
    os.path.dirname(os.path.abspath(__file__)),
)

# Reuse the full historical foundation + classifier, unchanged, up to (not including) the
# day-specific JSON-building steps.
_src = open(f'{PROJECT_ROOT}/build_full_recompute_real_day.py').read()
_cutoff = _src.split("print(f\"Loading real day export:")[0]
exec(_cutoff)

print(f"Loading real day export: {len(_real_patients_raw['patients'])} patients, "
      f"{len(_real_caregivers_raw['caregivers'])} caregivers")

# ---------------------------------------------------------------------------
# Extract today's data (same as the main script)
# ---------------------------------------------------------------------------
day_requirements = []
for p in _real_patients_raw['patients']:
    client = f"{p.get('name', '')} {p.get('lastname', '')}".strip()
    rw = p['request_window']
    day_requirements.append({
        'pid': p['pid'], 'prid': p['prid'], 'client': client,
        'start_min': rw['start_time_soft'], 'end_min': rw['end_time_soft'],
        'match_request_list': rw.get('match_request_list', []),
    })

carer_shift_input = {}
for c in _real_caregivers_raw['caregivers']:
    carer = f"{c.get('name', '')} {c.get('lastname', '')}".strip()
    carer_shift_input[carer] = {
        'cid': c['cid'], 'crid': c['crid'],
        'start_min': c['shift']['start_time'], 'end_min': c['shift']['end_time'],
    }

carers_today = sorted(carer_shift_input.keys())

daily_carers_all = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            daily_carers_all[v['start_dt'].date()].add(carer)
same_weekday_dates = [d for d in daily_carers_all if WEEKDAYS[d.weekday()] == TARGET_WEEKDAY]
recent_same_weekday = [d for d in same_weekday_dates if d >= TARGET_DATE - datetime.timedelta(weeks=8)]
avg_carers_recent = (sum(len(daily_carers_all[d]) for d in recent_same_weekday) / len(recent_same_weekday)
                      if recent_same_weekday else len(carer_shift_input))
today_carers_n = len(carer_shift_input)
staffing_ratio = today_carers_n / avg_carers_recent if avg_carers_recent else 1.0
day_is_short_staffed = staffing_ratio < 0.85

# ---------------------------------------------------------------------------
# Rebuild the slot-weight raw scores (same formula as the main script) so we can show the
# breakdown per (carer, patient) pair for today.
# ---------------------------------------------------------------------------
customer_totals = Counter()
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            customer_totals[v['client']] += 1

WINDOW_DAYS = 112
STATUS_FACTORS = {'Current Primary': 1.0, 'Support / Relief': 0.5, 'Former / Relief': 0.2}

def identify_carer_status(overall_pct, days_since_last_visit):
    if days_since_last_visit > 50:
        return "Former / Relief"
    if overall_pct >= 40:
        return "Current Primary"
    return "Support / Relief"

slot_breakdown = {}  # key -> {carer: {breakdown dict}}
slot_clusters_by_client_wd = defaultdict(list)
for (client, wd), clusters in client_slot_history.items():
    total_cust_visits = customer_totals.get(client, 0)
    if total_cust_visits <= 0:
        continue
    calls_per_day = total_cust_visits / float(WINDOW_DAYS)
    freq_factor = 1 + math.log1p(calls_per_day)
    for idx, cluster in enumerate(clusters):
        slot_total = len(cluster)
        by_carer = defaultdict(list)
        for carer, dt, a_s in cluster:
            by_carer[carer].append(a_s.date() if a_s is not None else dt.date())
        minutes = [dt.hour * 60 + dt.minute for _, dt, a_s in cluster]
        median_minute = sorted(minutes)[len(minutes) // 2]
        key = (client, wd, idx)
        slot_clusters_by_client_wd[(client, wd)].append((median_minute, key))
        entry = {}
        for carer, dates in by_carer.items():
            last_visit = max(dates)
            days_since_last_visit = max((TARGET_DATE - last_visit).days, 0)
            overall_pct = round((len(dates) / slot_total) * 100, 1)
            status = identify_carer_status(overall_pct, days_since_last_visit)
            consistency = overall_pct / 100.0
            recency_decay = math.exp(-days_since_last_visit / 21.0)
            status_factor = STATUS_FACTORS.get(status, 0.3)
            cf = concentration_factor(carer)
            raw = consistency * freq_factor * recency_decay * status_factor * cf
            entry[carer] = {
                'consistency_pct': overall_pct, 'status': status, 'days_since': days_since_last_visit,
                'freq_factor': freq_factor, 'recency_decay': recency_decay,
                'status_factor': status_factor, 'concentration_factor': cf, 'raw': raw,
            }
        slot_breakdown[key] = entry

def find_slot_breakdown(client, wd, start_minute):
    best_key, best_diff = None, None
    for median_minute, key in slot_clusters_by_client_wd.get((client, wd), []):
        diff = abs(median_minute - start_minute)
        if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
            best_key, best_diff = key, diff
    return slot_breakdown.get(best_key, {})

# ---------------------------------------------------------------------------
# Read the JSON files the main script already produced, for the final weights/values
# ---------------------------------------------------------------------------
with open(f'{PROJECT_ROOT}/output/patient.json') as f:
    out_patients = json.load(f)
with open(f'{PROJECT_ROOT}/output/caregivers.json') as f:
    out_caregivers = json.load(f)
with open(f'{PROJECT_ROOT}/output/crid_prid_feasible.json') as f:
    fp = json.load(f)

pid_to_client = {r['pid']: r['client'] for r in day_requirements}
crid_to_carer = {}
for carer, shift in carer_shift_input.items():
    crid_to_carer[shift['crid']] = carer
out_caregiver_by_crid = {c['crid']: c for c in out_caregivers}
feas_by_crid = defaultdict(list)
for row in fp:
    feas_by_crid[row['crid']].append(row)
prid_to_pid = {r['prid']: r['pid'] for r in day_requirements}

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
title_font = Font(name=FONT, bold=True, size=14)
section_font = Font(name=FONT, bold=True, size=12)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
good_fill = PatternFill('solid', fgColor='E2EFDA')
warn_fill = PatternFill('solid', fgColor='FCE4D6')

notes_ws = wb.create_sheet('Read Me')
notes = [
    f"Carer decision analysis -- {TARGET_DATE} ({TARGET_WEEKDAY})",
    "",
    "One sheet per carer working today. This is the visible reasoning behind every number "
    "in caregivers.json and crid_prid_feasible.json for this same day -- not a separate "
    "analysis, the same one, exposed.",
    "",
    "Each sheet has 3 sections:",
    "1. EXTEND_FEASIBILITY DECISION -- the classifier's verdict (open/closed to unfamiliar "
    "clients) and the actual evidence behind it: her off-routine trial count, how many "
    "converted into real relationships, the Wilson confidence bound, and which clearing "
    "check (if any) applied.",
    "2. DATA-DRIVEN TRAVEL BOUNDS -- her max_distance_km/max_time_minutes (recency-weighted "
    "90th percentile of her own real trips) and border-crossing allowance (95th percentile "
    "gap beyond that), with the number of real observations behind each.",
    "3. TODAY'S FEASIBILITY ASSIGNMENTS -- every patient she's a candidate for today, her "
    "final weight, and the full breakdown that produced it (consistency %, status, "
    "recency decay, concentration factor).",
    "",
    f"Day situation: {today_carers_n} carers worked today vs a {avg_carers_recent:.1f} "
    f"recent-{TARGET_WEEKDAY} average (ratio {staffing_ratio:.2f}) -- "
    f"{'SHORT-STAFFED' if day_is_short_staffed else 'a normal/adequate day'}.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Extend?', 'Why', 'Max Distance (km)', 'Max Time (min)',
                    'Border Distance (km)', 'Border Time (min)', 'Today\'s Assignments']
for col, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()
row_i = 2

for carer in carers_today:
    shift = carer_shift_input[carer]
    crid = shift['crid']
    out_c = out_caregiver_by_crid.get(crid)
    if not out_c:
        continue
    ef = out_c['extend_feasibility']

    # --- classifier evidence ---
    trial = carer_trials.get(carer)
    if trial:
        n, k = trial['n_judged'], trial['k_converted']
        wb_bound = wilson_upper_bound(k, n) if n else None
        is_picky_flagged = carer in confirmed_picky
        was_cleared = carer in cleared
        classifier_reason = (
            f"{n} judged off-routine trial(s), {k} converted into real relationships "
            f"(Wilson 90% upper bound: {wb_bound:.2f})" if n else "No off-routine trial history"
        )
        if is_picky_flagged:
            classifier_reason += " -- FLAGGED as selective (extend=False from classifier)"
        elif was_cleared:
            classifier_reason += " -- flagged then CLEARED (insufficient tenure/opportunity/evidence)"
    else:
        classifier_reason = "No off-routine trial history at all -- no evidence either way"

    dynamic = dynamic_extend_decision(carer)
    if dynamic is not None:
        current_cl = carer_current_caseload(carer, TARGET_DATE)
        typical_cl = carer_typical_caseload(carer)
        classifier_reason += (f" | Dynamic caseload check: current={current_cl}, "
                               f"typical={typical_cl} -> deficit-based decision={dynamic}")

    has_hist = len(carer_all_visited_clients.get(carer, set())) > 0
    if not has_hist:
        classifier_reason += " | Brand-new carer, zero history -- defaults OPEN"

    # --- travel bounds evidence ---
    obs = []
    for wd2, visits2 in roster.get(carer, {}).items():
        for v2 in visits2:
            d2 = carer_client_km.get((carer, v2['client']))
            t2 = carer_client_min.get((carer, v2['client']))
            if d2 is not None and t2 is not None:
                obs.append((d2, t2))
    n_obs = len(obs)

    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=ef['extend']).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=classifier_reason).font = note_font
    summary_ws.cell(row=row_i, column=4, value=ef['max_distance_km']).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=ef['max_time_minutes']).font = normal_font
    summary_ws.cell(row=row_i, column=6, value=ef['max_distance_border_crossings_km']).font = normal_font
    summary_ws.cell(row=row_i, column=7, value=ef['max_time_border_crossings_minutes']).font = normal_font
    summary_ws.cell(row=row_i, column=8, value=len(feas_by_crid.get(crid, []))).font = normal_font
    row_i += 1

    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    r = 3

    ws.cell(row=r, column=1, value='1. EXTEND_FEASIBILITY DECISION').font = section_font
    r += 1
    ws.cell(row=r, column=1, value=f"Verdict: extend = {ef['extend']}").font = normal_font
    r += 1
    ws.cell(row=r, column=1, value=classifier_reason).font = note_font
    r += 2

    ws.cell(row=r, column=1, value='2. DATA-DRIVEN TRAVEL BOUNDS').font = section_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"max_distance_km={ef['max_distance_km']}, max_time_minutes={ef['max_time_minutes']} "
                   f"(recency-weighted 90th percentile of her {n_obs} real historical trips)")).font = normal_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"border_crossing: {ef['max_distance_border_crossings_km']}km / "
                   f"{ef['max_time_border_crossings_minutes']}min (95th percentile gap beyond her normal range)")).font = normal_font
    r += 2

    ws.cell(row=r, column=1, value="3. TODAY'S FEASIBILITY ASSIGNMENTS").font = section_font
    r += 1
    headers = ['Client', 'Weight', 'Consistency %', 'Status', 'Days Since Last Visit',
               'Recency Decay', 'Concentration Factor']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    rows_here = sorted(feas_by_crid.get(crid, []), key=lambda x: -x['weight'])
    for row in rows_here:
        pid = prid_to_pid.get(row['prid'])
        client = pid_to_client.get(pid, '?')
        req = next((rq for rq in day_requirements if rq['prid'] == row['prid']), None)
        start_min = req['start_min'] if req else 0
        bd = find_slot_breakdown(client, TARGET_WEEKDAY, start_min).get(carer)
        vals = [client, row['weight']]
        if bd:
            vals += [f"{bd['consistency_pct']}%", bd['status'], bd['days_since'],
                      round(bd['recency_decay'], 3), round(bd['concentration_factor'], 3)]
        else:
            vals += ['(no direct history -- geographic extend candidate)', '', '', '', '']
        fill = good_fill if row['weight'] >= 0.8 else (warn_fill if row['weight'] <= 0.3 else None)
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = normal_font
            if fill:
                cell.fill = fill
        r += 1

    widths = [30, 10, 14, 16, 20, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['C'].width = 70
for col in 'DEFGH':
    summary_ws.column_dimensions[col].width = 16
summary_ws.freeze_panes = 'A2'

out_path = f'{PROJECT_ROOT}/output/Carer_Decision_Analysis_{TARGET_DATE.isoformat()}.xlsx'
wb.save(out_path)
print(f"\nSaved {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
