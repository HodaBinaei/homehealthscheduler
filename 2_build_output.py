import pickle
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
os.makedirs('./output', exist_ok=True)

with open('./roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

# Aggregate: carer -> weekday -> (client, start_hhmm, end_hhmm) -> {count, first_date, last_date}
agg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'count':0,'first':None,'last':None})))

for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            st = v['start_dt'].strftime('%H:%M') if v['start_dt'] else v['start']
            et = v['end_dt'].strftime('%H:%M') if v['end_dt'] else v['end']
            key = (v['client'], st, et)
            rec = agg[carer][wd][key]
            rec['count'] += 1
            d = v['start_dt']
            if d and (rec['first'] is None or d < rec['first']):
                rec['first'] = d
            if d and (rec['last'] is None or d > rec['last']):
                rec['last'] = d

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='D9E1F2')
day_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)

def sheet_name(name, used):
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Summary sheet ---
summary_ws = wb.create_sheet('Summary')
summary_ws.append(['Carer', 'Days Worked (with Personal Care visits)', 'Total Distinct Visit Slots', 'Total Occurrences (visit-weeks)'])
for c in range(1, 5):
    cell = summary_ws.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for carer in sorted(agg.keys()):
    days_present = [wd for wd in WEEKDAYS if wd in agg[carer]]
    total_slots = sum(len(agg[carer][wd]) for wd in agg[carer])
    total_occ = sum(rec['count'] for wd in agg[carer] for rec in agg[carer][wd].values())
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_present)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=total_slots).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=total_occ).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 55
summary_ws.column_dimensions['C'].width = 22
summary_ws.column_dimensions['D'].width = 24
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
for carer in sorted(agg.keys()):
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:F1')

    r = 3
    headers = ['Weekday', 'Start Time', 'End Time', 'Client', 'Weeks Seen', 'Date Range']
    for wd in WEEKDAYS:
        if wd not in agg[carer]:
            continue
        ws.cell(row=r, column=1, value=wd).font = day_font
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        entries = sorted(agg[carer][wd].items(), key=lambda kv: kv[0][1])  # sort by start time
        for (client, st, et), rec in entries:
            date_range = ''
            if rec['first'] and rec['last']:
                if rec['first'].date() == rec['last'].date():
                    date_range = rec['first'].strftime('%d/%m/%Y')
                else:
                    date_range = f"{rec['first'].strftime('%d/%m/%Y')} - {rec['last'].strftime('%d/%m/%Y')}"
            ws.cell(row=r, column=1, value=wd).font = normal_font
            ws.cell(row=r, column=2, value=st).font = normal_font
            ws.cell(row=r, column=3, value=et).font = normal_font
            ws.cell(row=r, column=4, value=client).font = normal_font
            ws.cell(row=r, column=5, value=rec['count']).font = normal_font
            ws.cell(row=r, column=6, value=date_range).font = normal_font
            r += 1
        r += 1  # blank spacer row

    widths = [12, 11, 11, 32, 11, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

out_path = './output/Carer_Weekday_Roster.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
