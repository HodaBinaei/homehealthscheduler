import pickle, json, math
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

PROJECT_ROOT = '/home/hoda/Desktop/homehealthscheduler_local_pipeline'
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)
with open(f'{PROJECT_ROOT}/carer_presence.pkl', 'rb') as f:
    carer_presence = pickle.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90
DATA_START = datetime.date(2025, 6, 30)
DATA_END = datetime.date(2026, 8, 2)

def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()

with open(f'{PROJECT_ROOT}/data/clients-new.json') as f:
    clients_json = json.load(f)['client']
client_dates = {}
client_coords = {}  # normalized "firstname lastname" (matches roster's 'client' field) -> (lat, lon)
for c in clients_json:
    if c.get('status') != 'Active':
        continue
    first = (c.get('name') or '').strip()
    last = (c.get('lastname') or '').strip()
    key = norm(f"{last}, {first}")
    sd = c.get('start_date')
    ed = c.get('end_date') or c.get('termination_date')
    client_dates[key] = (
        datetime.date.fromisoformat(sd) if sd else None,
        datetime.date.fromisoformat(ed) if ed else None,
    )
    lat, lon = c.get('latitude'), c.get('longitude')
    if lat and lon:
        client_coords[norm(f"{first} {last}")] = (float(lat), float(lon))

with open(f'{PROJECT_ROOT}/data/users-new.json') as f:
    users_json = json.load(f)['user']
carer_coords = {}
for u in users_json:
    if u.get('status') != 'Active' or not u.get('is_caregiver'):
        continue
    first = (u.get('name') or '').strip()
    last = (u.get('lastname') or '').strip()
    display = f"{first} {last}".strip()
    lat, lon = u.get('latitude'), u.get('longitude')
    if lat and lon:
        carer_coords[display] = (float(lat), float(lon))

def haversine_km(p1, p2):
    lat1, lon1 = p1
    lat2, lon2 = p2
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

# Real driving distances (OSRM road-network km, from the carer/client pid/cid matrix),
# preferred over straight-line haversine wherever available -- haversine badly understates
# actual travel on real roads (lakes, rivers, missing bridges, etc.).
try:
    with open(f'{PROJECT_ROOT}/carer_client_distances.pkl', 'rb') as f:
        carer_client_km = pickle.load(f)
except FileNotFoundError:
    carer_client_km = {}

def travel_km(carer, client, home_coord=None, client_coord=None):
    d = carer_client_km.get((carer, client))
    if d is not None:
        return d
    if home_coord is not None and client_coord is not None:
        return haversine_km(home_coord, client_coord)
    return None

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

# --- double-up detection: genuine time-overlap between DIFFERENT carers visiting the
# same client on the same date (e.g. one carer 09:00-21:00 and another 09:00-10:00 helping
# at the start) -- built once, globally, same logic as the client-facing workbook.
client_day_visits = defaultdict(lambda: defaultdict(list))  # client -> date -> [(carer, start, end)]
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt'] and v['end_dt']:
                client_day_visits[v['client']][v['start_dt'].date()].append((carer, v['start_dt'], v['end_dt']))

def _overlaps(a_start, a_end, b_start, b_end):
    return a_start < b_end and b_start < a_end

def double_up_partners(client, carer, start_dt, end_dt):
    """Distinct carer names whose visit to this client on this date genuinely overlaps in
    time with this visit (excluding the same carer)."""
    date = start_dt.date()
    partners = set()
    for other_carer, o_start, o_end in client_day_visits.get(client, {}).get(date, []):
        if other_carer == carer:
            continue
        if _overlaps(start_dt, end_dt, o_start, o_end):
            partners.add(other_carer)
    return sorted(partners)

def double_up_summary(carer, client, occurrences):
    """occurrences: list of (req_start, req_end, actual_start, actual_end). Returns
    (fraction_display, partners_display) across all occurrences of this slot."""
    hits, partner_names = 0, set()
    for o in occurrences:
        partners = double_up_partners(client, carer, o[0], o[1])
        if partners:
            hits += 1
            partner_names.update(partners)
    if hits == 0:
        return '', ''
    return f"{hits / len(occurrences):.0%}", ', '.join(sorted(partner_names))

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

def iso_to_monday(yw):
    return datetime.date.fromisocalendar(yw[0], yw[1], 1)

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

def format_time_with_variance(dts):
    """Median time, with a 'varies X-Y' note appended when it isn't a single fixed value."""
    if not dts:
        return ''
    times = sorted(set(d.strftime('%H:%M') for d in dts))
    disp = median_time(dts)
    if len(times) > 1:
        disp = f"{disp} (varies {times[0]}-{times[-1]})"
    return disp

def combined_range(starts, ends):
    s = format_time_with_variance(starts)
    e = format_time_with_variance(ends)
    if not s:
        return ''
    return f"{s} - {e}" if e else s

def actual_time_summary(visits):
    """visits: list of raw visit dicts. Returns (scheduled_range, actual_range) strings, each
    'HH:MM - HH:MM', with a 'varies' note on either side when it isn't fixed across visits."""
    scheduled = combined_range([v['start_dt'] for v in visits if v.get('start_dt')],
                                [v['end_dt'] for v in visits if v.get('end_dt')])
    actual = combined_range([v['actual_start_dt'] for v in visits if v.get('actual_start_dt')],
                             [v['actual_end_dt'] for v in visits if v.get('actual_end_dt')])
    return scheduled, actual

# --- carer's own active weeks/days per weekday (as before) ---
carer_weekday_active_weeks = defaultdict(set)
carer_active_days = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))
                carer_active_days[carer].add(v['start_dt'].date())

DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5

def classify_relative(carer, wd, client, occurrences, daily_fallback=None):
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    actual_starts = [o[2] for o in occurrences if o[2] is not None]
    actual_ends = [o[3] for o in occurrences if o[3] is not None]
    dates = sorted(set(o[0].date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]
    carer_start, carer_end = carer_presence.get(carer, (DATA_START, DATA_END))
    client_start, client_end = client_dates.get(norm(client), (None, None))
    window_start, window_end = observed_start, observed_end

    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    win_start_wk, win_end_wk = isoweek(window_start), isoweek(window_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)
    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)

    if n_active < 3:
        pattern, ratio = 'Insufficient history', (n_hits / n_active if n_active else None)
        if daily_fallback is not None:
            daily_ratio, daily_active_days = daily_fallback
            ratio = daily_ratio
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'

    scheduled_disp = combined_range(starts, ends)
    actual_disp = combined_range(actual_starts, actual_ends)

    return {
        'pattern': pattern, 'ratio': ratio, 'start': median_time(starts),
        'scheduled': scheduled_disp, 'actual': actual_disp,
        'dates': dates, 'hit_weeks': hit_weeks, 'window_start': window_start, 'window_end': window_end,
        'win_start_wk': win_start_wk, 'win_end_wk': win_end_wk,
    }

# --- build the SET ROSTER: Weekly-pattern (carer, weekday, client) slots ---
set_roster = defaultdict(list)  # (carer, weekday) -> list of slot-info dicts (Weekly only)

for carer, wd_map in roster.items():
    per_client_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                per_client_tagged[v['client']].append(
                    (v['start_dt'], v['end_dt'], v.get('actual_start_dt'), v.get('actual_end_dt'), wd)
                )
    for client, tagged_occ in per_client_tagged.items():
        for cluster in cluster_by_time(tagged_occ):
            dates = sorted(set(o[0].date() for o in cluster))
            span_start, span_end = dates[0], dates[-1]
            weekdays_covered = set(WEEKDAYS[d.weekday()] for d in dates)
            active_days = [d for d in carer_active_days.get(carer, set()) if span_start <= d <= span_end]
            daily_fallback = None
            if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
                daily_fallback = (len(dates) / len(active_days), len(active_days))
            by_weekday = defaultdict(list)
            for o in cluster:
                by_weekday[o[4]].append(o[:4])
            for wd, occs in by_weekday.items():
                info = classify_relative(carer, wd, client, occs, daily_fallback=daily_fallback)
                if info['pattern'] == 'Weekly':
                    du_fraction, du_partners = double_up_summary(carer, client, occs)
                    set_roster[(carer, wd)].append({
                        'client': client, 'double_up': du_fraction, 'double_up_partners': du_partners, **info
                    })

print(f"Set-roster (Weekly) slots: {sum(len(v) for v in set_roster.values())}")

# Invert set_roster to a client-keyed lookup, so for any extra visit we can ask "does this
# client actually have a regular/set carer for this weekday, and were they unavailable?"
client_regular_carers = defaultdict(list)  # (client, weekday) -> list of (regular_carer, slot_info)
for (carer, wd), slots in set_roster.items():
    for slot in slots:
        client_regular_carers[(slot['client'], wd)].append((carer, slot))

def explain_extra_visit(client, wd, date):
    """Checks whether this client has a regular/set carer for this weekday, and whether
    that regular carer was unavailable (didn't visit) the week of this extra visit.
    Returns (has_set_roster: bool, regular_carers: list[str], explained_by: str or None)."""
    regular_slots = client_regular_carers.get((client, wd), [])
    if not regular_slots:
        return False, [], None
    wk = isoweek(date)
    regular_names = sorted(set(r for r, _ in regular_slots))
    for regular_carer, slot in regular_slots:
        if slot['win_start_wk'] <= wk <= slot['win_end_wk'] and wk not in slot['hit_weeks']:
            return True, regular_names, regular_carer
    return True, regular_names, None

# ==========================================================================
# PART 1 -- extra visits beyond the set roster
# ==========================================================================
set_roster_clients = defaultdict(set)  # (carer, weekday) -> set of client names in the set roster
for (carer, wd), slots in set_roster.items():
    set_roster_clients[(carer, wd)] = set(s['client'] for s in slots)

# Company-wide daily census -- how many distinct carers worked, how many distinct clients
# were seen, and how many total visits happened, on each calendar date -- used to give extra
# visits some staffing context (was this a generally short-staffed/busy day, or a one-off).
daily_carers = defaultdict(set)
daily_clients = defaultdict(set)
daily_visit_count = defaultdict(int)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                d = v['start_dt'].date()
                daily_carers[d].add(carer)
                daily_clients[d].add(v['client'])
                daily_visit_count[d] += 1

extra = defaultdict(lambda: {'clients': set(), 'count': 0, 'detail': Counter()})
extra_detail_timed = defaultdict(lambda: defaultdict(list))  # (carer, weekday) -> (client, sched_time) -> list of visit dicts
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        roster_clients = set_roster_clients.get((carer, wd), set())
        for v in visits:
            if v['client'] not in roster_clients:
                rec = extra[(carer, wd)]
                rec['clients'].add(v['client'])
                rec['count'] += 1
                rec['detail'][v['client']] += 1
                sched_time = v['start_dt'].strftime('%H:%M')
                extra_detail_timed[(carer, wd)][(v['client'], sched_time)].append(v)

print(f"Carer-weekday combos with extra (off-roster) visits: {len(extra)}")

# ==========================================================================
# GEOGRAPHIC COVERAGE -- for each carer, her "biggest travel" (home -> furthest
# set-roster client) used as a radius to find every other nearby active client, and how
# many of them she covers (roster or extra) versus never visits at all.
# ==========================================================================
set_roster_clients_by_carer = defaultdict(set)
for (carer, wd), slots in set_roster.items():
    for s in slots:
        set_roster_clients_by_carer[carer].add(s['client'])

extra_clients_by_carer = defaultdict(set)
for (carer, wd), rec in extra.items():
    extra_clients_by_carer[carer].update(rec['clients'])

geo_results = {}  # carer -> dict of geographic stats
client_exact_names = [f"{(c.get('name') or '').strip()} {(c.get('lastname') or '').strip()}".strip()
                       for c in clients_json if c.get('status') == 'Active']

for carer, roster_clients in set_roster_clients_by_carer.items():
    home = carer_coords.get(carer)
    if not home:
        continue
    # "biggest travel" should reflect everywhere she actually visits, not just her fixed
    # set roster -- an extra/ad hoc visit can easily be her furthest trip of all.
    all_visited_clients = roster_clients | extra_clients_by_carer.get(carer, set())
    all_client_coords = [
        (client, travel_km(carer, client, home, client_coords.get(norm(client))))
        for client in all_visited_clients if norm(client) in client_coords
    ]
    all_client_coords = [(c, d) for c, d in all_client_coords if d is not None]
    if not all_client_coords:
        continue
    furthest_client, max_dist = max(all_client_coords, key=lambda x: x[1])

    # The nearby-clients SEARCH radius is her own "normal" travel (median distance across
    # everywhere she actually visits) plus a 5km buffer -- not her single furthest trip.
    # A carer whose usual patch is tight but who has one rare far-flung outlier visit (e.g.
    # a one-off weekend cover 30km away) would otherwise get an unrealistically huge search
    # radius that swallows a huge area she never really operates in. Biggest Travel / Furthest
    # Client above are unchanged -- this only affects what counts as "nearby".
    all_distances = sorted(d for _, d in all_client_coords)
    n = len(all_distances)
    median_dist = (all_distances[n // 2 - 1] + all_distances[n // 2]) / 2 if n % 2 == 0 else all_distances[n // 2]
    search_radius = median_dist + 5

    nearby_set = set()
    for client_name in client_exact_names:
        d = travel_km(carer, client_name, home, client_coords.get(norm(client_name)))
        if d is not None and d <= search_radius:
            nearby_set.add(norm(client_name))
    extra_norm = set(norm(c) for c in extra_clients_by_carer.get(carer, set()))
    roster_norm = set(norm(c) for c in roster_clients)
    already_roster_nearby = nearby_set & roster_norm
    covered_nearby = (nearby_set & extra_norm) - already_roster_nearby

    geo_results[carer] = {
        'max_dist_km': max_dist, 'furthest_client': furthest_client,
        'search_radius_km': search_radius,
        'n_roster_clients': len(roster_clients),
        'n_nearby_total': len(nearby_set),
        'n_nearby_already_roster': len(already_roster_nearby),
        'n_nearby_covered_extra': len(covered_nearby),
        'n_nearby_uncovered': len(nearby_set) - len(already_roster_nearby) - len(covered_nearby),
    }

print(f"Carers with full geographic data (home + roster client coords): {len(geo_results)}")

# ==========================================================================
# PART 2 -- substitute coverage: for set-roster slots, who covers when the
# regular carer isn't the one who shows up
# ==========================================================================
# client-side lookup: (client, weekday) -> list of (iso_week, carer, req_start_dt, req_end_dt, actual_start_dt, actual_end_dt)
client_weekday_visits = defaultdict(list)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                client_weekday_visits[(v['client'], wd)].append(
                    (isoweek(v['start_dt'].date()), carer, v['start_dt'], v['end_dt'],
                     v.get('actual_start_dt'), v.get('actual_end_dt'))
                )

coverage_events = []  # list of dicts, one per missed week on a set-roster slot

for (carer, wd), slots in set_roster.items():
    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    for slot in slots:
        client = slot['client']
        win_start_wk, win_end_wk = slot['win_start_wk'], slot['win_end_wk']
        # every ISO week in the slot's own span
        span_weeks = []
        cur = iso_to_monday(win_start_wk)
        end_monday = iso_to_monday(win_end_wk)
        while cur <= end_monday:
            span_weeks.append(isoweek(cur))
            cur += datetime.timedelta(days=7)

        target_minute = None
        # approximate slot time from its 'start' string
        hh, mm = slot['start'].split(':')
        target_minute = int(hh) * 60 + int(mm)

        for wk in span_weeks:
            if wk in slot['hit_weeks']:
                continue  # regular carer did the visit that week -- no gap
            carer_was_active = wk in active_weeks_all
            # look for who (if anyone) visited this client on this weekday in this ISO week
            # at a similar time-of-day, excluding the regular carer
            candidates = [
                (c, dt, edt, adt, aet) for (w, c, dt, edt, adt, aet) in client_weekday_visits.get((client, wd), [])
                if w == wk and c != carer
                and abs((dt.hour * 60 + dt.minute) - target_minute) <= TIME_GAP_MINUTES
            ]
            monday = iso_to_monday(wk)
            if candidates:
                sub_carer, sub_dt, sub_edt, sub_actual_start, sub_actual_end = candidates[0]
                status = 'Covered by another carer'
            elif any(w == wk and c == carer for (w, c, dt, edt, adt, aet) in client_weekday_visits.get((client, wd), [])):
                # regular carer DID visit this client that week, just not matched to this
                # exact time-of-day cluster (e.g. time moved a lot that week) -- not a real gap
                continue
            else:
                sub_carer, sub_dt, sub_edt, sub_actual_start, sub_actual_end = None, None, None, None, None
                status = 'No visit found that week'
            coverage_events.append({
                'carer': carer, 'weekday': wd, 'client': client, 'scheduled_time': slot['scheduled'],
                'week_of': monday, 'carer_active_that_weekday': carer_was_active,
                'status': status, 'covering_carer': sub_carer,
                'covering_scheduled_time': combined_range([sub_dt], [sub_edt]) if sub_dt else '',
                'covering_actual_time': combined_range([sub_actual_start], [sub_actual_end]) if sub_actual_start else '',
            })

print(f"Coverage gap events on set-roster slots: {len(coverage_events)}")
covered_by_someone = sum(1 for e in coverage_events if e['covering_carer'])
print(f"  ...of which covered by a named substitute: {covered_by_someone}")

# ==========================================================================
# Build workbook
# ==========================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
covered_fill = PatternFill('solid', fgColor='FCE4D6')
uncovered_fill = PatternFill('solid', fgColor='F2F2F2')
double_up_row_fill = PatternFill('solid', fgColor='FFE699')

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Roster coverage: extra visits and substitute cover",
    "",
    "SET ROSTER = every (carer, weekday, client) relationship classified WEEKLY in the main "
    "roster workbook (>=75% consistency in the carer's own active weeks, or the daily-pooled "
    "fallback -- see that workbook's Read Me for full methodology).",
    "",
    "Each carer's sheet has three sections:",
    "--- SET ROSTER ---",
    "The explicit list of every client on that carer's fixed weekly schedule, by weekday, with "
    "their scheduled time and exact consistency %. This is what 'Set Roster Size' on the Summary "
    "sheet is counting.",
    "",
    "Double-Up / Double-Up Partner(s) columns: whether ANOTHER carer genuinely overlaps in time "
    "with this slot on the same date (e.g. this carer 09:00-21:00 and another carer 09:00-10:00 "
    "helping at the start) -- not just a same-day coincidence, an actual overlapping window. "
    "Double-Up shows what fraction of this slot's visits had a real overlapping partner; "
    "Double-Up Partner(s) names who. Rows with any double-up are highlighted amber -- useful for "
    "spotting which visits genuinely need two carers (e.g. hoist transfers) versus which are solo.",
    "",
    "--- EXTRA VISITS ---",
    "A compact per-weekday summary (people/times), followed by an EXTRA VISITS DETAIL table "
    "showing exactly what time each off-roster visit happened, per client -- so 'Anthony Kelly "
    "x14' becomes visible as, e.g., 5 visits at 09:00 and 9 visits at 14:00.",
    "For each carer and weekday, this counts visits to clients who are NOT part of that carer's "
    "set roster for that weekday -- i.e. people outside their normal fixed schedule.",
    "- 'Extra People' = distinct clients visited outside the set roster (how many different "
    "people).",
    "- 'Extra Visit Occurrences' = total number of such visits (how many times), which can be "
    "higher than Extra People if the same off-roster client was covered more than once.",
    "",
    "EXTRA VISITS DETAIL also answers WHY each extra visit happened:",
    "- 'Carers Working That Day (avg)' / 'Clients Seen That Day (avg)' / 'Total Visits That Day "
    "(avg)': company-wide staffing context for the specific date(s) this extra visit happened "
    "(any carer, any client, not just this one) -- Carers Working and Clients Seen are DISTINCT "
    "counts (headcount), while Total Visits is the raw visit count, which is naturally much "
    "higher since most clients get more than one visit a day (morning/afternoon/evening). All "
    "three grow substantially over the export period as the roster expanded, so a low number "
    "usually means this extra visit happened early on (e.g. mid-2025), not that the day was "
    "unusually quiet. Useful for spotting whether extra visits cluster on generally "
    "short-staffed or unusually busy days versus being spread evenly across normal days.",
    "- 'Client Has Set Roster?': whether this client has a regular/set carer AT ALL for this "
    "weekday (from anyone's set roster, not just this carer).",
    "- 'Regular Carer(s)': who that regular carer normally is, if one exists.",
    "- 'Likely Reason': cross-checks the visit's date against the regular carer's own set-roster "
    "history for that client -- 'Covering for regular carer's absence' means the regular carer "
    "didn't visit that specific week (matches a Substitute Coverage gap); 'no set roster' means "
    "this is a genuinely ad hoc/occasional client with no fixed carer at all; 'reason unclear' "
    "means the regular carer WAS active that week too, so this extra visit looks like an "
    "additional need rather than a covering-for-absence case.",
    "",
    "--- Substitute Coverage ---",
    "For every set-roster slot, every week within its own span where the REGULAR carer did NOT "
    "do the visit is one row. For each such week:",
    "- 'Covered by another carer': someone else visited that client on that weekday at a similar "
    "time (within 90 minutes) that week -- shown in the Covering Carer / Covering Time columns. "
    "Highlighted orange.",
    "- 'No visit found that week': no one visited that client at that time-of-day that week at "
    "all, as far as the data shows -- the client may have gone without that visit, or it may have "
    "moved outside the normal time window. Highlighted grey.",
    "- 'Carer Active That Weekday' = TRUE if the regular carer was working that weekday at all "
    "that week (just not this particular visit) vs FALSE if they weren't working that weekday at "
    "all that week (more likely a genuine day off/leave).",
    "",
    "--- Geographic Coverage ---",
    "Each carer's own 'biggest travel' -- the distance from her home address to the furthest "
    "client she actually visits, whether that's a set-roster client or an extra/ad hoc one -- "
    "is shown as-is (Biggest Travel / Furthest Client), but is NOT what defines the nearby-client "
    "search radius. A carer whose usual patch is tight but who has one rare far-flung outlier "
    "visit (e.g. a one-off weekend cover 30km away) would otherwise get an unrealistically huge "
    "search radius that swallows a huge area she never really operates in. Instead, Search "
    "Radius = her median travel distance across everywhere she actually visits, plus a 5km "
    "buffer -- a much more realistic reflection of her normal working area. Every other active "
    "client within that radius counts as 'nearby', and the rest of the columns (Already On Her "
    "Roster, Covered By Extra Visits, Never Visited) are computed against that radius.",
    "",
    "Distances are REAL driving distances (road-network km, from an OSRM-computed carer/client "
    "distance matrix keyed by their pid/cid), not straight-line -- straight-line (haversine) "
    "distance is used only as a fallback for the handful of pairs missing from the matrix.",
    "",
    "Carers without a home coordinate, or without any coordinate-matched visited client, don't "
    "get this section.",
    "",
    f"Totals: {len(coverage_events)} coverage-gap weeks across all set-roster slots, of which "
    f"{covered_by_someone} were covered by a named substitute carer.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 130

def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Summary sheet ---
all_carers = sorted(set(c for (c, wd) in set_roster.keys()) | set(c for (c, wd) in extra.keys()))
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Set Roster (clients)', 'Extra People (total)', 'Extra Visits (total)',
                    'Coverage Gap Weeks', 'Covered by Substitute', 'Biggest Travel (km)',
                    '% Nearby Covered']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

coverage_by_carer = defaultdict(list)
for e in coverage_events:
    coverage_by_carer[e['carer']].append(e)

row_i = 2
for carer in all_carers:
    roster_size = len(set(s['client'] for wd in WEEKDAYS for s in set_roster.get((carer, wd), [])))
    carer_extra_sets = [extra[(carer, wd)]['clients'] for wd in WEEKDAYS if (carer, wd) in extra]
    extra_people = len(set().union(*carer_extra_sets)) if carer_extra_sets else 0
    extra_visits = sum(extra[(carer, wd)]['count'] for wd in WEEKDAYS if (carer, wd) in extra)
    gap_weeks = len(coverage_by_carer.get(carer, []))
    covered = sum(1 for e in coverage_by_carer.get(carer, []) if e['covering_carer'])
    geo = geo_results.get(carer)
    if geo:
        covered_total = geo['n_nearby_already_roster'] + geo['n_nearby_covered_extra']
        geo_pct = f"{covered_total / geo['n_nearby_total']:.0%}" if geo['n_nearby_total'] else ''
        geo_dist = round(geo['max_dist_km'], 1)
    else:
        geo_pct, geo_dist = '', ''
    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=roster_size).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=extra_people).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=extra_visits).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=gap_weeks).font = normal_font
    summary_ws.cell(row=row_i, column=6, value=covered).font = normal_font
    summary_ws.cell(row=row_i, column=7, value=geo_dist).font = normal_font
    summary_ws.cell(row=row_i, column=8, value=geo_pct).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
for col in 'BCDEFGH':
    summary_ws.column_dimensions[col].width = 20
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
ev_headers = ['Weekday', 'Set Roster Size', 'Extra People', 'Extra Visit Occurrences', 'Extra Clients (names)']
sc_headers = ['Weekday', 'Client', 'Service Required Start and End', 'Week Of', 'Carer Active That Weekday',
              'Status', 'Covering Carer', 'Covering Service Required Start and End', 'Covering Actual Start and End']

for carer in all_carers:
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:H1')

    r = 3
    ws.cell(row=r, column=1, value='SET ROSTER (the fixed weekly schedule)').font = Font(name=FONT, bold=True, size=12)
    r += 1
    set_roster_headers = ['Weekday', 'Client', 'Service Required Start and End', 'Actual Start and End',
                           'Consistency', 'Double-Up', 'Double-Up Partner(s)']
    for col, h in enumerate(set_roster_headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    for wd in WEEKDAYS:
        for slot in sorted(set_roster.get((carer, wd), []), key=lambda s: s['start']):
            row_vals = [wd, slot['client'], slot['scheduled'], slot['actual'],
                        f"{slot['ratio']:.0%}" if slot['ratio'] is not None else '',
                        slot['double_up'], slot['double_up_partners']]
            fill = double_up_row_fill if slot['double_up'] else None
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = normal_font
                if fill:
                    cell.fill = fill
            r += 1
    r += 2

    ws.cell(row=r, column=1, value='EXTRA VISITS (beyond set roster)').font = Font(name=FONT, bold=True, size=12)
    r += 1
    for col, h in enumerate(ev_headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    for wd in WEEKDAYS:
        if (carer, wd) not in extra:
            continue
        rec = extra[(carer, wd)]
        roster_size = len(set_roster_clients.get((carer, wd), set()))
        names = ', '.join(f"{name} (x{n})" for name, n in rec['detail'].most_common())
        row_vals = [wd, roster_size, len(rec['clients']), rec['count'], names]
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = note_font if col == 5 else normal_font
        r += 1
    r += 2

    ws.cell(row=r, column=1, value='EXTRA VISITS DETAIL (by time)').font = Font(name=FONT, bold=True, size=12)
    r += 1
    ev_detail_headers = ['Weekday', 'Client', 'Service Required Start and End', 'Actual Start and End', 'Visits',
                          'Carers Working That Day (avg)', 'Clients Seen That Day (avg)',
                          'Total Visits That Day (avg)', 'Client Has Set Roster?', 'Regular Carer(s)', 'Likely Reason']
    for col, h in enumerate(ev_detail_headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    for wd in WEEKDAYS:
        timed = extra_detail_timed.get((carer, wd))
        if not timed:
            continue
        for (client, sched_time), visits in sorted(timed.items(), key=lambda kv: (kv[0][1], kv[0][0])):
            scheduled_range, actual_range = actual_time_summary(visits)

            has_roster_flags, regular_names_all, explained_count, explaining_carers = [], set(), 0, Counter()
            for v in visits:
                has_roster, regular_names, explained_by = explain_extra_visit(client, wd, v['start_dt'].date())
                has_roster_flags.append(has_roster)
                regular_names_all.update(regular_names)
                if explained_by:
                    explained_count += 1
                    explaining_carers[explained_by] += 1

            has_set_roster = 'Yes' if any(has_roster_flags) else 'No'
            regular_carers_display = ', '.join(sorted(regular_names_all))
            total = len(visits)
            if not any(has_roster_flags):
                reason = 'Client has no set roster for this weekday (ad hoc/occasional client)'
            elif explained_count == total:
                who = ', '.join(f"{c} ({n}/{total})" for c, n in explaining_carers.most_common())
                reason = f"Covering for regular carer's absence -- {who}"
            elif explained_count > 0:
                who = ', '.join(f"{c} ({n}/{total})" for c, n in explaining_carers.most_common())
                reason = f"Partly covering absence ({explained_count}/{total}: {who}); rest unclear (possible additional need)"
            else:
                reason = 'Regular carer was present those weeks -- reason unclear (possibly an additional/extra need)'

            occ_dates = [v['start_dt'].date() for v in visits]
            avg_carers = sum(len(daily_carers.get(d, set())) for d in occ_dates) / len(occ_dates)
            avg_clients = sum(len(daily_clients.get(d, set())) for d in occ_dates) / len(occ_dates)
            avg_total_visits = sum(daily_visit_count.get(d, 0) for d in occ_dates) / len(occ_dates)

            row_vals = [wd, client, scheduled_range, actual_range, len(visits),
                        round(avg_carers, 1), round(avg_clients, 1), round(avg_total_visits, 1),
                        has_set_roster, regular_carers_display, reason]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 11 else normal_font
            r += 1
    r += 2

    ws.cell(row=r, column=1, value='SUBSTITUTE COVERAGE (set-roster weeks the regular carer missed)').font = Font(name=FONT, bold=True, size=12)
    r += 1
    for col, h in enumerate(sc_headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    events = sorted(coverage_by_carer.get(carer, []), key=lambda e: (WEEKDAYS.index(e['weekday']), e['week_of']))
    for e in events:
        row_vals = [e['weekday'], e['client'], e['scheduled_time'], e['week_of'].strftime('%d/%m/%Y'),
                    e['carer_active_that_weekday'], e['status'], e['covering_carer'] or '',
                    e['covering_scheduled_time'], e['covering_actual_time']]
        fill = covered_fill if e['covering_carer'] else uncovered_fill
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = normal_font
            cell.fill = fill
        r += 1
    r += 2

    geo = geo_results.get(carer)
    if geo:
        ws.cell(row=r, column=1, value='GEOGRAPHIC COVERAGE').font = Font(name=FONT, bold=True, size=12)
        r += 1
        geo_headers = ['Biggest Travel (km)', 'Furthest Client', 'Search Radius (km)', 'Set Roster Clients',
                        'Nearby Clients (within radius)', 'Already On Her Roster',
                        'Covered By Extra Visits', 'Never Visited', '% Nearby Covered (roster+extra)']
        for col, h in enumerate(geo_headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        covered_total = geo['n_nearby_already_roster'] + geo['n_nearby_covered_extra']
        pct = covered_total / geo['n_nearby_total'] if geo['n_nearby_total'] else 0
        row_vals = [round(geo['max_dist_km'], 1), geo['furthest_client'], round(geo['search_radius_km'], 1),
                    geo['n_roster_clients'], geo['n_nearby_total'], geo['n_nearby_already_roster'],
                    geo['n_nearby_covered_extra'], geo['n_nearby_uncovered'], f"{pct:.0%}"]
        for col, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=col, value=val).font = normal_font
        r += 1

    widths = [14, 30, 22, 22, 60, 24, 26, 24, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

out_path = f'{PROJECT_ROOT}/output/Carer_Roster_Coverage.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
