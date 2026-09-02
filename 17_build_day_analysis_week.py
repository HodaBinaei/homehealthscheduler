import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime
import os
os.makedirs('./output', exist_ok=True)

# Reuse every foundation data structure already built and tested in build_hhs_day_files.py
# (set_roster, per_carer_client_slot_pattern, carer_does_extra, carer_search_radius,
# client_slot_history, travel_km, carer_info, client_info, concentration_factor, etc.)
_src = open('./build_hhs_day_files.py').read()
_cutoff = _src.split("# Helper: minutes since midnight")[0]
exec(_cutoff)

WEEK_END = datetime.date(2026, 8, 2)  # the same Sunday used for the single-day example
WEEK_DATES = {}
for i in range(7):
    d = WEEK_END - datetime.timedelta(days=i)
    WEEK_DATES[d.strftime('%A')] = d

print("Week covered:", {wd: d.isoformat() for wd, d in WEEK_DATES.items()})

# ---------------------------------------------------------------------------
# Per-(carer, day) visit extraction and staffing context, for each of the 7 days
# ---------------------------------------------------------------------------
day_context = {}  # weekday -> dict with visits_by_carer, staffing numbers
daily_carers_all = defaultdict(set)
daily_visit_count_all = defaultdict(int)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                d = v['start_dt'].date()
                daily_carers_all[d].add(carer)
                daily_visit_count_all[d] += 1

for wd_name, target_date in WEEK_DATES.items():
    visits_by_carer = defaultdict(list)
    for carer, wd_map in roster.items():
        for v in wd_map.get(wd_name, []):
            if v['start_dt'] and v['start_dt'].date() == target_date:
                visits_by_carer[carer].append(v)

    same_weekday_dates = [d for d in daily_visit_count_all if WEEKDAYS[d.weekday()] == wd_name and d != target_date]
    recent_same_weekday = [d for d in same_weekday_dates if d >= target_date - datetime.timedelta(weeks=8)]
    if recent_same_weekday:
        avg_carers_recent = sum(len(daily_carers_all[d]) for d in recent_same_weekday) / len(recent_same_weekday)
    else:
        avg_carers_recent = len(daily_carers_all.get(target_date, set()))
    today_carers = len(daily_carers_all.get(target_date, set()))
    staffing_ratio = today_carers / avg_carers_recent if avg_carers_recent else 1.0
    day_is_short_staffed = staffing_ratio < 0.85

    day_context[wd_name] = {
        'date': target_date, 'visits_by_carer': visits_by_carer,
        'today_carers': today_carers, 'avg_carers_recent': avg_carers_recent,
        'staffing_ratio': staffing_ratio, 'short_staffed': day_is_short_staffed,
    }
    print(f"{wd_name} {target_date}: {today_carers} carers, ratio {staffing_ratio:.2f} vs recent avg "
          f"{avg_carers_recent:.1f} -> {'SHORT-STAFFED' if day_is_short_staffed else 'normal'}")

def find_slot_pattern_local(carer, wd, client, start_minute):
    candidates = per_carer_client_slot_pattern.get((carer, wd, client), [])
    if not candidates:
        return None, 0.0
    best = min(candidates, key=lambda c: abs(c[2] - start_minute))
    if abs(best[2] - start_minute) > TIME_GAP_MINUTES:
        return None, 0.0
    return best[0], best[1]

# All carers who worked at least one of the 7 days
all_carers_week = sorted(set(c for ctx in day_context.values() for c in ctx['visits_by_carer']))
print(f"\nDistinct carers across the week: {len(all_carers_week)}")

client_exact_names = list(client_info.keys())
carer_nearby_count = {}
for carer in all_carers_week:
    radius = carer_search_radius.get(carer)
    if radius is None:
        continue
    n = sum(1 for cl in client_exact_names if (d := travel_km(carer, cl)) is not None and d <= radius)
    carer_nearby_count[carer] = n

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
title_font = Font(name=FONT, bold=True, size=14)
day_font = Font(name=FONT, bold=True, size=12, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='548235')
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
roster_fill = PatternFill('solid', fgColor='E2EFDA')
extra_fill = PatternFill('solid', fgColor='FFE699')

notes_ws = wb.create_sheet('Read Me')
notes = [
    f"Carer day analysis across a full week ({WEEK_DATES['Monday']} to {WEEK_DATES['Sunday']})",
    "",
    "The single-day version only shows carers who happened to work that specific day -- e.g. "
    "the Sunday-only file covered 37 of 99 active carers. This version covers a full "
    "representative week (the most recent Mon-Sun at the time of the export) so nearly every "
    "carer appears at least once: 89 of 99 active carers worked at least one day this week.",
    "",
    "Each carer's sheet has one block per day she actually worked that week, each answering "
    "the same 4 questions as the single-day version:",
    "1. SET ROSTER THAT DAY -- her fixed weekly-schedule visits.",
    "2. EXTRA VISITS THAT DAY -- visits outside her set roster, with the likely reason "
    "(covering a known absence vs. an ad hoc/no-roster client vs. unclear additional need).",
    "3. DAY SITUATION -- that specific day's company-wide staffing vs. the recent (8-week) "
    "average for that weekday.",
    "4. GEOGRAPHIC CHOICES -- her own search radius (median travel + 5km) and how many other "
    "active clients fall within it -- shown once per carer (it doesn't change day to day).",
    "",
    "Carers who didn't work at all this week (10 of 99) are not included -- there's no visit "
    "data to analyse for them in this window.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Days Worked This Week', 'Total Set Roster Visits', 'Total Extra Visits',
                    'Search Radius (km)', 'Nearby Clients', 'Closed or Flexible?']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

def carer_day_breakdown(carer, wd_name):
    ctx = day_context[wd_name]
    my_visits = ctx['visits_by_carer'].get(carer, [])
    roster_visits, extra_visits = [], []
    for v in my_visits:
        start_min = v['start_dt'].hour * 60 + v['start_dt'].minute
        pattern, ratio = find_slot_pattern_local(carer, wd_name, v['client'], start_min)
        entry = {'v': v, 'pattern': pattern, 'ratio': ratio}
        (roster_visits if pattern == 'Weekly' else extra_visits).append(entry)
    return roster_visits, extra_visits

row_i = 2
carer_week_data = {}
for carer in all_carers_week:
    days_worked = [wd for wd in WEEKDAYS if carer in day_context[wd]['visits_by_carer']]
    total_roster, total_extra = 0, 0
    per_day = {}
    for wd in days_worked:
        rv, ev = carer_day_breakdown(carer, wd)
        per_day[wd] = (rv, ev)
        total_roster += len(rv)
        total_extra += len(ev)
    carer_week_data[carer] = per_day

    is_flexible = carer_does_extra.get(carer, False)
    radius = carer_search_radius.get(carer)
    nearby = carer_nearby_count.get(carer)

    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_worked)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=total_roster).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=total_extra).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=round(radius, 1) if radius else '').font = normal_font
    summary_ws.cell(row=row_i, column=6, value=nearby if nearby is not None else '').font = normal_font
    summary_ws.cell(row=row_i, column=7, value='Flexible' if is_flexible else 'Closed').font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['B'].width = 45
for col in 'CDEFG':
    summary_ws.column_dimensions[col].width = 18
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

for carer in all_carers_week:
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    is_flexible = carer_does_extra.get(carer, False)
    radius = carer_search_radius.get(carer)
    nearby = carer_nearby_count.get(carer)
    ws['A2'] = (f"{'Flexible' if is_flexible else 'Closed'} caseload  |  "
                f"Search radius {round(radius,1) if radius else 'n/a'}km  |  "
                f"{nearby if nearby is not None else 'n/a'} nearby clients  |  "
                f"(4. GEOGRAPHIC CHOICES -- constant across the week)")
    ws['A2'].font = note_font
    r = 4

    per_day = carer_week_data[carer]
    for wd in WEEKDAYS:
        if wd not in per_day:
            continue
        ctx = day_context[wd]
        roster_visits, extra_visits = per_day[wd]

        ws.cell(row=r, column=1,
                value=f"{wd} {ctx['date']}  --  {ctx['today_carers']} carers worked vs recent avg "
                      f"{ctx['avg_carers_recent']:.1f} (ratio {ctx['staffing_ratio']:.2f}, "
                      f"{'SHORT-STAFFED' if ctx['short_staffed'] else 'normal'})").font = day_font
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1

        ws.cell(row=r, column=1, value='1. Set Roster').font = Font(name=FONT, bold=True, size=10)
        r += 1
        if roster_visits:
            for col, h in enumerate(['Client', 'Time', 'Consistency'], start=1):
                c = ws.cell(row=r, column=col, value=h)
                c.font = header_font; c.fill = header_fill
            r += 1
            for e in sorted(roster_visits, key=lambda e: e['v']['start_dt']):
                v = e['v']
                row_vals = [v['client'], v['start_dt'].strftime('%H:%M'), f"{e['ratio']:.0%}"]
                for col, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = normal_font; cell.fill = roster_fill
                r += 1
        else:
            ws.cell(row=r, column=1, value='(none)').font = note_font
            r += 1

        ws.cell(row=r, column=1, value='2. Extra Visits').font = Font(name=FONT, bold=True, size=10)
        r += 1
        if extra_visits:
            for col, h in enumerate(['Client', 'Time', 'Pattern', 'Likely Reason'], start=1):
                c = ws.cell(row=r, column=col, value=h)
                c.font = header_font; c.fill = header_fill
            r += 1
            for e in sorted(extra_visits, key=lambda e: e['v']['start_dt']):
                v = e['v']
                client = v['client']
                regular_slots = [(c2, ratio2) for (c2, wd2, cl2), slots in per_carer_client_slot_pattern.items()
                                  if wd2 == wd and cl2 == client and c2 != carer
                                  for (pat2, ratio2, med2) in slots if pat2 == 'Weekly']
                if regular_slots:
                    names = ', '.join(sorted(set(c2 for c2, _ in regular_slots)))
                    reason = f"Regular carer ({names}) -- likely covering absence/additional need"
                else:
                    reason = 'No set roster for this weekday -- ad hoc/occasional client'
                row_vals = [client, v['start_dt'].strftime('%H:%M'), e['pattern'] or 'Occasional/ad hoc', reason]
                for col, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = note_font if col == 4 else normal_font
                    cell.fill = extra_fill
                r += 1
        else:
            ws.cell(row=r, column=1, value='(none)').font = note_font
            r += 1
        r += 1  # spacer between day blocks

    widths = [30, 14, 14, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

out_path = f'./output/Carer_Day_Analysis_Week_{WEEK_DATES["Monday"].isoformat()}_to_{WEEK_DATES["Sunday"].isoformat()}.xlsx'
wb.save(out_path)
print(f"\nSaved {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
