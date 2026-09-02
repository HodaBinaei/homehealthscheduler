import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

import os
PROJECT_ROOT = os.environ.get(
    'HHS_PROJECT_ROOT',
    os.path.dirname(os.path.abspath(__file__)),
)
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/cancellation_analysis.json') as f:
    results = json.load(f)

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
fills = {
    'Always cancelled': PatternFill('solid', fgColor='F2F2F2'),
    'Never cancelled': PatternFill('solid', fgColor='E2EFDA'),
    'Occasionally cancelled': PatternFill('solid', fgColor='FCE4D6'),
    'Insufficient history': PatternFill('solid', fgColor='FFFFFF'),
}

notes_ws = wb.create_sheet('Read Me')
notes = [
    "Cancellation pattern analysis",
    "",
    "For each (client, weekday, time-slot), this looks at every scheduled occurrence across "
    "the full export and classifies how reliably it actually happens, using TWO signals that "
    "were previously silently dropped from all other analysis:",
    "- Cancellation Description reasons that reflect a genuine OPERATIONAL failure to cover the "
    "call: VNR, Missed call, Cancelled <12h/24h, Covered by Another Agency.",
    "- Rows where Actual Employee Name is a system placeholder (*Ryan, Moira / *Management, "
    "Alert / *Management 2/3, Alert / *Temporary, Roster) -- these aren't real people (confirmed: "
    "none exist as any user record, active or inactive) and represent an unfulfilled requirement "
    "even though the Cancellation Description field itself is blank.",
    "",
    "Reasons that reflect the CLIENT not needing the visit that day, unrelated to staffing "
    "(Hospital, Holiday, Bank Holiday, Respite, Coronavirus-related) are excluded entirely from "
    "the cancellation-rate calculation -- including them would wrongly penalise clients who are "
    "simply often hospitalised as if their slot were unreliable.",
    "",
    "Classification (needs >=3 total fulfilled+cancelled occurrences to classify at all):",
    "- Never cancelled (<=5% rate): a firm, reliable commitment -- should get high scheduling "
    "priority and not be dropped.",
    "- Always cancelled (>=90% rate): this slot essentially never really happens -- low priority, "
    "may not be worth hard-committing scheduler resources to.",
    "- Occasionally cancelled (in between): the Fairness Flag notes whether cancellations were "
    "concentrated on ONE planned carer's shifts repeatedly (a possible equity concern -- the same "
    "person's calls keep being the one dropped) versus spread across several different planned "
    "carers (more consistent with genuine, evenly-distributed short-staffing).",
    "",
    f"Total (client, weekday, slot) combinations analysed: {len(results)}",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

ws = wb.create_sheet('Cancellation Patterns')
headers = ['Client', 'Weekday', 'Time', 'Fulfilled', 'Op. Cancelled', 'Excused (client-driven)',
           'Cancellation Rate', 'Classification', 'Fairness Flag']
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

results_sorted = sorted(results, key=lambda r: (r['client'], WEEKDAYS.index(r['weekday']), r['time']))
row_i = 2
for r in results_sorted:
    rate_display = f"{r['cancellation_rate']:.0%}" if r['cancellation_rate'] is not None else ''
    row_vals = [r['client'], r['weekday'], r['time'], r['fulfilled'], r['op_cancelled'], r['excused'],
                rate_display, r['classification'], r['fairness_flag']]
    fill = fills.get(r['classification'])
    for col, val in enumerate(row_vals, start=1):
        cell = ws.cell(row=row_i, column=col, value=val)
        cell.font = note_font if col == 9 else normal_font
        if fill:
            cell.fill = fill
    row_i += 1

widths = [30, 12, 10, 11, 14, 20, 16, 22, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

out_path = f'{PROJECT_ROOT}/output/Cancellation_Pattern_Analysis.xlsx'
wb.save(out_path)
print("Saved", out_path)
