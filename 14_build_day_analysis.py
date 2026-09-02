import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import os
os.makedirs('./output', exist_ok=True)

# Reuse every foundation data structure already built and tested in build_hhs_day_files.py
# (set_roster, per_carer_client_slot_pattern, carer_does_extra, carer_search_radius,
# day_visits, requests, day_is_short_staffed, staffing_ratio, client_slot_history, etc.)
# up to (but not including) the Patient/Caregiver/Feasibility generation itself.
_src = open('./build_hhs_day_files.py').read()
_cutoff = _src.split("# Helper: minutes since midnight")[0]
exec(_cutoff)

# ---------------------------------------------------------------------------
# Per-carer day analysis, answering the 4 original questions:
# 1. Does she have a set roster that day?
# 2. Does she cover extra?
# 3. If so, what's the day's situation, and does that explain it?
# 4. Does she have enough nearby choices to cover, or is she geographically boxed in?
# ---------------------------------------------------------------------------
carers_today = sorted(set(v['carer'] for v in day_visits))
visits_by_carer = defaultdict(list)
for v in day_visits:
    visits_by_carer[v['carer']].append(v)

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
title_font = Font(name=FONT, bold=True, size=14)
section_font = Font(name=FONT, bold=True, size=12)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
roster_fill = PatternFill('solid', fgColor='E2EFDA')
extra_fill = PatternFill('solid', fgColor='FFE699')

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    f"Day analysis for {TARGET_DATE} ({TARGET_WEEKDAY})",
    "",
    "This is the human-readable reasoning behind the phase-2 feasibility files "
    "(patient/caregivers/crid_prid_feasibility JSON) for this specific day -- one sheet per "
    "carer who worked, walking through the same 4 questions that drove the weight and "
    "extend_feasibility decisions in those files.",
    "",
    "1. SET ROSTER TODAY -- which of today's visits are on this carer's own fixed weekly "
    "schedule (Weekly-classified, >=75% consistency in her own active weeks).",
    "2. EXTRA VISITS TODAY -- which of today's visits are outside her set roster.",
    "3. DAY SITUATION -- today's company-wide staffing compared to the recent average for "
    f"this weekday: {today_carers} carers worked today vs a {avg_carers_recent:.1f} recent-"
    f"{TARGET_WEEKDAY} average (ratio {staffing_ratio:.2f}) -- "
    f"{'SHORT-STAFFED' if day_is_short_staffed else 'a normal/adequate day, not short-staffed'}. "
    "For each extra visit, whether the client's regular carer was absent that week (a real "
    "coverage gap) or present (an additional/unclear need) is shown.",
    "4. GEOGRAPHIC CHOICES -- her own search radius (median travel + 5km) and how many other "
    "active clients fall within it, i.e. how much real flexibility she has if asked to cover "
    "someone else nearby versus being geographically boxed into a small, isolated patch.",
    "",
    "This explains WHY the JSON files made the decisions they made -- e.g. a carer with a "
    "fully closed roster, few nearby alternatives, and a normal (not short-staffed) day is "
    "the clearest case for weight=2/extend=False; a flexible carer covering a real absence "
    "gap on a tight day with many nearby options is the clearest case for wide extension.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Carer', 'Set Roster Visits Today', 'Extra Visits Today', 'Search Radius (km)',
                    'Nearby Clients', 'Closed or Flexible?', 'Notes']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

def find_slot_pattern_local(carer, wd, client, start_minute):
    candidates = per_carer_client_slot_pattern.get((carer, wd, client), [])
    if not candidates:
        return None, 0.0
    best = min(candidates, key=lambda c: abs(c[2] - start_minute))
    if abs(best[2] - start_minute) > TIME_GAP_MINUTES:
        return None, 0.0
    return best[0], best[1]

# Precompute nearby-client counts per carer (reuse client_info/travel_km already loaded)
client_exact_names = list(client_info.keys())
carer_nearby_count = {}
for carer in carers_today:
    radius = carer_search_radius.get(carer)
    if radius is None:
        continue
    n = sum(1 for cl in client_exact_names if (d := travel_km(carer, cl)) is not None and d <= radius)
    carer_nearby_count[carer] = n

row_i = 2
per_carer_rows = {}
for carer in carers_today:
    my_visits = visits_by_carer[carer]
    roster_visits, extra_visits = [], []
    for v in my_visits:
        start_min = v['start_dt'].hour * 60 + v['start_dt'].minute
        pattern, ratio = find_slot_pattern_local(carer, TARGET_WEEKDAY, v['client'], start_min)
        entry = {'v': v, 'pattern': pattern, 'ratio': ratio}
        if pattern == 'Weekly':
            roster_visits.append(entry)
        else:
            extra_visits.append(entry)
    per_carer_rows[carer] = (roster_visits, extra_visits)

    is_flexible = carer_does_extra.get(carer, False)
    radius = carer_search_radius.get(carer)
    nearby = carer_nearby_count.get(carer)
    note = ''
    if not is_flexible and (nearby or 0) < 5:
        note = 'Closed caseload + few nearby options -- boxed in if her clients need cover'
    elif is_flexible and (nearby or 0) >= 20:
        note = 'Flexible with many nearby options -- a strong candidate for covering others'

    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=len(roster_visits)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=len(extra_visits)).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=round(radius, 1) if radius else '').font = normal_font
    summary_ws.cell(row=row_i, column=5, value=nearby if nearby is not None else '').font = normal_font
    summary_ws.cell(row=row_i, column=6, value='Flexible' if is_flexible else 'Closed').font = normal_font
    summary_ws.cell(row=row_i, column=7, value=note).font = note_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 28
for col in 'BCDEF':
    summary_ws.column_dimensions[col].width = 20
summary_ws.column_dimensions['G'].width = 60
summary_ws.freeze_panes = 'A2'

# --- Per-carer sheets ---
def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

for carer in carers_today:
    ws = wb.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    is_flexible = carer_does_extra.get(carer, False)
    radius = carer_search_radius.get(carer)
    nearby = carer_nearby_count.get(carer)
    ws['A2'] = (f"{'Flexible' if is_flexible else 'Closed'} caseload  |  "
                f"Search radius {round(radius,1) if radius else 'n/a'}km  |  "
                f"{nearby if nearby is not None else 'n/a'} nearby clients")
    ws['A2'].font = note_font
    r = 4

    roster_visits, extra_visits = per_carer_rows[carer]

    ws.cell(row=r, column=1, value='1. SET ROSTER TODAY').font = section_font
    r += 1
    if roster_visits:
        headers = ['Client', 'Time', 'Consistency']
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font; c.fill = header_fill
        r += 1
        for e in sorted(roster_visits, key=lambda e: e['v']['start_dt']):
            v = e['v']
            row_vals = [v['client'], v['start_dt'].strftime('%H:%M'), f"{e['ratio']:.0%}"]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = normal_font; cell.fill = roster_fill
            r += 1
    else:
        ws.cell(row=r, column=1, value='(none -- every visit today is off her set roster)').font = note_font
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='2. EXTRA VISITS TODAY').font = section_font
    r += 1
    if extra_visits:
        headers = ['Client', 'Time', 'Pattern', '3. Likely Reason (day situation)']
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font; c.fill = header_fill
        r += 1
        for e in sorted(extra_visits, key=lambda e: e['v']['start_dt']):
            v = e['v']
            client = v['client']
            wd = TARGET_WEEKDAY
            # who is the client's regular carer for this weekday (if any), and were they
            # active/did-they-visit that week -- same logic as Extra Visits Detail
            regular_slots = [(c2, ratio2) for (c2, wd2, cl2), slots in per_carer_client_slot_pattern.items()
                              if wd2 == wd and cl2 == client and c2 != carer
                              for (pat2, ratio2, med2) in slots if pat2 == 'Weekly']
            if regular_slots:
                names = ', '.join(sorted(set(c2 for c2, _ in regular_slots)))
                reason = f"Client has a regular carer ({names}) -- likely covering absence or an additional need"
            else:
                reason = 'Client has no set roster for this weekday -- ad hoc/occasional client'
            row_vals = [client, v['start_dt'].strftime('%H:%M'), e['pattern'] or 'Occasional/ad hoc', reason]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 4 else normal_font
                cell.fill = extra_fill
            r += 1
    else:
        ws.cell(row=r, column=1, value='(none -- every visit today is on her set roster)').font = note_font
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='3. DAY SITUATION').font = section_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"{today_carers} carers worked today vs recent-{TARGET_WEEKDAY} average "
                   f"{avg_carers_recent:.1f} (ratio {staffing_ratio:.2f}) -- "
                   f"{'SHORT-STAFFED, extension widened company-wide' if day_is_short_staffed else 'normal/adequate staffing, extension kept conservative'}")
            ).font = normal_font
    r += 2

    ws.cell(row=r, column=1, value='4. GEOGRAPHIC CHOICES').font = section_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"Search radius (her median travel + 5km): {round(radius,1) if radius else 'n/a'}km. "
                   f"{nearby if nearby is not None else 'n/a'} active clients fall within that radius -- "
                   f"{'a healthy set of alternatives if she needs to cover someone nearby' if (nearby or 0) >= 15 else 'a limited catchment -- few realistic nearby options'}.")
            ).font = normal_font
    r += 1

    widths = [30, 14, 14, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

out_path = f'./output/Carer_Day_Analysis_{TARGET_DATE.isoformat()}.xlsx'
wb.save(out_path)
print(f"Saved {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
