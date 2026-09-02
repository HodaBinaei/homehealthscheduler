import pickle, json
from collections import defaultdict, Counter
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

import os
PROJECT_ROOT = os.environ.get(
    'HHS_PROJECT_ROOT',
    os.path.dirname(os.path.abspath(__file__)),
)
os.makedirs('./output', exist_ok=True)

with open(f'{PROJECT_ROOT}/roster_data.pkl', 'rb') as f:
    roster = pickle.load(f)  # carer -> weekday -> list of visit dicts

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
TIME_GAP_MINUTES = 90
DATA_START = datetime.date(2025, 6, 30)
DATA_END = datetime.date(2026, 8, 2)

def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()

with open(f'{PROJECT_ROOT}/data/clients-new.json') as f:
    clients_json = json.load(f)['client']

client_dates = {}
for c in clients_json:
    if c.get('status') != 'Active':
        continue
    first = (c.get('name') or '').strip()
    last = (c.get('lastname') or '').strip()
    key = norm(f"{last}, {first}")
    sd = c.get('start_date')
    ed = c.get('end_date') or c.get('termination_date')
    client_dates[key] = (
        datetime.date.fromisoformat(sd) if sd else None,
        datetime.date.fromisoformat(ed) if ed else None,
    )

def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters

def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)

def median_time(dts):
    minutes = sorted(d.hour * 60 + d.minute for d in dts)
    n = len(minutes)
    mid = minutes[n // 2] if n % 2 else round((minutes[n // 2 - 1] + minutes[n // 2]) / 2)
    return f"{mid // 60:02d}:{mid % 60:02d}"

def fmt_minutes(m):
    sign = '+' if m >= 0 else ''
    return f"{sign}{int(round(m))}min"

def actual_variance_note(occurrences):
    """occurrences: list of (req_start, req_end, actual_start, actual_end, carer)."""
    diffs = [(o[2] - o[0]).total_seconds() / 60 for o in occurrences if o[2] is not None]
    if not diffs:
        return ''
    diffs_sorted = sorted(diffs)
    n = len(diffs_sorted)
    median = diffs_sorted[n // 2] if n % 2 else (diffs_sorted[n // 2 - 1] + diffs_sorted[n // 2]) / 2
    lo, hi = diffs_sorted[0], diffs_sorted[-1]
    if abs(median) <= 5 and (hi - lo) <= 20:
        return ''
    return f"actual runs {fmt_minutes(median)} vs scheduled (range {fmt_minutes(lo)} to {fmt_minutes(hi)})"

# --- invert the carer-keyed roster into a client-keyed one ---
client_roster = defaultdict(lambda: defaultdict(list))  # client -> weekday -> list of visit dicts (+carer)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            client_roster[v['client']][wd].append({**v, 'carer': carer})

# --- double-up detection: genuine time-overlap between DIFFERENT carers visiting the
# same client on the same date (not just "same date", and not just "similar time-of-day"
# clustering -- an actual overlapping window, e.g. one carer 09:00-21:00 and another
# 09:00-10:00 helping at the start). Built once, globally, from the raw visit data.
client_day_visits = defaultdict(lambda: defaultdict(list))  # client -> date -> [(carer, start, end)]
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt'] and v['end_dt']:
                client_day_visits[v['client']][v['start_dt'].date()].append((carer, v['start_dt'], v['end_dt']))

def _overlaps(a_start, a_end, b_start, b_end):
    return a_start < b_end and b_start < a_end

def double_up_partners(client, carer, start_dt, end_dt):
    """Returns a sorted list of distinct carer names whose visit to this client on this
    date genuinely overlaps in time with this visit (excluding the same carer)."""
    date = start_dt.date()
    partners = set()
    for other_carer, o_start, o_end in client_day_visits.get(client, {}).get(date, []):
        if other_carer == carer:
            continue
        if _overlaps(start_dt, end_dt, o_start, o_end):
            partners.add(other_carer)
    return sorted(partners)

# Client's own "received a visit this weekday" calendar and "received a visit this day" set --
# the denominators for judging consistency FROM THE CLIENT'S SIDE (regardless of which carer).
client_weekday_active_weeks = defaultdict(set)
client_active_days = defaultdict(set)
for client, wd_map in client_roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                client_weekday_active_weeks[(client, wd)].add(isoweek(v['start_dt'].date()))
                client_active_days[client].add(v['start_dt'].date())

DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5

def classify_client_relative(client, wd, occurrences, daily_fallback=None):
    # occurrences: list of (req_start, req_end, actual_start, actual_end, carer)
    starts = [o[0] for o in occurrences]
    ends = [o[1] for o in occurrences if o[1] is not None]
    dates = sorted(set(o[0].date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]

    client_start, client_end = client_dates.get(norm(client), (None, None))
    window_start, window_end = observed_start, observed_end
    left_bound = max(client_start or DATA_START, DATA_START)
    right_bound = min(client_end or DATA_END, DATA_END)
    left_censored = observed_start <= left_bound + datetime.timedelta(days=14)
    right_censored = observed_end >= right_bound - datetime.timedelta(days=14)

    active_weeks_all = client_weekday_active_weeks[(client, wd)]
    win_start_wk, win_end_wk = isoweek(window_start), isoweek(window_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)

    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)

    fallback_note = ''
    if n_active < 3:
        pattern, ratio = 'Insufficient history', (n_hits / n_active if n_active else None)
        if daily_fallback is not None:
            daily_ratio, daily_active_days = daily_fallback
            ratio = daily_ratio
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
            fallback_note = (f"insufficient per-weekday history -- consistency instead computed "
                              f"from all weekdays pooled ({round(daily_ratio * daily_active_days)}/"
                              f"{daily_active_days} of client's active days)")
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'

    # carer breakdown -- who actually does these visits
    carer_counts = Counter(o[4] for o in occurrences)
    total_visits = len(occurrences)
    n_dates = len(dates)
    visits_per_date = total_visits / n_dates  # >1 means multiple visit records same date/slot
    if len(carer_counts) == 1:
        carer_summary = list(carer_counts.keys())[0]
    else:
        top_carer, top_n = carer_counts.most_common(1)[0]
        if top_n / total_visits >= 0.75:
            carer_summary = f"mostly {top_carer} ({top_n}/{total_visits}), {len(carer_counts)} carers total"
        else:
            carer_summary = f"rotates across {len(carer_counts)} carers (top: {top_carer} {top_n}/{total_visits})"

    # double-up: genuine time-overlap with a DIFFERENT carer's visit on the same date
    double_up_hits = 0
    double_up_partner_names = set()
    for o in occurrences:
        partners = double_up_partners(client, o[4], o[0], o[1])
        if partners:
            double_up_hits += 1
            double_up_partner_names.update(partners)
    double_up_fraction = double_up_hits / total_visits if total_visits else 0

    start_times = sorted(set(d.strftime('%H:%M') for d in starts))
    time_note = f"scheduled time varies {start_times[0]}-{start_times[-1]}" if len(start_times) > 1 else ""
    variance_note = actual_variance_note(occurrences)
    censor_note = ''
    if left_censored and right_censored:
        censor_note = 'spans full eligible window'
    elif left_censored:
        censor_note = 'ongoing from window start'
    elif right_censored:
        censor_note = 'ongoing to window end'
    censor_note = ' / '.join(filter(None, [fallback_note, censor_note]))

    return {
        'pattern': pattern, 'ratio': ratio,
        'start': median_time(starts), 'end': median_time(ends) if ends else '',
        'time_note': time_note, 'variance_note': variance_note,
        'visits': total_visits, 'dates': n_dates, 'visits_per_date': visits_per_date,
        'active_weeks': n_active, 'carer_summary': carer_summary, 'n_carers': len(carer_counts),
        'double_up_fraction': double_up_fraction, 'double_up_partners': sorted(double_up_partner_names),
        'first': observed_start, 'last': observed_end, 'censor_note': censor_note,
    }

records = defaultdict(lambda: defaultdict(list))
pattern_counts = Counter()

for client, wd_map in client_roster.items():
    per_wd_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                per_wd_tagged[wd]  # ensure key exists
    # pool across weekdays for this client (for the daily fallback), tagging weekday
    tagged = []
    for wd, visits in wd_map.items():
        for v in visits:
            if v['start_dt']:
                tagged.append((v['start_dt'], v['end_dt'], v.get('actual_start_dt'),
                                v.get('actual_end_dt'), v['carer'], wd))

    for cluster in cluster_by_time(tagged):
        dates = sorted(set(o[0].date() for o in cluster))
        span_start, span_end = dates[0], dates[-1]
        weekdays_covered = set(WEEKDAYS[d.weekday()] for d in dates)
        active_days = [d for d in client_active_days.get(client, set()) if span_start <= d <= span_end]
        daily_fallback = None
        if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
            daily_fallback = (len(dates) / len(active_days), len(active_days))

        by_weekday = defaultdict(list)
        for o in cluster:
            by_weekday[o[5]].append(o[:5])  # drop weekday tag, keep (req_s,req_e,act_s,act_e,carer)
        for wd, occs in by_weekday.items():
            info = classify_client_relative(client, wd, occs, daily_fallback=daily_fallback)
            pattern_counts[info['pattern']] += 1
            records[client][wd].append(info)

print("Pattern counts (client-relative):", dict(pattern_counts))

wb = openpyxl.Workbook()
wb.remove(wb.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
day_fill = PatternFill('solid', fgColor='D9E1F2')
day_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
pattern_fills = {
    'Weekly': PatternFill('solid', fgColor='E2EFDA'),
    'Fortnightly': PatternFill('solid', fgColor='FCE4D6'),
    'Occasional': PatternFill('solid', fgColor='FFFFFF'),
    'Insufficient history': PatternFill('solid', fgColor='F2F2F2'),
}
double_up_fill = PatternFill('solid', fgColor='FFE699')  # visits_per_date > 1.05

def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()

# --- Read Me ---
notes_ws = wb.create_sheet('Read Me')
notes = [
    "Client weekday visit volume (across ALL carers)",
    "",
    "This is the client-facing counterpart to the carer-facing roster workbook: instead of asking "
    "'how consistently does this carer see this client', it asks 'how many visits does this client "
    "get in this slot per week, and who provides them' -- pooling across every carer who has ever "
    "covered that slot.",
    "",
    "Same methodology as the carer workbook, mirrored onto the client: consistency is judged "
    "against the CLIENT's own active weeks (weeks they received any visit at all on that weekday, "
    "from any carer), within the slot's own observed span. The same daily-pooled fallback applies "
    "when a single weekday alone has too little history but the client is seen on 4+ other "
    "weekdays too.",
    "",
    "Visits/Date column: 1.0 means exactly one visit record happens per occurrence of this "
    "time-of-day slot (the normal case). Values above 1.0 mean multiple visit records were "
    "logged for the same date/slot -- worth a look, but NOT the same thing as a genuine "
    "double-up (see below); it can also happen from data duplication or closely-spaced "
    "separate visits that happened to cluster together by start time.",
    "",
    "Double-Up / Double-Up Partner(s) columns: this is the precise measure -- computed by "
    "checking, for every actual visit in this slot, whether ANOTHER carer had a visit to the "
    "SAME client that genuinely OVERLAPS in time that day (e.g. one carer 09:00-21:00 and "
    "another 09:00-10:00 helping at the start). Double-Up shows what fraction of this slot's "
    "visits had a real overlapping partner; Double-Up Partner(s) names who. Rows with any "
    "double-up are highlighted amber. This catches cases a simple 'visits per date' count "
    "would miss or over-count -- e.g. a long visit with a short helper visit nested inside it "
    "the same day.",
    "",
    "Carer(s) column shows who actually performs the visits in this slot: a single name if it's "
    "always the same carer, 'mostly X' if one carer dominates but others have covered it, or "
    "'rotates across N carers' if no one carer is dominant.",
    "",
    f"Slot pattern totals across all clients: {dict(pattern_counts)}",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

# --- Summary ---
summary_ws = wb.create_sheet('Summary')
summary_headers = ['Client', 'Days Visited', 'Total Weekly Slots', 'Distinct Carers Involved']
for c, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_i = 2
for client in sorted(records.keys()):
    days_present = [wd for wd in WEEKDAYS if wd in records[client]]
    all_entries = [e for wd in records[client] for e in records[client][wd]]
    n_slots = len(all_entries)
    all_carers = set()
    for e in all_entries:
        cs = e['carer_summary']
    # recompute distinct carers count properly via n_carers field, take max as a rough signal
    n_carers_total = len(set().union(*[set([1]) for _ in all_entries])) if all_entries else 0
    summary_ws.cell(row=row_i, column=1, value=client).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=', '.join(days_present)).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=n_slots).font = normal_font
    summary_ws.cell(row=row_i, column=4, value=max((e['n_carers'] for e in all_entries), default=0)).font = normal_font
    row_i += 1

summary_ws.column_dimensions['A'].width = 30
summary_ws.column_dimensions['B'].width = 55
summary_ws.column_dimensions['C'].width = 20
summary_ws.column_dimensions['D'].width = 24
summary_ws.freeze_panes = 'A2'

# --- Per-client sheets ---
headers = ['Weekday', 'Scheduled Start', 'Scheduled End', 'Visits/Date', 'Total Visits', 'Dates',
           'Pattern', 'Consistency', 'Double-Up', 'Double-Up Partner(s)', 'Carer(s)', 'Date Range', 'Notes']

for client in sorted(records.keys()):
    ws = wb.create_sheet(sheet_name(client, used_names))
    ws['A1'] = client
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    ws.merge_cells('A1:M1')
    r = 3
    for wd in WEEKDAYS:
        entries = records[client].get(wd, [])
        if not entries:
            continue
        ws.cell(row=r, column=1, value=wd).font = day_font
        for col in range(1, 14):
            ws.cell(row=r, column=col).fill = day_fill
        r += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        r += 1
        entries = sorted(entries, key=lambda e: e['start'])
        for e in entries:
            date_range = (e['first'].strftime('%d/%m/%Y') if e['first'] == e['last']
                          else f"{e['first'].strftime('%d/%m/%Y')} - {e['last'].strftime('%d/%m/%Y')}")
            note = ' / '.join(filter(None, [e['time_note'], e['variance_note'], e['censor_note']]))
            double_up_display = f"{e['double_up_fraction']:.0%}" if e['double_up_fraction'] > 0 else ''
            partners_display = ', '.join(e['double_up_partners'])
            row_vals = [wd, e['start'], e['end'], round(e['visits_per_date'], 2), e['visits'], e['dates'],
                        e['pattern'], f"{e['ratio']:.0%}" if e['ratio'] is not None else '',
                        double_up_display, partners_display,
                        e['carer_summary'], date_range, note]
            is_double_up = e['double_up_fraction'] > 0
            fill = double_up_fill if is_double_up else pattern_fills.get(e['pattern'])
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = note_font if col == 13 else normal_font
                if fill:
                    cell.fill = fill
            r += 1
        r += 1

    widths = [12, 13, 13, 11, 11, 8, 18, 12, 10, 34, 34, 24, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

out_path = f'{PROJECT_ROOT}/output/Client_Weekday_Visits.xlsx'
wb.save(out_path)
print("Saved", out_path)
print("Sheets:", len(wb.sheetnames))
