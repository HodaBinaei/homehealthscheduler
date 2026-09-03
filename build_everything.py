"""
Build patient/caregivers/feasibility for a REAL day export (patients.json / caregivers.json
from your own system), but treating ONLY identity (pid/cid/prid/crid, gender, location) and
the raw need/availability time (patient's request_window start/end, caregiver's shift
start/end) as trusted input. Every other field -- duration, min_duration, soft/hard windows,
priorities, violation levels, extend_feasibility, caregiver_usage_priority -- is DISCARDED
from your export and recomputed from history, exactly the same way run_all_in_one.py builds
these fields from today_patients.csv / today_carers.csv. This is the correct behaviour: your
export's own analysis fields are not trusted here, only who/when.

Reuses the full historical foundation from run_all_in_one.py unchanged (roster building with
DISLIKES stripped, real distances, cancellation analysis, set-roster classification,
extend_feasibility classifier) -- none of that depends on which specific day you're
scheduling.
"""
import csv, sys, json, math, re, datetime, os
from collections import defaultdict, Counter

csv.field_size_limit(sys.maxsize)

# =============================================================================
# CONFIG
# =============================================================================
PROJECT_ROOT = '/home/hoda/Desktop/homehealthscheduler_local_pipeline'
CSV_PATH = f'{PROJECT_ROOT}/data/VisitExport.csv'
USERS_PATH = f'{PROJECT_ROOT}/data/users-new.json'
CLIENTS_PATH = f'{PROJECT_ROOT}/data/clients-new.json'
DRIVING_DATA_PATH = f'{PROJECT_ROOT}/data/driving_data.json'
DAY_EXPORT_DIR = f'{PROJECT_ROOT}/data_today'
DAY_PATIENTS_FILENAME = 'patients.json'      # <- exact filename in DAY_EXPORT_DIR, edit if yours differs
DAY_CAREGIVERS_FILENAME = 'caregivers.json'  # <- exact filename in DAY_EXPORT_DIR, edit if yours differs
OUTPUT_DIR = f'{PROJECT_ROOT}/output'
HHS_SCHEMA_PATH = PROJECT_ROOT

import os as _os
for _fname in (DAY_PATIENTS_FILENAME, DAY_CAREGIVERS_FILENAME):
    _fpath = f'{DAY_EXPORT_DIR}/{_fname}'
    if not _os.path.isfile(_fpath):
        _actual = _os.listdir(DAY_EXPORT_DIR) if _os.path.isdir(DAY_EXPORT_DIR) else []
        raise SystemExit(
            f"Expected '{_fpath}' but it doesn't exist. Files actually in {DAY_EXPORT_DIR}: "
            f"{_actual}. Update DAY_PATIENTS_FILENAME / DAY_CAREGIVERS_FILENAME near the top "
            f"of this script to match your real filenames."
        )

with open(f'{DAY_EXPORT_DIR}/{DAY_PATIENTS_FILENAME}') as f:
    _real_patients_raw = json.load(f)
with open(f'{DAY_EXPORT_DIR}/{DAY_CAREGIVERS_FILENAME}') as f:
    _real_caregivers_raw = json.load(f)

# Sanity check: your REAL day export is {"date": "...", "patients": [...]}. A plain list
# here almost always means this is actually a COMPUTED OUTPUT file (patient.json /
# patient_YYYY-MM-DD.json from an earlier run of this pipeline) that ended up in
# data_today/ by mistake instead of your original export -- fail clearly rather than
# crash on a cryptic TypeError three lines later.
if not isinstance(_real_patients_raw, dict) or 'date' not in _real_patients_raw:
    raise SystemExit(
        f"'{DAY_EXPORT_DIR}/{DAY_PATIENTS_FILENAME}' doesn't look like a real day export "
        f"(expected a {{'date': ..., 'patients': [...]}} object, got a "
        f"{'list' if isinstance(_real_patients_raw, list) else type(_real_patients_raw).__name__}). "
        f"This is very likely a previously COMPUTED output file (e.g. patient.json from an "
        f"earlier run) placed here by mistake instead of your original export -- check "
        f"{DAY_EXPORT_DIR}/{DAY_PATIENTS_FILENAME} and replace it with the real export."
    )

if not isinstance(_real_caregivers_raw, dict) or 'caregivers' not in _real_caregivers_raw:
    raise SystemExit(
        f"'{DAY_EXPORT_DIR}/{DAY_CAREGIVERS_FILENAME}' doesn't look like a real day export "
        f"(expected a {{'date': ..., 'caregivers': [...]}} object). Same likely cause as "
        f"above -- check this file is your original export, not a computed output file."
    )

TARGET_DATE = datetime.date.fromisoformat(_real_patients_raw['date'])
TARGET_WEEKDAY = TARGET_DATE.strftime('%A')

DISLIKES = [
    ('Fiona Buchannon (DS)', 'Bridget Madden'),
    ('Kathleen Nee', 'Mary Scullion'),
    ('Lourdes Conneely', 'Mary Scullion'),
    ('Geraldine King (DS)', 'Mary Scullion'),
    ('Bernie McEvaddy', 'Mary Scullion'),
    ('Mary Jo Flynn', 'Mary Scullion'),
    ("John O'Toole", 'Mary Scullion'),
]

# Manual extend_feasibility overrides -- for carers whose classifier result sits right on the
# threshold (isolated-override or off-routine-rate borderline) where the person managing this
# has direct knowledge the automatic result doesn't reflect. This does NOT change the
# classifier logic itself for anyone else -- it's applied as a final, targeted override after
# the normal computation, for exactly the carers listed here.
EXTEND_FEASIBILITY_OVERRIDES = {
    'Bernie Egan': False,
    'Heba El-Maleh': True,
}

WEEKDAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
TIME_GAP_MINUTES = 90
DAILY_FALLBACK_MIN_ACTIVE_DAYS = 5
MAX_SCHEMA_MINUTE = (23 * 60 + 59) * 2
SCHEMA_MAX_DURATION = 8 * 60

import os
os.makedirs(OUTPUT_DIR, exist_ok=True)


def norm(s):
    if s is None:
        return ''
    return ' '.join(s.strip().split()).lower()


def strip_loc_suffix(name):
    return re.sub(r'\s*\([^)]*\)', '', name or '')


def percentile(values, p):
    if not values:
        return None
    s = sorted(values)
    k = (len(s) - 1) * (p / 100)
    f, c = int(k), min(int(k) + 1, len(s) - 1)
    if f == c:
        return s[f]
    return s[f] + (s[c] - s[f]) * (k - f)


def gender_enum(g):
    g = (g or '').lower()
    return g if g in ('male', 'female') else 'prefer_not_to_say'


print("=" * 70)
print(f"Building historical foundation for {TARGET_DATE} ({TARGET_WEEKDAY})")
print("=" * 70)

# ---- roster (Service Requirement anchored) ----
with open(USERS_PATH) as f:
    users = json.load(f)['user']
active_carers = {}  # despite the name, this maps ALL known carers (any current status) to
# their display name, for matching historical CSV rows -- a visit that already happened is
# still real regardless of whether the carer is still employed today. Current employment
# status only matters for carer_info (who can be assigned TODAY), built separately below.
carer_termination_date = {}  # display name -> date.date, if terminationDate is set on record
for u in users:
    if u.get('is_caregiver'):
        first = (u.get('name') or '').strip()
        last = (u.get('lastname') or '').strip()
        active_carers[norm(f"{last}, {first}")] = f"{first} {last}".strip()
        _term = u.get('terminationDate')
        if _term:
            try:
                carer_termination_date[f"{first} {last}".strip()] = datetime.date.fromisoformat(_term[:10])
            except (ValueError, TypeError):
                pass

with open(CLIENTS_PATH) as f:
    clients_json = json.load(f)['client']
active_clients = {}
for c in clients_json:
    if c.get('status') == 'Active':
        first = (c.get('name') or '').strip()
        last = (c.get('lastname') or '').strip()
        active_clients[norm(f"{last}, {first}")] = f"{first} {last}".strip()

roster = defaultdict(lambda: defaultdict(list))
carer_dates = defaultdict(set)
kept = 0
with open(CSV_PATH, newline='', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for row in reader:
        if (row.get('Actual Service Type Description') or '').strip() != 'Personal Care':
            continue
        if (row.get('Cancellation Description') or '').strip():
            continue
        emp_name = (row.get('Actual Employee Name') or '').strip()
        loc_name = (row.get('Service Location Name') or '').strip()
        if not emp_name or not loc_name:
            continue
        emp_key = norm(emp_name)
        if emp_key not in active_carers:
            continue
        try:
            req_start_dt = datetime.datetime.strptime(
                (row.get('Service Requirement Start Date And Time') or '').strip(), '%d/%m/%Y %H:%M:%S')
            req_end_dt = datetime.datetime.strptime(
                (row.get('Service Requirement End Date And Time') or '').strip(), '%d/%m/%Y %H:%M:%S')
        except Exception:
            continue
        carer_dates[active_carers[emp_key]].add(req_start_dt.date())
        client_display = active_clients.get(norm(loc_name)) or active_clients.get(norm(strip_loc_suffix(loc_name)))
        if not client_display:
            continue
        try:
            actual_start_dt = datetime.datetime.strptime(
                (row.get('Actual Start Date And Time') or '').strip(), '%d/%m/%Y %H:%M:%S')
            actual_end_dt = datetime.datetime.strptime(
                (row.get('Actual End Date And Time') or '').strip(), '%d/%m/%Y %H:%M:%S')
        except Exception:
            actual_start_dt = actual_end_dt = None
        weekday = WEEKDAYS[req_start_dt.weekday()]
        roster[active_carers[emp_key]][weekday].append({
            'client': client_display, 'start_dt': req_start_dt, 'end_dt': req_end_dt,
            'actual_start_dt': actual_start_dt, 'actual_end_dt': actual_end_dt,
        })
        kept += 1

carer_presence = {c: (min(d), max(d)) for c, d in carer_dates.items()}
print(f"Roster: {kept} historical visits kept, {len(carer_presence)} carer presence windows")

MERGE_GAP_TOLERANCE_MIN = 15


def merge_split_visits(roster):
    merged_count = 0
    for carer, wd_map in roster.items():
        for wd, visits in wd_map.items():
            by_date = defaultdict(list)
            for v in visits:
                by_date[v['start_dt'].date()].append(v)
            windows_single, windows_multi = set(), set()
            for date, day_visits in by_date.items():
                clients_today = defaultdict(list)
                for v in day_visits:
                    clients_today[v['client']].append(v)
                for client, cvisits in clients_today.items():
                    cvisits.sort(key=lambda v: v['start_dt'])
                    if len(cvisits) == 1:
                        windows_single.add((client, (cvisits[0]['start_dt'].strftime('%H:%M'),
                                                      cvisits[0]['end_dt'].strftime('%H:%M'))))
                    else:
                        contiguous = all(
                            abs((cvisits[i + 1]['start_dt'] - cvisits[i]['end_dt']).total_seconds() / 60) <= MERGE_GAP_TOLERANCE_MIN
                            for i in range(len(cvisits) - 1))
                        if contiguous:
                            windows_multi.add((client, (cvisits[0]['start_dt'].strftime('%H:%M'),
                                                         cvisits[-1]['end_dt'].strftime('%H:%M'))))
            eligible = windows_single & windows_multi
            if not eligible:
                continue
            new_visits = []
            for date, day_visits in by_date.items():
                clients_today = defaultdict(list)
                for v in day_visits:
                    clients_today[v['client']].append(v)
                for client, cvisits in clients_today.items():
                    cvisits.sort(key=lambda v: v['start_dt'])
                    if len(cvisits) > 1:
                        window = (cvisits[0]['start_dt'].strftime('%H:%M'), cvisits[-1]['end_dt'].strftime('%H:%M'))
                        if (client, window) in eligible:
                            first, last = cvisits[0], cvisits[-1]
                            new_visits.append({
                                'client': client, 'start_dt': first['start_dt'], 'end_dt': last['end_dt'],
                                'actual_start_dt': first['actual_start_dt'], 'actual_end_dt': last['actual_end_dt'],
                            })
                            merged_count += 1
                            continue
                    new_visits.extend(cvisits)
            wd_map[wd] = new_visits
    return merged_count


merged = merge_split_visits(roster)
print(f"Split-visit merge: {merged}")

_dislike_removed = 0
for dislike_client, dislike_carer in DISLIKES:
    wd_map = roster.get(dislike_carer)
    if not wd_map:
        continue
    for wd, visits in wd_map.items():
        before = len(visits)
        wd_map[wd] = [v for v in visits if v['client'] != dislike_client]
        _dislike_removed += before - len(wd_map[wd])
if _dislike_removed:
    print(f"Removed {_dislike_removed} historical visit(s) for DISLIKES pairs")

# ---- distances ----
carer_id_by_name = {}
for u in users:
    if u.get('status') == 'Active' and u.get('is_caregiver'):
        first = (u.get('name') or '').strip()
        last = (u.get('lastname') or '').strip()
        carer_id_by_name[f"{first} {last}".strip()] = str(u.get('id'))
client_id_by_name = {}
for c in clients_json:
    if c.get('status') == 'Active':
        first = (c.get('name') or '').strip()
        last = (c.get('lastname') or '').strip()
        client_id_by_name[f"{first} {last}".strip()] = str(c.get('id'))

with open(DRIVING_DATA_PATH) as f:
    _driving = json.load(f)
_dist = _driving['distance']
_dur = _driving.get('duration', {})
carer_client_km = {}
carer_client_min = {}
for carer_name, cid in carer_id_by_name.items():
    for client_name, did in client_id_by_name.items():
        key1, key2 = f"{cid}_{did}", f"{did}_{cid}"
        d = _dist.get(key1) or _dist.get(key2)
        t = _dur.get(key1) or _dur.get(key2)
        if d is not None:
            carer_client_km[(carer_name, client_name)] = float(d)
        if t is not None:
            carer_client_min[(carer_name, client_name)] = float(t)
del _driving, _dist, _dur
print(f"Matched {len(carer_client_km)} carer-client distance pairs, {len(carer_client_min)} travel-time pairs")

carer_info = {}
for u in users:
    if u.get('status') == 'Active' and u.get('is_caregiver'):
        first = (u.get('name') or '').strip()
        last = (u.get('lastname') or '').strip()
        display = f"{first} {last}".strip()
        # a passed terminationDate overrides a stale 'Active' status -- she's genuinely gone
        # as of today even if status hasn't been updated to reflect it yet.
        _term = carer_termination_date.get(display)
        if _term and _term <= TARGET_DATE:
            continue
        carer_info[display] = {
            'id': str(u.get('id')), 'gender': (u.get('gender') or 'prefer_not_to_say').lower(),
            'lat': u.get('latitude'), 'lon': u.get('longitude'), 'postcode': u.get('postcode'),
            'travel_method': (u.get('travel_method') or 'Car'),
            'first_name': first, 'last_name': last,
        }
client_info = {}
for c in clients_json:
    if c.get('status') == 'Active':
        first = (c.get('name') or '').strip()
        last = (c.get('lastname') or '').strip()
        display = f"{first} {last}".strip()
        client_info[display] = {
            'id': str(c.get('id')), 'gender': (c.get('gender') or 'prefer_not_to_say').lower(),
            'lat': c.get('latitude'), 'lon': c.get('longitude'), 'postcode': c.get('postcode'),
            'service_priority': c.get('service_priority') or 'Medium',
            'first_name': first, 'last_name': last,
        }


def haversine_km(p1, p2):
    lat1, lon1 = p1
    lat2, lon2 = p2
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi, dlambda = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def travel_km(carer, client):
    d = carer_client_km.get((carer, client))
    if d is not None:
        return d
    ci, cj = carer_info.get(carer), client_info.get(client)
    if ci and cj and ci['lat'] and cj['lat']:
        return haversine_km((ci['lat'], ci['lon']), (cj['lat'], cj['lon']))
    return None


# ---- cancellation analysis ----
OPERATIONAL_CANCEL_REASONS = {
    'VNR', 'Missed call', 'Missed Call', 'Cancelled Less than 12h',
    'Cancelled with less than 24 hours notice', 'Covered  By Another Agency',
}
EXCUSED_REASONS = {
    'Hospital', 'Holiday', 'Bank Holiday', 'Respite',
    'ZzzCoronavirus – Financial Reasons', 'ZzzCoronavirus – Hospitalised', 'ZzzCoronavirus – Shielding',
}
PLACEHOLDER_PREFIX = '*'

slot_cancel_data = defaultdict(lambda: {'fulfilled': 0, 'op_cancelled': 0, 'excused': 0, 'op_cancel_planned': Counter()})
with open(CSV_PATH, newline='', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for row in reader:
        if (row.get('Actual Service Type Description') or '') != 'Personal Care':
            continue
        loc_raw = (row.get('Service Location Name') or '').strip()
        if not loc_raw:
            continue
        client_display = active_clients.get(norm(loc_raw)) or active_clients.get(norm(strip_loc_suffix(loc_raw)))
        if not client_display:
            continue
        req_start = (row.get('Service Requirement Start Date And Time') or '').strip()
        try:
            req_dt = datetime.datetime.strptime(req_start, '%d/%m/%Y %H:%M:%S')
        except Exception:
            continue
        wd = WEEKDAYS[req_dt.weekday()]
        half_hour = (req_dt.hour * 60 + req_dt.minute) // 30 * 30
        rec = slot_cancel_data[(client_display, wd, half_hour)]
        cancel_desc = (row.get('Cancellation Description') or '').strip()
        emp = (row.get('Actual Employee Name') or '').strip()
        if cancel_desc in EXCUSED_REASONS:
            rec['excused'] += 1
        elif cancel_desc in OPERATIONAL_CANCEL_REASONS or emp.startswith(PLACEHOLDER_PREFIX):
            rec['op_cancelled'] += 1
        elif not cancel_desc and norm(emp) in active_carers:
            rec['fulfilled'] += 1

cancellation_lookup = {}
for (client, wd, half_hour), rec in slot_cancel_data.items():
    denom = rec['fulfilled'] + rec['op_cancelled']
    if denom < 3:
        continue
    rate = rec['op_cancelled'] / denom
    if rate >= 0.90:
        classification = 'Always cancelled'
    elif rate <= 0.05:
        classification = 'Never cancelled'
    else:
        classification = 'Occasionally cancelled'
    cancellation_lookup[(client, wd, half_hour)] = {'classification': classification, 'cancellation_rate': rate}


def find_cancellation_record(client, wd, start_minute):
    bucket = (start_minute // 30) * 30
    for cand in (bucket, bucket - 30, bucket + 30):
        rec = cancellation_lookup.get((client, wd, cand))
        if rec:
            return rec
    return None


print(f"Classified {len(cancellation_lookup)} (client, weekday, slot) cancellation combinations")

# ---- carer-relative foundation ----
carer_weekday_active_weeks = defaultdict(set)
carer_active_days = defaultdict(set)


def isoweek(d):
    y, w, _ = d.isocalendar()
    return (y, w)


for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            carer_weekday_active_weeks[(carer, wd)].add(isoweek(v['start_dt'].date()))
            carer_active_days[carer].add(v['start_dt'].date())


def cluster_by_time(occ):
    occ_sorted = sorted(occ, key=lambda o: o[0].hour * 60 + o[0].minute)
    clusters, cur = [], [occ_sorted[0]]
    for prev, c in zip(occ_sorted, occ_sorted[1:]):
        if (c[0].hour * 60 + c[0].minute) - (prev[0].hour * 60 + prev[0].minute) > TIME_GAP_MINUTES:
            clusters.append(cur); cur = [c]
        else:
            cur.append(c)
    clusters.append(cur)
    return clusters


def classify_relative(carer, wd, occurrences, daily_fallback=None):
    dates = sorted(set(o.date() for o in occurrences))
    observed_start, observed_end = dates[0], dates[-1]
    active_weeks_all = carer_weekday_active_weeks[(carer, wd)]
    win_start_wk, win_end_wk = isoweek(observed_start), isoweek(observed_end)
    active_weeks_in_window = sorted(w for w in active_weeks_all if win_start_wk <= w <= win_end_wk)
    hit_weeks = set(isoweek(d) for d in dates)
    n_active = len(active_weeks_in_window)
    n_hits = len(set(active_weeks_in_window) & hit_weeks)
    if n_active < 3:
        pattern = 'Insufficient history'
        if daily_fallback is not None:
            daily_ratio, _ = daily_fallback
            pattern = 'Weekly' if daily_ratio >= 0.75 else 'Occasional'
    else:
        ratio = n_hits / n_active
        if ratio >= 0.75:
            pattern = 'Weekly'
        elif ratio >= 0.35:
            idx = {w: i for i, w in enumerate(active_weeks_in_window)}
            hit_positions = sorted(idx[w] for w in active_weeks_in_window if w in hit_weeks)
            if len(hit_positions) >= 2:
                steps = [b - a for a, b in zip(hit_positions, hit_positions[1:])]
                alt_like = sum(1 for s in steps if s == 2) / len(steps)
            else:
                alt_like = 0
            pattern = 'Fortnightly' if alt_like >= 0.5 else 'Occasional'
        else:
            pattern = 'Occasional'
    return pattern


set_roster = defaultdict(set)
per_carer_client_slot_pattern = defaultdict(list)
for carer, wd_map in roster.items():
    per_client_tagged = defaultdict(list)
    for wd, visits in wd_map.items():
        for v in visits:
            per_client_tagged[v['client']].append((v['start_dt'], wd))
    for client, tagged in per_client_tagged.items():
        for cluster in cluster_by_time([(dt,) for dt, wd in tagged]):
            cluster_set = set(cluster)
            members = [(dt, wd) for dt, wd in tagged if (dt,) in cluster_set]
            dates_all = [dt for dt, wd in members]
            span_start, span_end = min(d.date() for d in dates_all), max(d.date() for d in dates_all)
            weekdays_covered = set(wd for _, wd in members)
            active_days = [d for d in carer_active_days.get(carer, set()) if span_start <= d <= span_end]
            daily_fallback = None
            if len(weekdays_covered) >= 4 and len(active_days) >= DAILY_FALLBACK_MIN_ACTIVE_DAYS:
                n_dates = len(set(d.date() for d in dates_all))
                daily_fallback = (n_dates / len(active_days), len(active_days))
            by_wd = defaultdict(list)
            for dt, wd in members:
                by_wd[wd].append(dt)
            for wd, dts in by_wd.items():
                pattern = classify_relative(carer, wd, dts, daily_fallback=daily_fallback)
                median_minute = sorted(d.hour * 60 + d.minute for d in dts)[len(dts) // 2]
                per_carer_client_slot_pattern[(carer, wd, client)].append((pattern, median_minute))
                if pattern == 'Weekly':
                    set_roster[(carer, wd)].add(client)

print(f"Set roster slots: {sum(len(v) for v in set_roster.values())}")

carer_does_extra = defaultdict(bool)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        roster_clients = set_roster.get((carer, wd), set())
        for v in visits:
            if v['client'] not in roster_clients:
                carer_does_extra[carer] = True

carer_all_visited_clients = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            carer_all_visited_clients[carer].add(v['client'])

carer_search_radius = {}
for carer, cl in carer_all_visited_clients.items():
    dists = sorted(d for d in (travel_km(carer, c) for c in cl) if d is not None)
    if not dists:
        continue
    n = len(dists)
    median = (dists[n // 2 - 1] + dists[n // 2]) / 2 if n % 2 == 0 else dists[n // 2]
    carer_search_radius[carer] = median + 5

_breadths = sorted(len(cl) for cl in carer_all_visited_clients.values())
MEDIAN_BREADTH = _breadths[len(_breadths) // 2] if _breadths else 1


def concentration_factor(carer):
    breadth = max(len(carer_all_visited_clients.get(carer, set())), 1)
    factor = 1 + 0.15 * math.log2(MEDIAN_BREADTH / breadth)
    return max(0.7, min(factor, 1.3))


client_slot_history = defaultdict(list)
client_slot_actual_times = defaultdict(list)
client_slot_durations = defaultdict(list)  # (client, weekday, idx) -> [(req_duration_min, actual_duration_min)]
client_visits_by_wd = defaultdict(lambda: defaultdict(list))
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            client_visits_by_wd[v['client']][wd].append(
                (carer, v['start_dt'], v['end_dt'], v['actual_start_dt'], v['actual_end_dt']))
# Visits are grouped into recurring slots using Service Requirement time (stable, intended
# schedule); every per-carer STATISTIC (last visit date, recency, on-time variance, duration
# compression) uses the ACTUAL date/time instead -- a visit can genuinely land on a different
# calendar day than its requirement, so anchoring stats on requirement date alone risks a
# slightly wrong reference point.
for client, wd_map in client_visits_by_wd.items():
    for wd, tagged in wd_map.items():
        occ = [(dt,) for carer, dt, e_dt, a_s, a_e in tagged]
        for idx, cluster_occ in enumerate(cluster_by_time(occ)):
            cluster_times = set(cluster_occ)
            members = [(carer, dt, a_s) for carer, dt, e_dt, a_s, a_e in tagged if (dt,) in cluster_times]
            actual_pairs = [(a_s, a_e) for carer, dt, e_dt, a_s, a_e in tagged
                             if (dt,) in cluster_times and a_s is not None and a_e is not None]
            duration_pairs = [((e_dt - dt).total_seconds() / 60, (a_e - a_s).total_seconds() / 60)
                               for carer, dt, e_dt, a_s, a_e in tagged
                               if (dt,) in cluster_times and a_s is not None and a_e is not None]
            client_slot_history[(client, wd)].append(members)
            client_slot_actual_times[(client, wd, idx)] = actual_pairs
            client_slot_durations[(client, wd, idx)] = duration_pairs

print(f"Carer breadth median: {MEDIAN_BREADTH}, carers with search radius: {len(carer_search_radius)}")

# ---- extend_feasibility classifier ----
# extend = True for: (a) new carers (not enough tenure yet to judge), (b) carers who take on
# new/off-routine patients at a rate at or above the peer median (compared against the
# population of other carers, not a fixed number -- there is no "right" absolute count),
# (c) carers who are geographically isolated (very few other active clients near her at all,
# so restricting her wouldn't give the solver any real alternative regardless of her own
# pattern). "Routine" = any client where she has a Weekly-classified slot (see weight=2
# below) -- the SAME definition drives both the weight=2 rule and this classifier,
# deliberately, since they're the same underlying concept (her fixed clients vs everyone else).
NEW_CARER_TENURE_DAYS = 90
CASELOAD_WINDOW_DAYS = 182  # the "per 6 months" period the off-routine rate is normalized to

carer_routine_clients = defaultdict(set)
for (carer, wd, client), slots in per_carer_client_slot_pattern.items():
    for pattern, median_minute in slots:
        if pattern == 'Weekly':
            carer_routine_clients[carer].add(client)

carer_off_routine_rate = {}
for carer, all_clients in carer_all_visited_clients.items():
    off_routine = all_clients - carer_routine_clients.get(carer, set())
    presence = carer_presence.get(carer)
    tenure_days = (presence[1] - presence[0]).days if presence else 0
    if tenure_days < NEW_CARER_TENURE_DAYS:
        carer_off_routine_rate[carer] = None  # too new to judge -- handled as extend=True below
    else:
        carer_off_routine_rate[carer] = len(off_routine) / max(tenure_days / CASELOAD_WINDOW_DAYS, 0.1)

_established_rates = sorted(r for r in carer_off_routine_rate.values() if r is not None)
MEDIAN_OFF_ROUTINE_RATE = _established_rates[len(_established_rates) // 2] if _established_rates else 0.0
print(f"Median off-routine (new-patient) rate across established carers: "
      f"{MEDIAN_OFF_ROUTINE_RATE:.2f} per {CASELOAD_WINDOW_DAYS}-day period")

# "Nobody around her" means genuinely nowhere left to expand -- NOT just a small raw count
# of nearby clients. A carer with only 16 reachable clients but who's visited just 1 of them
# had 15 real, untapped opportunities -- that's a closed pattern, not isolation. True
# isolation needs BOTH: (a) few total clients reachable to her at all (bottom quartile
# across carers), AND (b) she's already covering most of that small pool (so there's
# realistically nothing left to reach). Either alone is not enough.
carer_nearby_total_count = {}
carer_nearby_coverage_ratio = {}
for carer in carer_all_visited_clients:
    radius = carer_search_radius.get(carer)
    if radius is None:
        continue
    visited = carer_all_visited_clients.get(carer, set())
    nearby = [cl for cl in client_info if (d := travel_km(carer, cl)) is not None and d <= radius]
    total = len(nearby)
    covered = sum(1 for cl in nearby if cl in visited)
    carer_nearby_total_count[carer] = total
    carer_nearby_coverage_ratio[carer] = (covered / total) if total > 0 else 0.0

_nearby_totals = sorted(carer_nearby_total_count.values())
# 5th percentile, not bottom quartile -- the quartile (58 clients) was too lenient: it let
# carers with a genuinely meaningful number of real nearby alternatives (e.g. 26, with real
# spread across them) qualify as "isolated", masking a real low engagement pattern. This is
# meant to catch carers with almost nothing around them at all (like 1-10 nearby clients),
# not anyone below-average.
ISOLATED_TOTAL_THRESHOLD = _nearby_totals[len(_nearby_totals) // 20] if _nearby_totals else 0  # 5th percentile
ISOLATED_COVERAGE_THRESHOLD = 0.5  # must already cover at least half of her small nearby pool
print(f"Isolated-carer thresholds: <= {ISOLATED_TOTAL_THRESHOLD} total nearby clients "
      f"AND >= {ISOLATED_COVERAGE_THRESHOLD:.0%} of them already visited")

# A carer with an extremely broad overall caseload (top quartile of breadth across all
# carers) getting the FIXED weight=1.0 "routine, dedicated" tier for many different clients
# at once doesn't reflect genuine dedication -- it just means she shows up often enough
# within her own active weeks for each one, which the formula already correctly penalizes
# via concentration_factor. This only affects the weight=1.0 tier (Weekly but not
# exclusive) -- genuine exclusivity (weight=2.0: whole-day-consuming or fully exclusive) is
# a separate, stronger, per-slot signal that stays valid regardless of her overall breadth.
_breadths = sorted(len(v) for v in carer_all_visited_clients.values())
BREADTH_DISQUALIFY_THRESHOLD = _breadths[3 * len(_breadths) // 4] if _breadths else 0  # 75th percentile
print(f"Breadth disqualification threshold for the fixed weight=1.0 tier: "
      f">= {BREADTH_DISQUALIFY_THRESHOLD} distinct clients ever")

confirmed_picky = set()  # carers whose own routine-based rate lands them closed, for reporting
for carer, rate in carer_off_routine_rate.items():
    if rate is not None and rate < MEDIAN_OFF_ROUTINE_RATE:
        confirmed_picky.add(carer)


def carer_extend_feasibility_flag(carer):
    if (carer_nearby_total_count.get(carer, 0) <= ISOLATED_TOTAL_THRESHOLD
            and carer_nearby_coverage_ratio.get(carer, 0) >= ISOLATED_COVERAGE_THRESHOLD):
        return True  # nobody around her to be selective about
    rate = carer_off_routine_rate.get(carer)
    if rate is None:
        return True  # new carer -- not enough tenure to judge
    return rate >= MEDIAN_OFF_ROUTINE_RATE


customer_totals = Counter()
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            customer_totals[v['client']] += 1

# ---------------------------------------------------------------------------
# Carer HISTORY analysis -- the full VisitExport.csv analysis per carer, independent of
# any specific target day. Saved every run, alongside the day-specific outputs.
# ---------------------------------------------------------------------------
print("\n" + "=" * 70)
print("Build per-carer HISTORY ANALYSIS workbook (independent of any target day)")
print("=" * 70)

import openpyxl as _openpyxl_hist
from openpyxl.styles import Font as _Font_hist, PatternFill as _PatternFill_hist
from openpyxl.utils import get_column_letter as _get_col_hist

_FONT = 'Arial'
_header_fill = _PatternFill_hist('solid', fgColor='2F5597')
_header_font = _Font_hist(name=_FONT, bold=True, color='FFFFFF')
_title_font = _Font_hist(name=_FONT, bold=True, size=14)
_section_font = _Font_hist(name=_FONT, bold=True, size=12)
_normal_font = _Font_hist(name=_FONT, size=10)
_note_font = _Font_hist(name=_FONT, size=9, italic=True, color='808080')
_weekly_fill = _PatternFill_hist('solid', fgColor='E2EFDA')

hist_wb = _openpyxl_hist.Workbook()
hist_wb.remove(hist_wb.active)

_hist_notes = hist_wb.create_sheet('Read Me')
_notes_text = [
    "Carer HISTORY analysis -- full VisitExport.csv analysis, one sheet per carer",
    "",
    "This is independent of any specific day -- it's the same per-carer analysis you had "
    "before, rebuilt on the current logic. Every carer who has ever appeared in VisitExport.csv "
    "gets a sheet here, not just those working on any particular target day.",
    "",
    "Each sheet has 4 sections:",
    "1. SET ROSTER -- every (client, weekday) slot pattern she's ever had, with the "
    "requirement start AND end time (median), plus the ACTUAL start AND end time (median) "
    "and their variance (std dev in minutes) -- how differently the visit really happens "
    "day to day compared to what was scheduled -- visit count, and the pattern "
    "classification (Weekly / Occasional / etc.).",
    "2. OFF-ROUTINE CLIENTS -- every client outside her Weekly roster she's ever visited, "
    "with visit count and first/last visit dates.",
    "3. GEOGRAPHIC REACH -- her search radius, how many active clients are reachable to her, "
    "how many of those she already covers, and whether she's classified as isolated.",
    "4. EXTEND_FEASIBILITY -- her off-routine rate vs. the peer median, and the resulting "
    "verdict (the same classifier that drives caregivers.json, shown here independent of "
    "any target day).",
]
for _i, _line in enumerate(_notes_text, start=1):
    _c = _hist_notes.cell(row=_i, column=1, value=_line)
    _c.font = _Font_hist(name=_FONT, bold=(_i == 1), size=13 if _i == 1 else 10)
_hist_notes.column_dimensions['A'].width = 130

_hist_summary = hist_wb.create_sheet('Summary')
_summary_headers = ['Carer', 'Tenure (days)', 'Weekly Clients', 'Off-Routine Clients',
                     'Off-Routine Rate', 'Nearby Clients', 'Coverage %', 'Extend?']
for _col, _h in enumerate(_summary_headers, start=1):
    _cell = _hist_summary.cell(row=1, column=_col, value=_h)
    _cell.font = _header_font
    _cell.fill = _header_fill

# Per-carer slot detail: for each (carer, weekday, client), re-cluster her actual raw visits
# by time-of-day (same clustering used throughout the pipeline) to get the FULL picture per
# slot -- not just a single median start time. For each cluster: requirement start AND end
# (median), start-time variance (std dev, in minutes -- how consistent is she really?), visit
# count, and the pattern classification matched from per_carer_client_slot_pattern.
import statistics as _statistics

_carer_client_wd_visits = defaultdict(list)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            _carer_client_wd_visits[(carer, wd, v['client'])].append(v)

_carer_slots = defaultdict(list)  # carer -> [(wd, client, req_start, req_end, variance, n, pattern, act_start, act_end, act_start_var, act_end_var)]
for (carer, wd, client), visits in _carer_client_wd_visits.items():
    occ = [(v['start_dt'],) for v in visits]
    for cluster_occ in cluster_by_time(occ):
        cluster_starts = [o[0] for o in cluster_occ]
        cluster_visits = [v for v in visits if v['start_dt'] in cluster_starts]
        start_minutes = sorted(d.hour * 60 + d.minute for d in cluster_starts)
        end_minutes = sorted(v['end_dt'].hour * 60 + v['end_dt'].minute for v in cluster_visits)
        med_start = start_minutes[len(start_minutes) // 2]
        med_end = end_minutes[len(end_minutes) // 2]
        variance = round(_statistics.pstdev(start_minutes), 1) if len(start_minutes) > 1 else 0.0
        # ACTUAL start/end times -- how differently a visit really happens vs. its
        # requirement time, since the two can genuinely drift apart day to day.
        act_start_minutes = sorted(v['actual_start_dt'].hour * 60 + v['actual_start_dt'].minute
                                    for v in cluster_visits if v['actual_start_dt'] is not None)
        act_end_minutes = sorted(v['actual_end_dt'].hour * 60 + v['actual_end_dt'].minute
                                  for v in cluster_visits if v['actual_end_dt'] is not None)
        if act_start_minutes:
            act_med_start = act_start_minutes[len(act_start_minutes) // 2]
            act_start_str = f"{act_med_start // 60:02d}:{act_med_start % 60:02d}"
            act_start_var = round(_statistics.pstdev(act_start_minutes), 1) if len(act_start_minutes) > 1 else 0.0
        else:
            act_start_str, act_start_var = 'N/A', 'N/A'
        if act_end_minutes:
            act_med_end = act_end_minutes[len(act_end_minutes) // 2]
            act_end_str = f"{act_med_end // 60:02d}:{act_med_end % 60:02d}"
            act_end_var = round(_statistics.pstdev(act_end_minutes), 1) if len(act_end_minutes) > 1 else 0.0
        else:
            act_end_str, act_end_var = 'N/A', 'N/A'
        # match the pattern classification from the existing slot-pattern step, by nearest
        # median start time (same TIME_GAP_MINUTES proximity matching used everywhere else)
        pattern = 'Insufficient history'
        best_diff = None
        for pat, pat_median in per_carer_client_slot_pattern.get((carer, wd, client), []):
            diff = abs(pat_median - med_start)
            if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
                pattern, best_diff = pat, diff
        _carer_slots[carer].append((
            wd, client,
            f"{med_start // 60:02d}:{med_start % 60:02d}",
            f"{med_end // 60:02d}:{med_end % 60:02d}",
            variance, len(cluster_visits), pattern,
            act_start_str, act_end_str, act_start_var, act_end_var,
        ))

# Per-(carer, client) visit stats for the off-routine section
_carer_client_visits = defaultdict(list)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            _carer_client_visits[(carer, v['client'])].append(v['start_dt'].date())

_used_hist_names = set()


def _hist_sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand


_row_i = 2
for carer in sorted(carer_all_visited_clients.keys()):
    presence = carer_presence.get(carer)
    tenure_days = (presence[1] - presence[0]).days if presence else 0
    routine = carer_routine_clients.get(carer, set())
    all_clients = carer_all_visited_clients.get(carer, set())
    off_routine = all_clients - routine
    rate = carer_off_routine_rate.get(carer)
    total_nearby = carer_nearby_total_count.get(carer, 0)
    coverage = carer_nearby_coverage_ratio.get(carer, 0.0)
    extend_verdict = carer_extend_feasibility_flag(carer)

    _hist_summary.cell(row=_row_i, column=1, value=carer).font = _normal_font
    _hist_summary.cell(row=_row_i, column=2, value=tenure_days).font = _normal_font
    _hist_summary.cell(row=_row_i, column=3, value=len(routine)).font = _normal_font
    _hist_summary.cell(row=_row_i, column=4, value=len(off_routine)).font = _normal_font
    _hist_summary.cell(row=_row_i, column=5, value=round(rate, 2) if rate is not None else 'N/A (new)').font = _normal_font
    _hist_summary.cell(row=_row_i, column=6, value=total_nearby).font = _normal_font
    _hist_summary.cell(row=_row_i, column=7, value=f"{coverage:.0%}").font = _normal_font
    _hist_summary.cell(row=_row_i, column=8, value=extend_verdict).font = _normal_font
    _row_i += 1

    ws = hist_wb.create_sheet(_hist_sheet_name(carer, _used_hist_names))
    ws['A1'] = carer
    ws['A1'].font = _title_font
    ws.merge_cells('A1:E1')
    r = 3
    ws.cell(row=r, column=1, value=f"Tenure: {tenure_days} days "
            f"({presence[0]} to {presence[1]})" if presence else "Tenure: no presence data").font = _note_font
    r += 2

    ws.cell(row=r, column=1, value='1. SET ROSTER').font = _section_font
    r += 1
    _headers = ['Client', 'Weekday', 'Requirement Start', 'Requirement End', 'Time Variance (min)',
                'Actual Start', 'Actual End', 'Actual Start Variance (min)', 'Actual End Variance (min)',
                'Visits', 'Pattern']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1
    for wd, client, start_str, end_str, variance, n_visits, pattern, act_start, act_end, act_start_var, act_end_var in sorted(
            _carer_slots.get(carer, []), key=lambda x: (x[1], x[0])):
        vals = [client, wd, start_str, end_str, variance, act_start, act_end, act_start_var, act_end_var, n_visits, pattern]
        fill = _weekly_fill if pattern == 'Weekly' else None
        for _col, v in enumerate(vals, start=1):
            _cell = ws.cell(row=r, column=_col, value=v)
            _cell.font = _normal_font
            if fill:
                _cell.fill = fill
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='2. OFF-ROUTINE CLIENTS').font = _section_font
    r += 1
    _headers = ['Client', 'Visit Count', 'First Visit', 'Last Visit']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1
    for client in sorted(off_routine):
        dates = sorted(_carer_client_visits.get((carer, client), []))
        if not dates:
            continue
        vals = [client, len(dates), dates[0], dates[-1]]
        for _col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=_col, value=v).font = _normal_font
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='3. GEOGRAPHIC REACH').font = _section_font
    r += 1
    ws.cell(row=r, column=1, value=f"Search radius: {carer_search_radius.get(carer, 'N/A')} km").font = _normal_font
    r += 1
    ws.cell(row=r, column=1,
            value=f"{total_nearby} active client(s) reachable to her, {coverage:.0%} already covered "
                  f"({'ISOLATED' if total_nearby <= ISOLATED_TOTAL_THRESHOLD and coverage >= ISOLATED_COVERAGE_THRESHOLD else 'not isolated'})").font = _normal_font
    r += 2

    ws.cell(row=r, column=1, value='4. EXTEND_FEASIBILITY').font = _section_font
    r += 1
    if rate is None:
        ws.cell(row=r, column=1, value=f"NEW CARER (tenure {tenure_days}d < {NEW_CARER_TENURE_DAYS}d) -- defaults open").font = _normal_font
    else:
        verdict_text = "AT/ABOVE" if rate >= MEDIAN_OFF_ROUTINE_RATE else "BELOW"
        ws.cell(row=r, column=1,
                value=f"Off-routine rate {rate:.2f} per {CASELOAD_WINDOW_DAYS}d vs peer median "
                      f"{MEDIAN_OFF_ROUTINE_RATE:.2f} -- {verdict_text} median").font = _normal_font
    r += 1
    ws.cell(row=r, column=1, value=f"Final verdict: extend = {extend_verdict}").font = _normal_font

    _widths = [30, 12, 16, 16, 18, 14, 14, 18, 18, 8, 20]
    for _i, _w in enumerate(_widths, start=1):
        ws.column_dimensions[_get_col_hist(_i)].width = _w

_hist_summary.column_dimensions['A'].width = 30
for _col in 'BCDEFGH':
    _hist_summary.column_dimensions[_col].width = 16
_hist_summary.freeze_panes = 'A2'

_hist_path = f'{OUTPUT_DIR}/Carer_History_Analysis.xlsx'
import os as _os_hist
_os_hist.makedirs(OUTPUT_DIR, exist_ok=True)
hist_wb.save(_hist_path)
print(f"Saved {_hist_path} ({len(hist_wb.sheetnames)} sheets, {len(carer_all_visited_clients)} carers)")

# ---------------------------------------------------------------------------
# Carer Roster Coverage -- the fuller roster/substitute-coverage analysis. Four sections per
# carer: SET ROSTER (with real double-up overlap detection), EXTRA VISITS (with a
# per-occurrence "Regular Carer(s)" and "Likely Reason" classification), SUBSTITUTE COVERAGE
# (every week within a set-roster slot's span where the regular carer didn't do the visit --
# who covered it, if anyone), and GEOGRAPHIC COVERAGE (nearby clients broken into already on
# her roster / covered via extras / never visited).
# ---------------------------------------------------------------------------
print("\n" + "=" * 70)
print("Build CARER ROSTER COVERAGE workbook (set roster, extra visits, substitute coverage)")
print("=" * 70)

rc_wb = _openpyxl_hist.Workbook()
rc_wb.remove(rc_wb.active)

_rc_notes = rc_wb.create_sheet('Read Me')
_rc_notes_text = [
    "Roster coverage: extra visits and substitute cover",
    "",
    "SET ROSTER = every (carer, weekday, client) relationship classified WEEKLY (>=75% "
    "consistency in the carer's own active weeks, or the daily-pooled fallback).",
    "",
    "Each carer's sheet has four sections:",
    "1. SET ROSTER -- the fixed weekly schedule, with requirement and actual (median + real "
    "range) times, consistency%, and real Double-Up detection: whether ANOTHER carer "
    "genuinely overlapped in time with this same client on the same date (not just a "
    "same-day coincidence) -- shown as what fraction of visits had a real overlapping "
    "partner, and who. Rows with any double-up are highlighted.",
    "2. EXTRA VISITS -- clients visited outside the set roster. A per-weekday summary "
    "(names and counts), then a DETAIL table with exact times per visit, company-wide "
    "staffing context for those specific dates, whether the client has a set roster at all, "
    "who the regular carer is if so, and a Likely Reason -- cross-checking whether the "
    "regular carer was actually active that same week (covering an absence) or not "
    "(reason unclear -- possibly a genuine extra need).",
    "3. SUBSTITUTE COVERAGE -- for every set-roster slot, every week within its own span "
    "where the regular carer did NOT do the visit: covered by someone else (who, when) or "
    "no visit found that week, plus whether the regular carer was even working that "
    "weekday at all that week.",
    "4. GEOGRAPHIC COVERAGE -- her real search radius (median travel distance + 5km "
    "buffer, not skewed by rare outlier trips), and every other active client within that "
    "radius broken into: already on her set roster, covered via extra visits, or never "
    "visited.",
]
for _i, _line in enumerate(_rc_notes_text, start=1):
    _c = _rc_notes.cell(row=_i, column=1, value=_line)
    _c.font = _Font_hist(name=_FONT, bold=(_i == 1), size=13 if _i == 1 else 10)
_rc_notes.column_dimensions['A'].width = 130

# --- shared precomputation, reused across every carer's sheet ---
_client_wd_regular_carers = defaultdict(set)  # (client, wd) -> set of carer names with a Weekly slot
for (_carer_k, _wd_k), _clients_k in set_roster.items():
    for _client_k in _clients_k:
        _client_wd_regular_carers[(_client_k, _wd_k)].add(_carer_k)

_date_carers = defaultdict(set)
_date_clients = defaultdict(set)
_date_visit_count = defaultdict(int)
for _carer_k, _wd_map_k in roster.items():
    for _wd_k, _visits_k in _wd_map_k.items():
        for _v_k in _visits_k:
            _d_k = _v_k['start_dt'].date()
            _date_carers[_d_k].add(_carer_k)
            _date_clients[_d_k].add(_v_k['client'])
            _date_visit_count[_d_k] += 1

_client_date_visits = defaultdict(list)  # (client, date) -> [(carer, start_dt, end_dt), ...]
for _carer_k, _wd_map_k in roster.items():
    for _wd_k, _visits_k in _wd_map_k.items():
        for _v_k in _visits_k:
            _client_date_visits[(_v_k['client'], _v_k['start_dt'].date())].append(
                (_carer_k, _v_k['actual_start_dt'] or _v_k['start_dt'], _v_k['actual_end_dt'] or _v_k['end_dt']))

_regular_carer_client_wd_weeks = defaultdict(set)  # (carer, client, wd) -> set of isoweeks she visited that client that weekday
for _carer_k, _wd_map_k in roster.items():
    for _wd_k, _visits_k in _wd_map_k.items():
        for _v_k in _visits_k:
            _regular_carer_client_wd_weeks[(_carer_k, _v_k['client'], _wd_k)].add(isoweek(_v_k['start_dt'].date()))

_used_rc_names = set()

for carer in sorted(carer_all_visited_clients.keys()):
    ws = rc_wb.create_sheet(_hist_sheet_name(carer, _used_rc_names))
    ws['A1'] = carer
    ws['A1'].font = _title_font
    ws.merge_cells('A1:K1')
    r = 3

    # ============ SECTION 1: SET ROSTER ============
    ws.cell(row=r, column=1, value='SET ROSTER (the fixed weekly schedule)').font = _section_font
    r += 1
    _headers = ['Weekday', 'Client', 'Service Required Start and End', 'Actual Start and End',
                'Consistency', 'Double-Up', 'Double-Up Partner(s)']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1

    for wd in WEEKDAYS:
        for client in sorted(set_roster.get((carer, wd), set())):
            visits = [v for v in roster.get(carer, {}).get(wd, []) if v['client'] == client]
            if not visits:
                continue
            req_starts = sorted(v['start_dt'].hour * 60 + v['start_dt'].minute for v in visits)
            req_ends = sorted(v['end_dt'].hour * 60 + v['end_dt'].minute for v in visits)
            act_starts = sorted(v['actual_start_dt'].hour * 60 + v['actual_start_dt'].minute
                                 for v in visits if v['actual_start_dt'] is not None)
            act_ends = sorted(v['actual_end_dt'].hour * 60 + v['actual_end_dt'].minute
                               for v in visits if v['actual_end_dt'] is not None)
            req_s_med, req_e_med = req_starts[len(req_starts) // 2], req_ends[len(req_ends) // 2]
            req_str = f"{req_s_med // 60:02d}:{req_s_med % 60:02d} - {req_e_med // 60:02d}:{req_e_med % 60:02d}"
            if act_starts and act_ends:
                a_s_med, a_e_med = act_starts[len(act_starts) // 2], act_ends[len(act_ends) // 2]
                act_str = (f"{a_s_med // 60:02d}:{a_s_med % 60:02d} (varies {act_starts[0] // 60:02d}:{act_starts[0] % 60:02d}-"
                           f"{act_starts[-1] // 60:02d}:{act_starts[-1] % 60:02d}) - "
                           f"{a_e_med // 60:02d}:{a_e_med % 60:02d} (varies {act_ends[0] // 60:02d}:{act_ends[0] % 60:02d}-"
                           f"{act_ends[-1] // 60:02d}:{act_ends[-1] % 60:02d})")
            else:
                act_str = 'N/A'
            _slot_dates = sorted(set(v['start_dt'].date() for v in visits))
            _win_start_wk, _win_end_wk = isoweek(_slot_dates[0]), isoweek(_slot_dates[-1])
            _active_weeks_in_window = sorted(w for w in carer_weekday_active_weeks.get((carer, wd), set())
                                              if _win_start_wk <= w <= _win_end_wk)
            _hit_weeks = set(isoweek(d) for d in _slot_dates)
            _n_active = len(_active_weeks_in_window)
            _n_hits = len(set(_active_weeks_in_window) & _hit_weeks)
            consistency_pct = f"{_n_hits / _n_active * 100:.0f}%" if _n_active else 'N/A'

            dates_for_slot = sorted(set(v['start_dt'].date() for v in visits))
            dup_dates = 0
            dup_partners = set()
            for d in dates_for_slot:
                this_visit = next((v for v in visits if v['start_dt'].date() == d), None)
                if not this_visit:
                    continue
                this_s = this_visit['actual_start_dt'] or this_visit['start_dt']
                this_e = this_visit['actual_end_dt'] or this_visit['end_dt']
                found = False
                for other_carer, o_s, o_e in _client_date_visits.get((client, d), []):
                    if other_carer == carer:
                        continue
                    if o_s < this_e and this_s < o_e:
                        found = True
                        dup_partners.add(other_carer)
                if found:
                    dup_dates += 1
            dup_frac = f"{dup_dates / len(dates_for_slot) * 100:.0f}%" if dates_for_slot else ''
            dup_partner_str = ', '.join(sorted(dup_partners)) if dup_partners else ''

            vals = [wd, client, req_str, act_str, consistency_pct,
                    dup_frac if dup_dates else None, dup_partner_str if dup_dates else None]
            fill = _PatternFill_hist('solid', fgColor='FFE0B2') if dup_dates else None
            for _col, v in enumerate(vals, start=1):
                _cell = ws.cell(row=r, column=_col, value=v)
                _cell.font = _normal_font
                if fill:
                    _cell.fill = fill
            r += 1
    r += 1

    # ============ SECTION 2: EXTRA VISITS ============
    ws.cell(row=r, column=1, value='EXTRA VISITS (beyond set roster)').font = _section_font
    r += 1
    _headers = ['Weekday', 'Set Roster Size', 'Extra People', 'Extra Visit Occurrences', 'Extra Clients (names)']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1

    _extra_by_wd = {}
    for wd in WEEKDAYS:
        roster_clients = set_roster.get((carer, wd), set())
        visits = roster.get(carer, {}).get(wd, [])
        extra_counts = Counter(v['client'] for v in visits if v['client'] not in roster_clients)
        _extra_by_wd[wd] = extra_counts
        if not extra_counts:
            continue
        names_str = ', '.join(f"{c} (x{n})" for c, n in extra_counts.most_common())
        vals = [wd, len(roster_clients), len(extra_counts), sum(extra_counts.values()), names_str]
        for _col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=_col, value=v).font = _normal_font
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='EXTRA VISITS DETAIL (by time)').font = _section_font
    r += 1
    _headers = ['Weekday', 'Client', 'Service Required Start and End', 'Actual Start and End', 'Visits',
                'Carers Working That Day (avg)', 'Clients Seen That Day (avg)', 'Total Visits That Day (avg)',
                'Client Has Set Roster?', 'Regular Carer(s)', 'Likely Reason']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1

    for wd in WEEKDAYS:
        roster_clients = set_roster.get((carer, wd), set())
        visits = [v for v in roster.get(carer, {}).get(wd, []) if v['client'] not in roster_clients]
        by_client = defaultdict(list)
        for v in visits:
            by_client[v['client']].append(v)
        for client, cvisits in sorted(by_client.items()):
            for cluster_occ in cluster_by_time([(v['start_dt'],) for v in cvisits]):
                cluster_starts = set(o[0] for o in cluster_occ)
                cluster_visits = [v for v in cvisits if v['start_dt'] in cluster_starts]
                req_starts = sorted(v['start_dt'].hour * 60 + v['start_dt'].minute for v in cluster_visits)
                req_ends = sorted(v['end_dt'].hour * 60 + v['end_dt'].minute for v in cluster_visits)
                act_starts = sorted(v['actual_start_dt'].hour * 60 + v['actual_start_dt'].minute
                                     for v in cluster_visits if v['actual_start_dt'] is not None)
                act_ends = sorted(v['actual_end_dt'].hour * 60 + v['actual_end_dt'].minute
                                   for v in cluster_visits if v['actual_end_dt'] is not None)
                req_s_med, req_e_med = req_starts[len(req_starts) // 2], req_ends[len(req_ends) // 2]
                req_str = f"{req_s_med // 60:02d}:{req_s_med % 60:02d} - {req_e_med // 60:02d}:{req_e_med % 60:02d}"
                if act_starts and act_ends:
                    if len(act_starts) > 1:
                        a_s_med = act_starts[len(act_starts) // 2]
                        a_e_med = act_ends[len(act_ends) // 2]
                        act_str = (f"{a_s_med // 60:02d}:{a_s_med % 60:02d} (varies {act_starts[0] // 60:02d}:{act_starts[0] % 60:02d}-"
                                   f"{act_starts[-1] // 60:02d}:{act_starts[-1] % 60:02d}) - "
                                   f"{a_e_med // 60:02d}:{a_e_med % 60:02d} (varies {act_ends[0] // 60:02d}:{act_ends[0] % 60:02d}-"
                                   f"{act_ends[-1] // 60:02d}:{act_ends[-1] % 60:02d})")
                    else:
                        act_str = f"{act_starts[0] // 60:02d}:{act_starts[0] % 60:02d} - {act_ends[0] // 60:02d}:{act_ends[0] % 60:02d}"
                else:
                    act_str = 'N/A'

                cluster_dates = sorted(set(v['start_dt'].date() for v in cluster_visits))
                n_visits = len(cluster_visits)
                avg_carers = sum(len(_date_carers.get(d, set())) for d in cluster_dates) / len(cluster_dates)
                avg_clients = sum(len(_date_clients.get(d, set())) for d in cluster_dates) / len(cluster_dates)
                avg_total_visits = sum(_date_visit_count.get(d, 0) for d in cluster_dates) / len(cluster_dates)

                regulars = _client_wd_regular_carers.get((client, wd), set())
                has_roster = 'Yes' if regulars else 'No'
                regulars_str = ', '.join(sorted(regulars)) if regulars else None

                if not regulars:
                    reason = 'Client has no set roster for this weekday (ad hoc/occasional client)'
                else:
                    covering_counts = Counter()
                    unclear = 0
                    for d in cluster_dates:
                        wk = isoweek(d)
                        any_regular_active = False
                        for reg in regulars:
                            if wk in _regular_carer_client_wd_weeks.get((reg, client, wd), set()):
                                any_regular_active = True
                        if any_regular_active:
                            unclear += 1
                        else:
                            for reg in regulars:
                                if wk not in _regular_carer_client_wd_weeks.get((reg, client, wd), set()):
                                    covering_counts[reg] += 1
                    total = len(cluster_dates)
                    if covering_counts and not unclear:
                        parts = ', '.join(f"{c} ({n}/{total})" for c, n in covering_counts.most_common())
                        reason = f"Covering for regular carer's absence -- {parts}"
                    elif covering_counts and unclear:
                        parts = ', '.join(f"{c} ({n}/{total})" for c, n in covering_counts.most_common())
                        reason = f"Partly covering absence ({sum(covering_counts.values())}/{total}: {parts}); rest unclear (possible additional need)"
                    else:
                        reason = 'Regular carer was present those weeks -- reason unclear (possibly an additional/extra need)'

                vals = [wd, client, req_str, act_str, n_visits, round(avg_carers, 1), round(avg_clients, 1),
                        round(avg_total_visits, 1), has_roster, regulars_str, reason]
                for _col, v in enumerate(vals, start=1):
                    ws.cell(row=r, column=_col, value=v).font = _normal_font
                r += 1
    r += 1

    # ============ SECTION 3: SUBSTITUTE COVERAGE ============
    ws.cell(row=r, column=1, value='SUBSTITUTE COVERAGE').font = _section_font
    r += 1
    _headers = ['Weekday', 'Client', 'Week', 'Status', 'Covering Carer', 'Covering Time', 'Carer Active That Weekday']
    for _col, _h in enumerate(_headers, start=1):
        _c = ws.cell(row=r, column=_col, value=_h)
        _c.font = _header_font
        _c.fill = _header_fill
    r += 1

    for wd in WEEKDAYS:
        for client in sorted(set_roster.get((carer, wd), set())):
            visits = [v for v in roster.get(carer, {}).get(wd, []) if v['client'] == client]
            if not visits:
                continue
            slot_dates = sorted(v['start_dt'].date() for v in visits)
            span_start, span_end = slot_dates[0], slot_dates[-1]
            typical_minute = sorted(v['start_dt'].hour * 60 + v['start_dt'].minute for v in visits)[len(visits) // 2]
            visited_weeks = set(isoweek(d) for d in slot_dates)
            active_weeks_in_span = sorted(w for w in carer_weekday_active_weeks.get((carer, wd), set())
                                           if isoweek(span_start) <= w <= isoweek(span_end))
            for wk in active_weeks_in_span:
                if wk in visited_weeks:
                    continue
                carer_active_this_wd_that_week = any(
                    isoweek(d) == wk for d in (v['start_dt'].date() for v in roster.get(carer, {}).get(wd, [])))
                # was the carer active on THIS weekday during this week AT ALL (any client)?
                covering_carer, covering_time = None, None
                for other_carer, other_wd_map in roster.items():
                    if other_carer == carer:
                        continue
                    for v2 in other_wd_map.get(wd, []):
                        if v2['client'] != client:
                            continue
                        if isoweek(v2['start_dt'].date()) != wk:
                            continue
                        this_min = v2['start_dt'].hour * 60 + v2['start_dt'].minute
                        if abs(this_min - typical_minute) <= TIME_GAP_MINUTES:
                            covering_carer = other_carer
                            covering_time = f"{v2['start_dt'].hour:02d}:{v2['start_dt'].minute:02d}"
                            break
                    if covering_carer:
                        break
                status = 'Covered by another carer' if covering_carer else 'No visit found that week'
                vals = [wd, client, f"{wk[0]}-W{wk[1]:02d}", status, covering_carer, covering_time,
                        carer_active_this_wd_that_week]
                fill = _PatternFill_hist('solid', fgColor='FFE0B2') if covering_carer else _PatternFill_hist('solid', fgColor='E0E0E0')
                for _col, v in enumerate(vals, start=1):
                    _cell = ws.cell(row=r, column=_col, value=v)
                    _cell.font = _normal_font
                    _cell.fill = fill
                r += 1
    r += 1

    # ============ SECTION 4: GEOGRAPHIC COVERAGE ============
    ws.cell(row=r, column=1, value='GEOGRAPHIC COVERAGE').font = _section_font
    r += 1
    all_clients_ever = carer_all_visited_clients.get(carer, set())
    dists = [(c, travel_km(carer, c)) for c in all_clients_ever]
    dists = [(c, d) for c, d in dists if d is not None]
    if dists:
        furthest_client, furthest_dist = max(dists, key=lambda x: x[1])
        ws.cell(row=r, column=1, value=f"Biggest Travel: {furthest_dist:.1f} km, Furthest Client: {furthest_client}").font = _normal_font
        r += 1
        radius = carer_search_radius.get(carer)
        ws.cell(row=r, column=1, value=f"Search Radius: {radius:.1f} km" if radius else "Search Radius: N/A").font = _normal_font
        r += 2
        if radius:
            roster_clients_ever = set()
            for wd in WEEKDAYS:
                roster_clients_ever |= set_roster.get((carer, wd), set())
            _headers = ['Client', 'Distance (km)', 'Status']
            for _col, _h in enumerate(_headers, start=1):
                _c = ws.cell(row=r, column=_col, value=_h)
                _c.font = _header_font
                _c.fill = _header_fill
            r += 1
            for client_name, info in sorted(client_info.items()):
                d = travel_km(carer, client_name)
                if d is None or d > radius:
                    continue
                if client_name in roster_clients_ever:
                    status = 'Already On Her Roster'
                elif client_name in all_clients_ever:
                    status = 'Covered By Extra Visits'
                else:
                    status = 'Never Visited'
                ws.cell(row=r, column=1, value=client_name).font = _normal_font
                ws.cell(row=r, column=2, value=round(d, 1)).font = _normal_font
                ws.cell(row=r, column=3, value=status).font = _normal_font
                r += 1
    else:
        ws.cell(row=r, column=1, value="No coordinate-matched visited clients -- geographic coverage unavailable.").font = _note_font

    _widths = [16, 32, 24, 40, 12, 12, 20]
    for _i, _w in enumerate(_widths, start=1):
        ws.column_dimensions[_get_col_hist(_i)].width = _w

_rc_path = f'{OUTPUT_DIR}/Carer_Roster_Coverage.xlsx'
rc_wb.save(_rc_path)
print(f"Saved {_rc_path} ({len(rc_wb.sheetnames)} sheets, {len(carer_all_visited_clients)} carers)")

# ---------------------------------------------------------------------------
# Coverage Risk Analysis -- for any carer who's left (detected automatically once their
# status is marked Deactive, or listed manually in KNOWN_DEPARTED_CARERS for departures the
# source data hasn't caught up with yet), find every client where she was the dominant
# historical carer, and for each: the full weekday-by-weekday breakdown of who else has real
# history with them, plus each candidate's own workload on that weekday (visits/day) so it's
# clear who genuinely has room versus who's already fully loaded. This is informational only
# -- it doesn't change any feasibility weight; it's for a person to review and decide.
# ---------------------------------------------------------------------------
print("\n" + "=" * 70)
print("Build COVERAGE RISK ANALYSIS workbook (departed carers' clients, informational only)")
print("=" * 70)

# "Departed" is based on the real terminationDate field, scoped to RECENT/UPCOMING
# departures only -- a carer counts here if her termination date falls within 14 days
# before OR after today, so this stays focused on transitions that actually need managing
# now, not decades of past staff turnover. A future terminationDate is included on purpose:
# if she's known to be leaving soon and simply isn't working today, the whole week's roster
# for her clients should be sorted out proactively, not only after she's actually gone.
# Whether she's ACTUALLY working today is checked separately, once today's real caregiver
# list is known (see below, after the day export loads) -- if she IS working today despite
# having an upcoming termination date, she just covers her own clients as normal; the
# rotation only applies on days she's genuinely absent.
RECENT_TERMINATION_WINDOW_DAYS = 14
_departed_carers = sorted(
    c for c, term_date in carer_termination_date.items()
    if abs((term_date - TARGET_DATE).days) <= RECENT_TERMINATION_WINDOW_DAYS
    and c in carer_all_visited_clients
)
print(f"Departed carers considered (terminationDate within {RECENT_TERMINATION_WINDOW_DAYS} days "
      f"before or after {TARGET_DATE}): {_departed_carers}")

_at_risk_clients = {}  # client -> departed carer who was dominant for them
for _dep_carer in _departed_carers:
    if _dep_carer not in carer_all_visited_clients:
        continue
    for _client in carer_all_visited_clients[_dep_carer]:
        _totals = Counter()
        for _wd_all in WEEKDAYS:
            for _cluster in client_slot_history.get((_client, _wd_all), []):
                for _c, _dt, _a_s in _cluster:
                    _totals[_c] += 1
        if _totals and _totals.most_common(1)[0][0] == _dep_carer:
            _at_risk_clients[_client] = _dep_carer

print(f"At-risk clients found (departed carer was their dominant historical carer): {len(_at_risk_clients)}")

# For each departed carer, and each weekday, automatically pick ONE carer to take over that
# weekday's at-risk clients FOR THAT DEPARTED CARER SPECIFICALLY -- computed separately per
# departed carer so one person's rotation is never diluted by an unrelated departure's
# pattern. Balances genuine historical connection (has she really covered these clients
# before) against genuine capacity (is she already fully loaded, or does she have real room).
#
# For each departed carer, compute a score per (candidate, weekday), then assign weekdays to
# candidates with a FAIRNESS CAP so the load is spread across the real pool rather than one
# strong historical candidate taking every single day. Score = her (emergency-filtered) visit
# count to this departed carer's at-risk clients on that weekday, divided by her own average
# daily workload on that weekday (higher = strong genuine connection AND real room).
#
# Before scoring, EMERGENCY-ONLY coverage is filtered out: for each covering instance, check
# whether any OTHER pool member was also working that same date. If nobody else was around,
# that instance was a one-off emergency, not a genuine pattern, and doesn't count.
#
# Fairness cap: no candidate is assigned more than ceil(7 / pool size) weekdays (minimum 1) --
# once she hits her share, the next-best genuine candidate for a remaining weekday takes it
# instead. All 7 weekdays are assigned in one pass (highest-scoring pairs first), not one
# weekday at a time, so the cap is respected globally, not just locally.
AT_RISK_ROTATION = {}  # (departed_carer, weekday) -> selected carer
MIN_DISTINCT_CLIENTS_FOR_POOL = 2  # a candidate must have real history with at least this
# many of the departed carer's at-risk clients (not just one) to be considered a genuine
# pool member -- someone with a single, thin visit to just one client isn't a real
# established backup, just a coincidence.
for _dep_carer in _departed_carers:
    _dep_clients = {c for c, dc in _at_risk_clients.items() if dc == _dep_carer}
    if not _dep_clients:
        continue

    _candidate_client_counts = Counter()
    for _client in _dep_clients:
        for _cand, _cand_clients in carer_all_visited_clients.items():
            if _cand not in _departed_carers and _client in _cand_clients:
                _candidate_client_counts[_cand] += 1
    _eligible_candidates = {c for c, n in _candidate_client_counts.items() if n >= MIN_DISTINCT_CLIENTS_FOR_POOL}

    _wd_scores = {}  # weekday -> {candidate: score}
    _all_candidates = set()
    for _wd in WEEKDAYS:
        _pool_dates = defaultdict(set)
        for _client in _dep_clients:
            for _cluster in client_slot_history.get((_client, _wd), []):
                for _c, _dt, _a_s in _cluster:
                    if _c in _eligible_candidates:
                        _pool_dates[_c].add(_dt.date())
        if not _pool_dates:
            continue
        _pool_members = set(_pool_dates.keys())
        _wd_candidate_visits = Counter()
        for _cand, _dates in _pool_dates.items():
            for _d in _dates:
                _others_working = any(
                    any(_v['start_dt'].date() == _d for _v in roster.get(_other, {}).get(_wd, []))
                    for _other in _pool_members if _other != _cand
                )
                if _others_working:
                    _wd_candidate_visits[_cand] += 1
        if not _wd_candidate_visits:
            continue
        _scores = {}
        for _cand, _n in _wd_candidate_visits.items():
            _cand_visits = roster.get(_cand, {}).get(_wd, [])
            _cand_dates = set(_v['start_dt'].date() for _v in _cand_visits)
            _cand_load = len(_cand_visits) / max(len(_cand_dates), 1) if _cand_dates else 0.01
            # concentration_factor rewards focused carers and penalizes generalists (same
            # factor used everywhere else in the pipeline) -- so a candidate with genuinely
            # substantial, focused history is preferred over a broad generalist even when
            # their raw visits/workload ratio comes out close or slightly behind.
            _scores[_cand] = _n / (_cand_load + 0.01) * concentration_factor(_cand)
        _wd_scores[_wd] = _scores
        _all_candidates.update(_scores.keys())

    if not _wd_scores or not _all_candidates:
        continue

    _cap = max(1, -(-7 // len(_all_candidates)))  # ceil(7 / pool size)
    _assigned_count = Counter()
    _all_pairs = [(_score, _wd, _cand) for _wd, _scores in _wd_scores.items() for _cand, _score in _scores.items()]
    _all_pairs.sort(key=lambda x: -x[0])
    _assigned_days = set()
    for _score, _wd, _cand in _all_pairs:
        if _wd in _assigned_days:
            continue
        if _assigned_count[_cand] >= _cap:
            continue
        AT_RISK_ROTATION[(_dep_carer, _wd)] = _cand
        _assigned_count[_cand] += 1
        _assigned_days.add(_wd)
    # Second pass: coverage beats strict fairness. A weekday can be left unassigned if its
    # only qualified candidate already hit her cap on a higher-scoring day elsewhere -- fill
    # any such gap with the best available candidate for that day, even if it means she
    # exceeds her "fair share". An uncovered day is worse than an uneven load.
    for _wd, _scores in _wd_scores.items():
        if _wd in _assigned_days:
            continue
        _best_cand = max(_scores, key=_scores.get)
        AT_RISK_ROTATION[(_dep_carer, _wd)] = _best_cand
        _assigned_count[_best_cand] += 1
        _assigned_days.add(_wd)

print("Automatic at-risk rotation (one carer per departed carer per weekday, balancing history against real capacity):")
for (_dep_carer, _wd), _carer in AT_RISK_ROTATION.items():
    print(f"  {_dep_carer} / {_wd}: {_carer}")

if _at_risk_clients:
    cov_wb = _openpyxl_hist.Workbook()
    cov_wb.remove(cov_wb.active)

    _cov_notes = cov_wb.create_sheet('Read Me')
    _cov_notes_text = [
        "Coverage Risk Analysis -- clients whose dominant historical carer has left",
        "",
        f"Departed carers considered: {', '.join(_departed_carers)}",
        "",
        "This is informational only -- nothing here changes any feasibility weight. Each "
        "sheet is one at-risk client, showing every other carer with real historical "
        "involvement, broken down by weekday, alongside each candidate's own workload "
        "(visits/day) on that same weekday -- so it's clear who has genuine room versus who's "
        "already fully loaded elsewhere. Use this to decide how to redistribute coverage; "
        "nothing here is applied automatically.",
    ]
    for _i, _line in enumerate(_cov_notes_text, start=1):
        _c = _cov_notes.cell(row=_i, column=1, value=_line)
        _c.font = _Font_hist(name=_FONT, bold=(_i == 1), size=13 if _i == 1 else 10)
    _cov_notes.column_dimensions['A'].width = 130

    _cov_summary = cov_wb.create_sheet('Summary')
    _cov_summary_headers = ['Client', 'Departed Carer', 'Days with Any Real Alternative']
    for _col, _h in enumerate(_cov_summary_headers, start=1):
        _cell = _cov_summary.cell(row=1, column=_col, value=_h)
        _cell.font = _header_font
        _cell.fill = _header_fill

    _used_cov_names = set()
    _cov_row_i = 2
    for _client in sorted(_at_risk_clients.keys()):
        _dep_carer = _at_risk_clients[_client]
        _days_with_alt = 0
        _ws = cov_wb.create_sheet(_hist_sheet_name(_client, _used_cov_names))
        _ws['A1'] = _client
        _ws['A1'].font = _title_font
        _ws.merge_cells('A1:F1')
        _ws.cell(row=3, column=1, value=f"Dominant/departed carer: {_dep_carer}").font = _note_font
        _r = 5
        for _wd in WEEKDAYS:
            _wd_totals = Counter()
            _wd_last_seen = {}
            for _idx, _cluster in enumerate(client_slot_history.get((_client, _wd), [])):
                for _c, _dt, _a_s in _cluster:
                    if _c == _dep_carer:
                        continue
                    _wd_totals[_c] += 1
                    _d = _dt.date()
                    if _c not in _wd_last_seen or _d > _wd_last_seen[_c]:
                        _wd_last_seen[_c] = _d
            _ws.cell(row=_r, column=1, value=_wd).font = _section_font
            _r += 1
            if not _wd_totals:
                _ws.cell(row=_r, column=1, value="No other carer has real history for this day.").font = _note_font
                _r += 2
                continue
            _days_with_alt += 1
            _headers = ['Candidate Carer', 'Visits (this weekday)', 'Last Seen', "Candidate's Own Workload That Weekday (visits/day)"]
            for _col, _h in enumerate(_headers, start=1):
                _c2 = _ws.cell(row=_r, column=_col, value=_h)
                _c2.font = _header_font
                _c2.fill = _header_fill
            _r += 1
            for _cand, _n in _wd_totals.most_common():
                _cand_visits = roster.get(_cand, {}).get(_wd, [])
                _cand_dates = set(_v['start_dt'].date() for _v in _cand_visits)
                _cand_load = len(_cand_visits) / max(len(_cand_dates), 1)
                _ws.cell(row=_r, column=1, value=_cand).font = _normal_font
                _ws.cell(row=_r, column=2, value=_n).font = _normal_font
                _ws.cell(row=_r, column=3, value=_wd_last_seen.get(_cand)).font = _normal_font
                _ws.cell(row=_r, column=4, value=round(_cand_load, 1)).font = _normal_font
                _r += 1
            _r += 1
        for _col, _w in zip('ABCD', [30, 22, 14, 40]):
            _ws.column_dimensions[_col].width = _w

        _cov_summary.cell(row=_cov_row_i, column=1, value=_client).font = _normal_font
        _cov_summary.cell(row=_cov_row_i, column=2, value=_dep_carer).font = _normal_font
        _cov_summary.cell(row=_cov_row_i, column=3, value=_days_with_alt).font = _normal_font
        _cov_row_i += 1

    _cov_summary.column_dimensions['A'].width = 30
    _cov_summary.column_dimensions['B'].width = 22
    _cov_summary.column_dimensions['C'].width = 28
    _cov_summary.freeze_panes = 'A2'

    _cov_path = f'{OUTPUT_DIR}/Coverage_Risk_Analysis.xlsx'
    cov_wb.save(_cov_path)
    print(f"Saved {_cov_path} ({len(cov_wb.sheetnames)} sheets, {len(_at_risk_clients)} at-risk clients)")
else:
    print("No at-risk clients found -- skipping Coverage Risk Analysis workbook.")

print("\n" + "=" * 70)
print(f"Loading real day export: {len(_real_patients_raw['patients'])} patients, "
      f"{len(_real_caregivers_raw['caregivers'])} caregivers")
print("Only identity + need/availability time is trusted from this export -- everything")
print("else (duration, windows, priorities, extend_feasibility) is recomputed from history.")
print("=" * 70)

# ---------------------------------------------------------------------------
# Extract ONLY identity + need/availability time from the real export.
#
# Two input formats are supported: the older start_time_soft/end_time_soft naming, and the
# newer requestStartTime/requestEndTime naming (the exact "Service Requirement start/end
# time" concept -- the SAME concept used to build history from VisitExport.csv, so matching
# today's requests against historical slots is done on a consistent basis).
# ---------------------------------------------------------------------------
day_requirements = []  # {pid, prid, client, start_min, end_min, match_request_list}
for p in _real_patients_raw['patients']:
    client = f"{p.get('name', '')} {p.get('lastname', '')}".strip()
    rw = p['request_window']
    start_min = rw.get('requestStartTime', rw.get('start_time_soft'))
    end_min = rw.get('requestEndTime', rw.get('end_time_soft'))
    day_requirements.append({
        'pid': p['pid'], 'prid': p['prid'], 'client': client,
        'start_min': start_min, 'end_min': end_min,
        'match_request_list': list(rw.get('match_request_list', [])),
        'gender': p.get('gender'),
        'first_name': (p.get('name') or '').strip(), 'last_name': (p.get('lastname') or '').strip(),
        'location': p.get('location'),
    })

# If the input doesn't already provide match_request_list, detect double-ups ourselves:
# two requests for the SAME pid today with genuinely OVERLAPPING [start_min, end_min]
# windows both need a carer at the same time -- that's a double-up, regardless of whether
# the export says so explicitly.
_by_pid_today = defaultdict(list)
for _r in day_requirements:
    _by_pid_today[_r['pid']].append(_r)
_detected_double_ups = 0
for _pid, _reqs in _by_pid_today.items():
    if len(_reqs) < 2:
        continue
    for _i in range(len(_reqs)):
        for _j in range(_i + 1, len(_reqs)):
            _a, _b = _reqs[_i], _reqs[_j]
            if _a['start_min'] < _b['end_min'] and _b['start_min'] < _a['end_min']:
                if _b['prid'] not in _a['match_request_list']:
                    _a['match_request_list'].append(_b['prid'])
                    _detected_double_ups += 1
                if _a['prid'] not in _b['match_request_list']:
                    _b['match_request_list'].append(_a['prid'])
print(f"Detected {_detected_double_ups} double-up link(s) via time overlap "
      f"(for requests where the export didn't already specify match_request_list)")

carer_shift_input = {}  # carer_name -> (cid, crid, start_min, end_min, gender)
for c in _real_caregivers_raw['caregivers']:
    carer = f"{c.get('name', '')} {c.get('lastname', '')}".strip()
    carer_shift_input[carer] = {
        'cid': c['cid'], 'crid': c['crid'],
        'start_min': c['shift']['start_time'], 'end_min': c['shift']['end_time'],
        'gender': c.get('gender'),
        'first_name': (c.get('name') or '').strip(), 'last_name': (c.get('lastname') or '').strip(),
    }

print(f"{len(day_requirements)} patient requirements, {len(carer_shift_input)} carer shifts extracted")

daily_carers_all = defaultdict(set)
for carer, wd_map in roster.items():
    for wd, visits in wd_map.items():
        for v in visits:
            daily_carers_all[v['start_dt'].date()].add(carer)
same_weekday_dates = [d for d in daily_carers_all if WEEKDAYS[d.weekday()] == TARGET_WEEKDAY]
recent_same_weekday = [d for d in same_weekday_dates if d >= TARGET_DATE - datetime.timedelta(weeks=8)]
avg_carers_recent = (sum(len(daily_carers_all[d]) for d in recent_same_weekday) / len(recent_same_weekday)
                      if recent_same_weekday else len(carer_shift_input))
today_carers_n = len(carer_shift_input)
staffing_ratio = today_carers_n / avg_carers_recent if avg_carers_recent else 1.0
day_is_short_staffed = staffing_ratio < 0.85
print(f"Today's carers: {today_carers_n}, recent-{TARGET_WEEKDAY}-average: {avg_carers_recent:.1f}, "
      f"ratio: {staffing_ratio:.2f} -> {'SHORT-STAFFED' if day_is_short_staffed else 'normal/adequate'}")

PRIORITY_BY_SERVICE_PRIORITY = {'Medium': 0.5, 'High': 0.75, 'Very High': 0.9}

print("\n" + "=" * 70)
print("Build PATIENT.JSON (only client+need-time trusted; rest recomputed from history)")
print("=" * 70)

patients = []
long_visit_flags = []

for req in day_requirements:
    client = req['client']
    info = client_info.get(client)
    prid = req['prid']
    if not info:
        # no match in active clients -- keep identity, use given time as-is, no history-based
        # refinement possible; extend_feasibility left wide open since we know nothing. The
        # export's own location IS trusted here (it's identity/address data, not analysis --
        # same category as name/gender) since our clients-new.json snapshot simply hasn't
        # caught up with a brand-new client yet; falling back to 0.0/0.0 would silently
        # discard perfectly good location data the export already gave us.
        _loc = req.get('location') or {}
        patients.append({
            'pid': req['pid'], 'prid': prid, 'gender': gender_enum(req['gender']),
            'name': req['first_name'], 'lastname': req['last_name'],
            'location_id': req['pid'],
            'location': {
                'latitude': _loc.get('latitude', 0.0) or 0.0,
                'longitude': _loc.get('longitude', 0.0) or 0.0,
                'postcode': _loc.get('postcode', '') or '',
            },
            'request_window': {
                'start_time_hard': max(req['start_min'] - 20, 0),
                'end_time_hard': min(req['end_min'] + 20, MAX_SCHEMA_MINUTE),
                'start_time_soft': req['start_min'], 'end_time_soft': req['end_min'],
                'duration': max(req['end_min'] - req['start_min'], 15),
                'min_duration': max(req['end_min'] - req['start_min'] - 15, 15),
                'duration_reduction_priority': 0.3, 'request_window_priority': 0.5,
                'soft_window_violation_level': 0.5,
                'match_request_list': req['match_request_list'],
            },
            'extend_feasibility': {
                'extend': True, 'max_distance_km': 20.0, 'max_time_minutes': 60,
                'max_distance_border_crossings_km': 10, 'max_time_border_crossings_minutes': 60,
            },
        })
        continue

    start_min, end_min = req['start_min'], req['end_min']
    raw_duration = end_min - start_min
    if raw_duration > SCHEMA_MAX_DURATION:
        long_visit_flags.append((prid, raw_duration))
        duration = SCHEMA_MAX_DURATION
        end_min = start_min + duration
    else:
        duration = max(raw_duration, 15)
    # Schema requires min_duration>=15 AND duration-min_duration>=10, so duration must be
    # at least 25 regardless of which specific min_duration formula is used below.
    if duration < 25:
        duration = 25
    # Data-driven min_duration: the shortest the solver may compress this visit to is 2/3 of
    # its duration (a 30-min call floors at 20, a 60-min call at 40, etc.) -- but never closer
    # than the schema's own 10-minute gap, which binds tighter for shorter visits, and never
    # below the schema's absolute floor of 15.
    min_duration = max(min(round(duration * 2 / 3), duration - 10), 15)

    # Soft window recomputed from historical ACTUAL times for this slot (not trusted from
    # the export) -- same methodology as run_all_in_one.py.
    best_key, best_diff = None, None
    for cidx, cluster in enumerate(client_slot_history.get((client, TARGET_WEEKDAY), [])):
        minutes = [dt.hour * 60 + dt.minute for _, dt, a_s in cluster]
        med = sorted(minutes)[len(minutes) // 2]
        diff = abs(med - start_min)
        if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
            best_key, best_diff = (client, TARGET_WEEKDAY, cidx), diff
    actual_pairs = client_slot_actual_times.get(best_key, []) if best_key else []
    if actual_pairs:
        actual_start_mins = [(a_s.hour * 60 + a_s.minute) for a_s, a_e in actual_pairs]
        actual_end_mins = [(a_e.hour * 60 + a_e.minute) for a_s, a_e in actual_pairs]
        p25_start = percentile(actual_start_mins, 25)
        p75_end = percentile(actual_end_mins, 75)
        start_soft = round(p25_start / 5) * 5
        end_soft = max(round(p75_end / 5) * 5, start_soft + duration)
    else:
        start_soft, end_soft = start_min, end_min
        if end_soft - start_soft < duration:
            end_soft = start_soft + duration

    hard_buffer = 20
    start_hard = max(min(start_soft - 15, start_soft - hard_buffer), 0)
    end_hard = min(max(end_soft + 15, end_soft + hard_buffer), MAX_SCHEMA_MINUTE)

    # duration_reduction_priority: mean actual/requirement duration ratio for this slot --
    # how much it's historically been compressed in practice.
    duration_pairs = client_slot_durations.get(best_key, []) if best_key else []
    if duration_pairs:
        ratios = [a / r for r, a in duration_pairs if r > 0]
        mean_ratio = sum(ratios) / len(ratios) if ratios else 1.0
        duration_reduction_priority = round(min(max(1 - mean_ratio, 0.0), 1.0), 2)
    else:
        duration_reduction_priority = 0.3

    # request_window_priority: driven directly by this slot's own operational cancellation
    # rate (VNR / Missed call / Covered by another agency / Management-Moira-Temporary
    # placeholder). service_priority is only a fallback when there's no cancellation history.
    cancel_rec = find_cancellation_record(client, TARGET_WEEKDAY, start_min)
    if cancel_rec and cancel_rec['cancellation_rate'] is not None:
        request_window_priority = round(min(max(1 - cancel_rec['cancellation_rate'], 0.05), 0.95), 2)
    else:
        request_window_priority = PRIORITY_BY_SERVICE_PRIORITY.get(info['service_priority'], 0.5)

    # soft_window_violation_level: std deviation of real historical ACTUAL start times --
    # low = reliably punctual (strict, more penalty if violated), high = historically
    # variable (more tolerance).
    if actual_pairs and len(actual_start_mins) >= 2:
        mean_m = sum(actual_start_mins) / len(actual_start_mins)
        variance = sum((m - mean_m) ** 2 for m in actual_start_mins) / len(actual_start_mins)
        std_dev = math.sqrt(variance)
        soft_window_violation_level = round(min(max(std_dev / 90, 0.0), 1.0), 2)
    else:
        soft_window_violation_level = 0.5

    all_carers_this_client_wd = set()
    for cluster in client_slot_history.get((client, TARGET_WEEKDAY), []):
        all_carers_this_client_wd.update(c for c, dt, a_s in cluster)
    rotates = len(all_carers_this_client_wd) >= 3

    patients.append({
        'pid': info['id'], 'prid': prid, 'gender': gender_enum(info['gender']),
        'name': info['first_name'], 'lastname': info['last_name'],
        'location_id': info['id'],
        'location': {'latitude': info['lat'], 'longitude': info['lon'], 'postcode': info['postcode']},
        'request_window': {
            'start_time_hard': start_hard, 'end_time_hard': end_hard,
            'start_time_soft': start_soft, 'end_time_soft': end_soft,
            'duration': duration, 'min_duration': min_duration,
            'duration_reduction_priority': duration_reduction_priority,
            'request_window_priority': request_window_priority,
            'soft_window_violation_level': soft_window_violation_level,
            'match_request_list': req['match_request_list'],
        },
        'extend_feasibility': {
            # Per the other chat's own design, patient-side extend is a flat default -- it's
            # never meant to conditionally restrict a patient's own feasibility (extension
            # permission is a CAREGIVER-side concept: whether she's allowed to travel outside
            # her usual range). Always True; only max_distance_km/max_time_minutes vary,
            # reflecting whether this specific client's slot has historically rotated across
            # multiple carers (needs a wider net) or stayed with one (tighter is fine).
            'extend': True,
            'max_distance_km': 15.0 if rotates else 5.0,
            'max_time_minutes': 45 if rotates else 20,
            'max_distance_border_crossings_km': 10,
            'max_time_border_crossings_minutes': 60,
        },
        '_slot_key': best_key,  # internal only -- stripped before writing; lets the
                                 # double-up tightening pass below look up this exact
                                 # visit's own historical actual-duration data
    })

print(f"Built {len(patients)} patient records")

# ---------------------------------------------------------------------------
# Tighten double-up windows so the solver can't schedule them sequentially.
#
# A double-up (match_request_list non-empty) needs BOTH carers physically present at the
# SAME time -- e.g. for a hoist transfer. But each member's window was computed
# independently above, from ITS OWN historical actual-time percentiles -- the two legs'
# windows can drift apart, and if the resulting window is wider than one visit's duration,
# the solver can legally schedule the two legs back-to-back instead of concurrently
# (cheaper on driving time, but not what a double-up means).
#
# Some double-ups are a genuine partial overlap, not a full one -- e.g. a 30min task nested
# inside a 90min visit, not two 90min visits at once. Forcing every member to the SAME
# duration (an earlier version of this fix) would wrongly stretch the 30min leg out to
# 90min. Instead: share the same WINDOW across the group (sized to the longest member's own
# duration), but leave every member's own `duration` exactly as computed -- untouched. The
# longest member's duration then exactly fills the shared window (zero slack), so THEY are
# pinned to occupy the whole window; any other member placed anywhere inside that same
# window necessarily overlaps them, however short their own visit is and wherever within
# the window the solver puts it. This holds regardless of whether the true historical
# overlap is at the start, middle, or end of the longer visit.
# ---------------------------------------------------------------------------
patient_by_prid = {p['prid']: p for p in patients}
seen_groups = set()
tightened = 0
for p in patients:
    others = p['request_window']['match_request_list']
    if not others:
        continue
    group_prids = frozenset([p['prid']] + others)
    if group_prids in seen_groups:
        continue
    seen_groups.add(group_prids)
    group = [patient_by_prid[prid] for prid in group_prids if prid in patient_by_prid]
    if len(group) < 2:
        continue

    # Window size = MEAN of the biggest member's own historical ACTUAL durations (how long
    # that visit genuinely tends to run in practice), not just its computed `duration` field
    # -- a more direct, real-world-grounded anchor for how tight the shared window can be.
    # Falls back to the `duration` field itself if this slot has no actual-duration history.
    # Never allowed to come out SMALLER than any member's own duration, though -- the mean
    # actual duration can legitimately be shorter than the stated duration (visits often run
    # under time), but the schema requires the window to fit every member's own duration
    # regardless, so the largest individual duration in the group is always a floor.
    biggest = max(group, key=lambda g: g['request_window']['duration'])
    dur_pairs = client_slot_durations.get(biggest.get('_slot_key'), []) if biggest.get('_slot_key') else []
    if dur_pairs:
        window_span = round(sum(a for r, a in dur_pairs) / len(dur_pairs))
    else:
        window_span = biggest['request_window']['duration']
    window_span = max(window_span, biggest['request_window']['duration'])
    if window_span - 15 < 10:  # same schema constraint as elsewhere: duration >= min_duration(15) + 10
        window_span = 25
    group_start_soft = min(g['request_window']['start_time_soft'] for g in group)
    group_end_soft = group_start_soft + window_span
    # Hard window margin: the schema now allows a tighter 5-minute margin specifically for
    # match_request_list (double-up) requests -- the general 15-minute minimum only applies
    # to ordinary single visits.
    group_start_hard = max(group_start_soft - 5, 0)
    group_end_hard = min(group_end_soft + 5, MAX_SCHEMA_MINUTE)

    for g in group:
        g['request_window']['start_time_soft'] = group_start_soft
        g['request_window']['end_time_soft'] = group_end_soft
        g['request_window']['start_time_hard'] = group_start_hard
        g['request_window']['end_time_hard'] = group_end_hard
        # duration is NOT touched -- each member keeps its own correct visit length.
        # min_duration uses the same data-driven 2/3-of-duration rule as everywhere else,
        # still bounded by the schema's own 10-minute gap and 15-minute floor. Note: for
        # visits >=30 min this can allow slightly MORE compression slack than the old
        # duration-10 floor did specifically for double-ups (which wanted the least possible
        # slack, to minimize the risk of a solver compressing one leg just enough to schedule
        # the two legs sequentially instead of concurrently) -- applied here for consistency
        # with the rule everywhere else, but worth revisiting if that risk shows up in practice.
        _d = g['request_window']['duration']
        g['request_window']['min_duration'] = max(min(round(_d * 2 / 3), _d - 10), 15)
    tightened += len(group)

if tightened:
    print(f"Tightened {tightened} double-up patient windows across {len(seen_groups)} groups "
          f"(forced identical, minimally-wide windows so they can't be scheduled sequentially)")
if long_visit_flags:
    print(f"NOTE: {len(long_visit_flags)} visit(s) exceeded 8h max duration, capped:")
    for prid, raw_dur in long_visit_flags:
        print(f"   {prid}: {raw_dur / 60:.1f}h -> capped to 8h")

print("\n" + "=" * 70)
print("Build CAREGIVERS.JSON (only shift time trusted; rest recomputed from history)")
print("=" * 70)

def weighted_percentile(values, weights, pct):
    """Weighted percentile (pct in [0,100]); values/weights are parallel lists."""
    if not values:
        return None
    pairs = sorted(zip(values, weights))
    total_weight = sum(w for v, w in pairs)
    if total_weight <= 0:
        return pairs[-1][0]
    target = total_weight * (pct / 100.0)
    cum = 0.0
    for v, w in pairs:
        cum += w
        if cum >= target:
            return v
    return pairs[-1][0]


def data_driven_extend_feasibility(carer, min_observations=1):
    """Ported from the '3 year history data' chat's own data_driven_extend_feasibility:
    derive a carer's real operating travel radius from her OWN historical trips (home ->
    every client she's actually visited, any weekday, any date), instead of a flat default
    or a generic search-radius heuristic. 'Normal' bounds use the recency-weighted 90th
    percentile of what she typically travels; 'border crossing' bounds are the EXTRA
    allowance beyond that -- the gap between her single furthest-ever trip and her normal
    90th-percentile range -- not a second independent cap. Recency weight per observation:
    0.5^(days_since_that_visit/180), so a visit from a year ago counts far less than one
    from last month, but nothing is hard-cut-off. min_observations=1: even a single real
    observation is more informative than an arbitrary flat default -- if she's only ever
    made one trip, that trip itself is both her 90th percentile AND her max, so the border
    allowance correctly comes out as 0 (no evidence she's ever gone further than that one
    trip), rather than a made-up 10km/60min. Returns None (caller falls back to the flat
    default) only if she has literally zero distance-matched visits at all."""
    obs = []
    for wd, visits in roster.get(carer, {}).items():
        for v in visits:
            client = v['client']
            d = carer_client_km.get((carer, client))
            t = carer_client_min.get((carer, client))
            if d is None or t is None:
                continue
            days_since = max((TARGET_DATE - v['start_dt'].date()).days, 0)
            obs.append((d, t, days_since))
    if len(obs) < min_observations:
        return None

    dists = [o[0] for o in obs]
    times = [o[1] for o in obs]
    weights = [0.5 ** (o[2] / 180.0) for o in obs]

    max_distance_km = min(max(weighted_percentile(dists, weights, 90), 0.0), 100.0)
    max_time_minutes = int(round(min(max(weighted_percentile(times, weights, 90), 0), 120)))
    # Border-crossing anchor: 95th percentile, not the literal single furthest-ever trip.
    # Using the true max means one rare outlier journey (e.g. a single one-off 60km cover)
    # can inflate the whole border allowance to an unrepresentative number. The 95th
    # percentile is still genuinely data-driven, but reflects her real "occasional further
    # trip" pattern rather than being entirely dictated by one extreme data point.
    p95_distance = weighted_percentile(dists, weights, 95)
    p95_time = weighted_percentile(times, weights, 95)
    border_distance_km = int(round(min(max(p95_distance - max_distance_km, 0), 50)))
    border_time_minutes = int(round(min(max(p95_time - max_time_minutes, 0), 120)))

    return {
        'max_distance_km': round(max_distance_km, 2),
        'max_time_minutes': max_time_minutes,
        'max_distance_border_crossings_km': border_distance_km,
        'max_time_border_crossings_minutes': border_time_minutes,
    }


caregivers = []
carer_to_crid = {}
for carer, shift in carer_shift_input.items():
    info = carer_info.get(carer)
    if not info or not info['lat']:
        print(f"WARNING: '{carer}' doesn't match an active carer with a home coordinate -- skipped.")
        continue
    shift_start = max(shift['start_min'], 0)
    shift_end = min(shift['end_min'], MAX_SCHEMA_MINUTE)
    crid = shift['crid']
    carer_to_crid[carer] = crid

    is_flexible = carer_does_extra.get(carer, False)
    extend = EXTEND_FEASIBILITY_OVERRIDES.get(carer, carer_extend_feasibility_flag(carer))

    data_driven = data_driven_extend_feasibility(carer)
    if data_driven:
        max_dist = data_driven['max_distance_km']
        max_time = data_driven['max_time_minutes']
        border_dist = data_driven['max_distance_border_crossings_km']
        border_time = data_driven['max_time_border_crossings_minutes']
    else:
        radius = carer_search_radius.get(carer, 10.0)
        if extend:
            max_dist, max_time = radius, 45
        else:
            max_dist = radius if day_is_short_staffed else 2.0
            max_time = 30 if day_is_short_staffed else 10
        border_dist, border_time = 10, 60

    travel_mode_map = {'Car': 'driving', 'Walk': 'walking'}
    caregivers.append({
        'cid': shift['cid'], 'crid': crid, 'gender': gender_enum(info['gender']),
        'name': info['first_name'], 'lastname': info['last_name'],
        'travel_mode': travel_mode_map.get(info['travel_method'], 'driving'),
        'location_id': info['id'],
        'location': {'latitude': info['lat'], 'longitude': info['lon'], 'postcode': info['postcode']},
        'current_location_id': info['id'], 'start_location_id': info['id'], 'end_location_id': info['id'],
        'shift': {'start_time': shift_start, 'end_time': shift_end},
        'extend_feasibility': {
            'extend': extend, 'max_distance_km': round(max_dist, 1), 'max_time_minutes': int(max_time),
            'max_distance_border_crossings_km': int(border_dist), 'max_time_border_crossings_minutes': int(border_time),
        },
        'caregiver_usage_priority': 0.75 if is_flexible else 0.5,
    })

carers_today = sorted(carer_to_crid.keys())
print(f"Built {len(caregivers)} caregiver records")
print(f"  Flexible (does extras): {sum(1 for c in carers_today if carer_does_extra.get(c))}")
print(f"  Classifier-flagged picky (extend=False): {len(confirmed_picky)}")

print("\n" + "=" * 70)
print("Build CRID_PRID_FEASIBILITY.JSON")
print("=" * 70)

WINDOW_DAYS = 112
STATUS_FACTORS = {'Current Primary': 1.0, 'Support / Relief': 0.5, 'Former / Relief': 0.2}


def identify_carer_status(overall_pct, days_since_last_visit):
    if days_since_last_visit > 50:
        return "Former / Relief"
    if overall_pct >= 40:
        return "Current Primary"
    return "Support / Relief"


slot_raw = {}  # key -> {carer: raw_score} -- UNNORMALIZED (see run_all_in_one.py for the
# full rationale). Normalizing per request, against only carers working today, means an
# absent regular carer's dominant historical score can't keep suppressing everyone else's
# weight on a day she isn't working at all.
slot_clusters_by_client_wd = defaultdict(list)
for (client, wd), clusters in client_slot_history.items():
    total_cust_visits = customer_totals.get(client, 0)
    if total_cust_visits <= 0:
        continue
    calls_per_day = total_cust_visits / float(WINDOW_DAYS)
    freq_factor = 1 + math.log1p(calls_per_day)
    for idx, cluster in enumerate(clusters):
        slot_total = len(cluster)
        by_carer = defaultdict(list)
        for carer, dt, a_s in cluster:
            by_carer[carer].append(a_s.date() if a_s is not None else dt.date())
        minutes = [dt.hour * 60 + dt.minute for _, dt, a_s in cluster]
        median_minute = sorted(minutes)[len(minutes) // 2]
        key = (client, wd, idx)
        slot_clusters_by_client_wd[(client, wd)].append((median_minute, key))
        raw_by_carer = {}
        for carer, dates in by_carer.items():
            last_visit = max(dates)
            days_since_last_visit = max((TARGET_DATE - last_visit).days, 0)
            overall_pct = round((len(dates) / slot_total) * 100, 1)
            status = identify_carer_status(overall_pct, days_since_last_visit)
            consistency = overall_pct / 100.0
            recency_decay = math.exp(-days_since_last_visit / 21.0)
            status_factor = STATUS_FACTORS.get(status, 0.3)
            raw = consistency * freq_factor * recency_decay * status_factor * concentration_factor(carer)
            raw_by_carer[carer] = raw
        slot_raw[key] = raw_by_carer

print(f"Computed carer affinity raw scores for {len(slot_raw)} historical (client, weekday, slot) combinations")


def find_slot_raw_scores(client, wd, start_minute):
    best_key, best_diff = None, None
    for median_minute, key in slot_clusters_by_client_wd.get((client, wd), []):
        diff = abs(median_minute - start_minute)
        if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
            best_key, best_diff = key, diff
    return slot_raw.get(best_key, {})


# weight = 2 means "the ONLY valid assignment" to the solver, not just "strongly preferred" --
# so it must be reserved for genuine exclusivity, not just high consistency. A carer whose
# long shift has plenty of spare room, and who does see other people even occasionally, must
# NOT get 2 for a routine slot -- that would tell the solver she can never be considered for
# anyone else, hiding real, lower-priority options for other patients she actually could take.
# weight = 2 requires Weekly consistency for this exact slot PLUS at least one of:
#   (a) on every historical occurrence of this slot, it was the ONLY client she saw that
#       day (the visit consumes her whole working day -- e.g. a 12h/18h call), or
#   (b) across her entire history, she has NEVER visited anyone outside her routine at all
#       (fully exclusive, e.g. a carer with exactly one client, ever).
# Weekly-but-not-exclusive carers fall through to the normal formula-based weight instead --
# which, given her real consistency, is usually still high, just not a hard lock.
#
# "Consumes her whole day" is measured by TIME DOMINANCE, not a strict "literally the only
# client" check -- real data has cases like a 12-hour visit alongside a genuinely separate
# 1-hour stop for a co-located/nearby client (e.g. two people sharing a home or facility).
# Requiring 100% exclusivity would wrongly disqualify that obviously-dominant pattern over a
# minor, consistently small addition. Instead: on the median historical occurrence of this
# slot, does it account for at least DOMINANCE_THRESHOLD of her total working time that day?
DOMINANCE_THRESHOLD = 0.7

carer_date_total_duration = defaultdict(float)
for _carer, _wd_map in roster.items():
    for _wd, _visits in _wd_map.items():
        for _v in _visits:
            _mins = (_v['end_dt'] - _v['start_dt']).total_seconds() / 60
            carer_date_total_duration[(_carer, _v['start_dt'].date())] += _mins


def carer_slot_consumes_whole_day(carer, client, wd):
    visits = [v for v in roster.get(carer, {}).get(wd, []) if v['client'] == client]
    if not visits:
        return False
    fractions = []
    for v in visits:
        date = v['start_dt'].date()
        this_duration = (v['end_dt'] - v['start_dt']).total_seconds() / 60
        total_duration = carer_date_total_duration.get((carer, date), 0.0)
        if total_duration > 0:
            fractions.append(this_duration / total_duration)
    if not fractions:
        return False
    fractions.sort()
    median_fraction = fractions[len(fractions) // 2]
    return median_fraction >= DOMINANCE_THRESHOLD


def carer_slot_typical_duration(carer, client, wd):
    """Median historical duration (minutes) of this carer's real visits for this exact
    (client, weekday) slot. Used to check that today's requested duration is plausibly the
    same commitment -- not a dramatically shorter, different task that just happens to start
    at the same time as her real, established slot (start-time matching alone can't tell
    those apart, since slot-matching only checks proximity in start time)."""
    durations = [(v['end_dt'] - v['start_dt']).total_seconds() / 60
                 for v in roster.get(carer, {}).get(wd, []) if v['client'] == client]
    if not durations:
        return None
    durations.sort()
    return durations[len(durations) // 2]


def carer_fully_exclusive(carer):
    routine = carer_routine_clients.get(carer, set())
    all_clients = carer_all_visited_clients.get(carer, set())
    return len(all_clients - routine) == 0


def carer_is_weekly_for_slot(carer, client, wd, start_minute):
    """Broad check: is she Weekly-classified for this exact slot at all (regardless of
    exclusivity)? Used to separate the routine (1.0/2.0) pool from the occasional pool."""
    for pattern, median_minute in per_carer_client_slot_pattern.get((carer, wd, client), []):
        if pattern == 'Weekly' and abs(median_minute - start_minute) <= TIME_GAP_MINUTES:
            return True
    return False


def carer_deserves_weight_2(carer, client, wd, start_minute):
    """Stricter check: Weekly for this slot AND exclusive (see carer_slot_consumes_whole_day
    / carer_fully_exclusive above) -- the 2.0 tier specifically."""
    if not carer_is_weekly_for_slot(carer, client, wd, start_minute):
        return False
    return carer_slot_consumes_whole_day(carer, client, wd) or carer_fully_exclusive(carer)


def carer_slot_last_occurrence(carer, client, wd, start_minute):
    """The most recent date this carer actually covered this exact slot, or None if she
    never has. Used to break ties when more than one carer is independently Weekly-eligible
    for the same slot -- only the most recently active one should hold the fixed 1.0/2.0
    tier; a carer who was consistent in the past but hasn't done this slot in months has
    likely moved on, and shouldn't sit at full confidence alongside someone still active."""
    best_key, best_diff = None, None
    for median_minute, key in slot_clusters_by_client_wd.get((client, wd), []):
        diff = abs(median_minute - start_minute)
        if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
            best_key, best_diff = key, diff
    if best_key is None:
        return None
    _, _, idx = best_key
    cluster = client_slot_history.get((client, wd), [])[idx]
    dates = [(a_s.date() if a_s is not None else dt.date()) for c, dt, a_s in cluster if c == carer]
    return max(dates) if dates else None


feasibility_pairs = []
seen_pairs = set()
# Tracks which carers got weight=2 specifically because THIS VISIT consumes their whole
# working day (carer_slot_consumes_whole_day) -- a genuine time/capacity constraint. This is
# deliberately separate from carer_fully_exclusive (has she ever served anyone besides this
# one client, in her whole career) -- that's a dedication PATTERN, not proof she's busy all
# day; a fully-exclusive carer with a short visit can genuinely still have capacity left.
# Only carers in this set get stripped from every other patient later.
_whole_day_committed_carers = set()


def add_pair(prid, crid, weight):
    key = (prid, crid)
    if key in seen_pairs:
        return
    seen_pairs.add(key)
    feasibility_pairs.append({'prid': prid, 'crid': crid, 'weight': round(min(max(weight, 0.0), 2.0), 2)})


# A feasibility pair exists ONLY for a carer and patient who have real history together --
# no invented geographic-proximity fallback. A patient nobody has ever seen may legitimately
# end up with zero feasibility pairs here; that's a correct outcome of this definition, not
# a gap to paper over.
#
# Three tiers per slot:
#   2.0 -- Weekly for this exact slot AND exclusive (whole-day-consuming, or she's never
#          seen anyone outside her routine at all).
#   1.0 -- Weekly for this exact slot but NOT exclusive -- her real, regular go-to, fixed at
#          1.0 rather than run through the formula, but she genuinely has spare capacity for
#          others too.
#   <1.0 -- everyone else with real history for this slot (occasional/substitute carers),
#          scored by the existing formula (consistency/freq/recency/status/concentration),
#          normalized to land strictly below whichever routine (1.0 or 2.0) carer is present
#          today. If NO routine carer is present today at all (she's off, or none exists),
#          this group falls back to the substitute-boost logic instead: a single available
#          cover still gets full 1.0, but multiple covers stay capped below the true
#          historical ceiling rather than one being arbitrarily crowned.
for p in patients:
    prid = p['prid']
    client = next((r['client'] for r in day_requirements if r['prid'] == prid), None)
    start_minute = p['request_window']['start_time_soft']

    raw_here = find_slot_raw_scores(client, TARGET_WEEKDAY, start_minute) if client in client_info else {}
    raw_today = {c: r for c, r in raw_here.items() if c in carer_to_crid}

    weekly_all = {c for c in raw_here if carer_is_weekly_for_slot(c, client, TARGET_WEEKDAY, start_minute)}
    weekly_today = weekly_all & set(raw_today.keys())

    # Today's requested duration must be a plausible match to a carer's real historical
    # pattern for this slot -- a 60-minute request shouldn't inherit the fixed tier (1.0 or
    # 2.0) of someone's real 720-minute (12h) commitment just because it happens to start at
    # the same time. Slot-matching only checks start-time proximity, so this is a necessary
    # second check. A carer failing this falls through to the normal formula-based pool,
    # where her genuine (thinner, more uncertain) affinity for this specific short task gets
    # a more honest score instead of inheriting a commitment that isn't really the same one.
    def _duration_plausible(c):
        _typical = carer_slot_typical_duration(c, client, TARGET_WEEKDAY)
        if _typical is None:
            return True
        return p['request_window']['duration'] >= _typical * 0.5

    _duration_implausible = {c for c in weekly_today if not _duration_plausible(c)}
    weekly_all = weekly_all - _duration_implausible
    weekly_today = weekly_today - _duration_implausible

    # If more than one carer independently qualifies as Weekly for this exact slot, only
    # the MOST RECENTLY active one keeps the fixed tier -- the others fall back into the
    # normal formula-based pool below (they may still score well there, just not as a
    # guaranteed 1.0/2.0). This prevents a carer who's effectively stopped covering a slot
    # months ago from sitting at full confidence alongside someone still actively doing it.
    if len(weekly_today) > 1:
        _dated = [(c, carer_slot_last_occurrence(c, client, TARGET_WEEKDAY, start_minute)) for c in weekly_today]
        _dated = [(c, d) for c, d in _dated if d is not None]
        if _dated:
            _most_recent = max(_dated, key=lambda x: x[1])[0]
            _demoted = weekly_today - {_most_recent}
            weekly_all = weekly_all - _demoted
            weekly_today = {_most_recent}

    weight_2_today = {c for c in weekly_today
                       if carer_slot_consumes_whole_day(c, client, TARGET_WEEKDAY) or carer_fully_exclusive(c)}
    for _c in weight_2_today:
        if carer_slot_consumes_whole_day(_c, client, TARGET_WEEKDAY):
            _whole_day_committed_carers.add(_c)
    weight_1_today = weekly_today - weight_2_today

    # Overly-broad carers (extreme generalists) don't get the fixed weight=1.0 tier -- demote
    # them back into the normal formula-based pool, where concentration_factor already
    # accounts for their breadth honestly. Genuine weight=2 exclusivity is untouched.
    _breadth_disqualified = {c for c in weight_1_today
                              if len(carer_all_visited_clients.get(c, set())) >= BREADTH_DISQUALIFY_THRESHOLD}
    weight_1_today = weight_1_today - _breadth_disqualified
    weekly_all = weekly_all - _breadth_disqualified

    others_today = {c: r for c, r in raw_today.items() if c not in weekly_all}
    others_here = {c: r for c, r in raw_here.items() if c not in weekly_all}

    _sole_substitute = False
    if weight_1_today or weight_2_today:
        # A routine carer (fixed 1.0 or 2.0) is present today -- everyone else normalizes
        # strictly below her.
        norm_base = max(raw_here[c] for c in (weight_1_today | weight_2_today))
    else:
        # No routine carer present today at all -- fall back to substitute-boost among the
        # occasional/non-weekly pool only.
        if others_here:
            main_other = max(others_here, key=others_here.get)
            overall_max_other = others_here[main_other]
        else:
            main_other, overall_max_other = None, 0.0
        if len(others_here) <= 1 or main_other in others_today:
            norm_base = max(others_today.values(), default=0.0)
        elif len(others_today) <= 1:
            norm_base = max(others_today.values(), default=0.0)
        else:
            norm_base = overall_max_other
        # She's the only real candidate at all today for this slot, with no established
        # routine carer to compare against -- genuinely useful, but this is a weaker signal
        # than either a real routine carer (1.0) or being ranked below one (up to 0.99), so
        # it gets its own, lower ceiling rather than sharing the general 0.99 cap.
        _sole_substitute = len(others_today) == 1

    _cap = 0.90 if _sole_substitute else 0.99
    weights_here = {
        # capped below 1.0 -- the schema only allows a weight strictly between 0.0 and 1.0,
        # or exactly 2.0, nothing in between and nothing tying the fixed 1.0 (routine) tier.
        # An "other" carer's raw score can occasionally exceed the routine carer's raw score
        # she's being normalized against (Weekly classification is about consistency within
        # her own active weeks, not raw score magnitude), so this cap is a real safety net.
        c: min(max(round(r / norm_base, 4), 0.01), _cap) if norm_base > 0 else 0.0
        for c, r in others_today.items()
    }

    # weight = 2 means exclusive -- a carer physically can't also be doing someone else's
    # visit at the same time, so if ANY carer gets 2.0 for this patient, she must be the
    # ONLY feasibility entry for it. No other carer gets listed alongside her.
    if weight_2_today:
        for carer in weight_2_today:
            crid = carer_to_crid.get(carer)
            if crid:
                add_pair(prid, crid, 2.0)
    else:
        for carer in weight_1_today:
            crid = carer_to_crid.get(carer)
            if crid:
                add_pair(prid, crid, 1.0)
        for carer, w in weights_here.items():
            crid = carer_to_crid.get(carer)
            if crid:
                add_pair(prid, crid, w)

# ---------------------------------------------------------------------------
# Cross-weekday fallback: if a patient ended up with ZERO feasible carers because their
# usual regular is absent today and nobody else has SAME-WEEKDAY history either, look at
# the client's OTHER weekdays to find who else has a real, substantial relationship with
# them -- and use the EXACT weight she'd genuinely hold on that other weekday (reusing the
# same tier logic: 2.0 exclusive / 1.0 routine / formula-based), rather than inventing a
# separate pooled-across-weekdays score. If she's the same person doing the same real job on
# a different day, her weight today should be exactly what it is on that day.
# ---------------------------------------------------------------------------
def best_weight_on_other_weekday(carer, client, exclude_wd):
    best = 0.0
    for wd in WEEKDAYS:
        if wd == exclude_wd:
            continue
        for pattern, median_minute in per_carer_client_slot_pattern.get((carer, wd, client), []):
            if pattern != 'Weekly':
                continue
            if carer_slot_consumes_whole_day(carer, client, wd) or carer_fully_exclusive(carer):
                best = max(best, 2.0)
            elif len(carer_all_visited_clients.get(carer, set())) < BREADTH_DISQUALIFY_THRESHOLD:
                best = max(best, 1.0)
            else:
                best = max(best, 0.99)
    return best


_prids_with_pairs = {r['prid'] for r in feasibility_pairs}
_zero_feasibility_prids = [p['prid'] for p in patients if p['prid'] not in _prids_with_pairs]

_cross_weekday_fallback_count = 0
for _prid in _zero_feasibility_prids:
    _client = next((r['client'] for r in day_requirements if r['prid'] == _prid), None)
    if not _client or _client not in client_info:
        continue
    _total_visits = Counter()
    for _wd_all in WEEKDAYS:
        for _cluster in client_slot_history.get((_client, _wd_all), []):
            for _carer, _dt, _a_s in _cluster:
                _total_visits[_carer] += 1
    if not _total_visits:
        continue
    _grand_total = sum(_total_visits.values())
    _ranked = _total_visits.most_common()
    _reference_carer, _reference_n = _ranked[0]
    _fallback_carer = next((c for c, n in _ranked if c in carer_to_crid), None)
    if not _fallback_carer:
        continue
    _crid = carer_to_crid.get(_fallback_carer)
    if not _crid:
        continue
    # First: does the actual fallback carer have a genuine Weekly relationship with this
    # client on ANY other weekday? If so, use that EXACT weight -- she's doing the same real
    # job, just on a different day.
    _other_day_weight = best_weight_on_other_weekday(_fallback_carer, _client, TARGET_WEEKDAY)
    if _other_day_weight > 0:
        _fallback_weight = _other_day_weight
    else:
        # She's not Weekly for this client on any weekday herself -- fall back to the
        # weaker, genuinely-thin-history case: does the TRUE reference carer's exclusivity
        # transfer, or does this need a proportional score from her own thin cross-weekday
        # history? (Same logic as before for this narrower case.)
        _fallback_carer_breadth = len(carer_all_visited_clients.get(_fallback_carer, set()))
        if carer_fully_exclusive(_reference_carer) and _fallback_carer_breadth < BREADTH_DISQUALIFY_THRESHOLD:
            _fallback_weight = 2.0
        else:
            _her_dates = [_v2['start_dt'].date() for _wd2, _visits2 in roster.get(_fallback_carer, {}).items()
                          for _v2 in _visits2 if _v2['client'] == _client]
            if _her_dates:
                _total_cust_visits = customer_totals.get(_client, 0)
                _last_visit = max(_her_dates)
                _days_since = max((TARGET_DATE - _last_visit).days, 0)
                _recency_decay = math.exp(-_days_since / 21.0)
                _calls_per_day = _total_cust_visits / float(WINDOW_DAYS)
                _freq_factor = 1 + math.log1p(_calls_per_day)
                _consistency = len(_her_dates) / max(_grand_total, 1)
                _cf = concentration_factor(_fallback_carer)
                _raw_fallback = _consistency * _freq_factor * _recency_decay * 0.5 * _cf
                _fallback_weight = min(max(round(_raw_fallback, 4), 0.01), 0.99)
            else:
                _fallback_weight = 0.01  # safety net; shouldn't normally happen
    add_pair(_prid, _crid, _fallback_weight)
    _cross_weekday_fallback_count += 1

if _cross_weekday_fallback_count:
    print(f"Cross-weekday fallback applied for {_cross_weekday_fallback_count} patient(s) with "
          f"zero same-weekday feasibility -- substituted with their most historically-involved "
          f"carer (from other weekdays) who is working today, as the sole candidate.")

# ---------------------------------------------------------------------------
# Cross-weekday fallback for CARERS: a carer working today can have real, substantial
# history with several of today's patients on OTHER weekdays, even with zero same-weekday
# matches (e.g. she simply doesn't normally work this weekday, but is scheduled today
# regardless). Add her as a supplementary candidate for those patients -- but ONLY when the
# patient doesn't already have a strong same-weekday match. A patient who already has their
# own real, dedicated carer today doesn't need extra candidates piled on from someone who's
# never even worked this weekday before; this is only meant to help patients who'd otherwise
# have thin or no real options at all.
# ---------------------------------------------------------------------------
STRONG_MATCH_THRESHOLD = 0.5  # existing weight at/above this = don't add a supplement

_carers_with_pairs = {r['crid'] for r in feasibility_pairs}
_today_carers_zero_pairs = [c for c, _crid in carer_to_crid.items() if _crid not in _carers_with_pairs]

_carer_fallback_count = 0
for _carer in _today_carers_zero_pairs:
    _her_clients = carer_all_visited_clients.get(_carer, set())
    for _p in patients:
        _prid = _p['prid']
        _client = next((r['client'] for r in day_requirements if r['prid'] == _prid), None)
        if _client not in _her_clients:
            continue
        _existing_rows = [r for r in feasibility_pairs if r['prid'] == _prid]
        _existing_max = max((r['weight'] for r in _existing_rows), default=0.0)
        if _existing_max >= STRONG_MATCH_THRESHOLD:
            continue  # patient already has a strong match -- no redundant candidates needed
        if _existing_max >= 2.0:
            continue  # weight=2 is exclusive -- never add alongside it (redundant given the
                       # check above, kept as an explicit safety net)

        _her_dates = [_v['start_dt'].date() for _wd_all, _visits in roster.get(_carer, {}).items()
                      for _v in _visits if _v['client'] == _client]
        if not _her_dates:
            continue
        _total_cust_visits = customer_totals.get(_client, 0)
        if _total_cust_visits <= 0:
            continue
        _last_visit = max(_her_dates)
        _days_since = max((TARGET_DATE - _last_visit).days, 0)
        _recency_decay = math.exp(-_days_since / 21.0)
        _calls_per_day = _total_cust_visits / float(WINDOW_DAYS)
        _freq_factor = 1 + math.log1p(_calls_per_day)
        _consistency = min(len(_her_dates) / max(_total_cust_visits, 1), 1.0)
        _cf = concentration_factor(_carer)
        _raw = _consistency * _freq_factor * _recency_decay * 0.5 * _cf

        _cap = round(_existing_max * 0.9, 4) if _existing_max > 0 else 0.99
        _w = min(max(round(_raw, 4), 0.01), _cap)
        _crid = carer_to_crid.get(_carer)
        if _crid:
            add_pair(_prid, _crid, _w)
            _carer_fallback_count += 1

if _carer_fallback_count:
    print(f"Cross-weekday fallback applied for {_carer_fallback_count} carer-patient pair(s) -- "
          f"carers working today with zero same-weekday matches but real cross-weekday history "
          f"were added as candidates ONLY for patients who didn't already have a strong "
          f"(>= {STRONG_MATCH_THRESHOLD}) same-weekday match.")

# ---------------------------------------------------------------------------
# DOUBLE-UP same-carer conflict: a true double-up (linked via match_request_list) needs TWO
# DIFFERENT carers at the SAME TIME. The same carer holding a "guaranteed" weight (1.0 or
# 2.0) on more than one leg of the same simultaneous group is a logical impossibility -- she
# can only actually do one of them. Keep her guaranteed status on whichever leg she has the
# strongest underlying raw affinity for, and demote her on the other leg(s) to just below the
# guaranteed tier (0.99), same treatment used everywhere else a carer doesn't hold a genuine
# guaranteed claim -- she's still a real, valid option there, just not simultaneously "the"
# routine carer for two things happening at once.
#
# Defined as a function and called TWICE: once here, and once more after the at-risk
# rotation is applied below -- the rotation step can independently assign the same carer to
# both legs of a double-up (it has no awareness of double-up grouping), so this must run
# again afterward as a guarantee, not rely on it happening to not occur.
# ---------------------------------------------------------------------------
def resolve_double_up_conflicts():
    _double_up_groups = set()
    for _r in day_requirements:
        if _r['match_request_list']:
            _double_up_groups.add(frozenset([_r['prid']] + _r['match_request_list']))

    _demoted = 0
    for _group in _double_up_groups:
        _group_prids = list(_group)
        _by_carer_in_group = defaultdict(list)
        for _row in feasibility_pairs:
            if _row['prid'] in _group_prids and _row['weight'] >= 1.0:
                _by_carer_in_group[_row['crid']].append(_row)
        for _crid, _rows in _by_carer_in_group.items():
            if len(_rows) < 2:
                continue
            _carer_name = next((c for c, cr in carer_to_crid.items() if cr == _crid), None)

            def _rank_key(_row):
                _client_ = next((r['client'] for r in day_requirements if r['prid'] == _row['prid']), None)
                _start_min_ = next((p['request_window']['start_time_soft'] for p in patients if p['prid'] == _row['prid']), 0)
                _raw_here_ = find_slot_raw_scores(_client_, TARGET_WEEKDAY, _start_min_) if _carer_name else {}
                _r_ = _raw_here_.get(_carer_name, 0)
                return (_row['weight'], _r_)  # weight tier first (2.0 beats 1.0), raw score as tiebreaker

            _best_row = max(_rows, key=_rank_key)
            for _row in _rows:
                if _row is not _best_row:
                    _row['weight'] = 0.99
                    _demoted += 1
    return _demoted


_double_up_demoted = resolve_double_up_conflicts()
if _double_up_demoted:
    print(f"Demoted {_double_up_demoted} pair(s) where the same carer held a guaranteed "
          f"(>=1.0) weight on multiple legs of the same simultaneous double-up -- kept her "
          f"guaranteed status on her strongest leg only.")

# ---------------------------------------------------------------------------
# CARER-WIDE weight=2 exclusivity -- but ONLY when weight=2 came from genuinely consuming
# her whole working day (carer_slot_consumes_whole_day), not merely from being fully
# exclusive to one client over her whole career (carer_fully_exclusive). Those are different
# things: a carer who's never diversified beyond one client, but whose visit with them is
# short, can genuinely still have real capacity for other patients that day -- her weight=2
# reflects a career-long DEDICATION pattern, not a same-day TIME constraint, so stripping her
# from everyone else would be wrong. Only carers whose weight=2 is a real time-capacity claim
# get stripped from other patients; the main loop above already enforces exclusivity from the
# PATIENT's side regardless (a patient with a weight=2 carer never shows anyone else).
# ---------------------------------------------------------------------------
_whole_day_committed_crids = {carer_to_crid[c] for c in _whole_day_committed_carers if c in carer_to_crid}

_crid_to_twos = defaultdict(list)
for r in feasibility_pairs:
    if r['weight'] == 2.0:
        _crid_to_twos[r['crid']].append(r['prid'])

_stripped_count = 0
for _crid, _two_prids in _crid_to_twos.items():
    if len(_two_prids) > 1:
        print(f"WARNING: crid {_crid} holds weight=2.0 for MULTIPLE patients today "
              f"({_two_prids}) -- a genuine scheduling conflict (she can't be exclusively "
              f"committed to more than one at once). Left as-is for manual review rather "
              f"than silently picking one.")
    if _crid not in _whole_day_committed_crids:
        continue  # career-long exclusivity only -- not proof she's busy all day, don't strip
    _before = len(feasibility_pairs)
    feasibility_pairs = [r for r in feasibility_pairs
                          if r['crid'] != _crid or r['prid'] in _two_prids]
    _stripped_count += _before - len(feasibility_pairs)

if _stripped_count:
    print(f"Stripped {_stripped_count} lower-weight pair(s) belonging to carers whose "
          f"weight=2.0 elsewhere today genuinely consumes their whole working day.")

for dislike_client, dislike_carer in DISLIKES:
    dislike_prids = {r['prid'] for r in day_requirements if r['client'] == dislike_client}
    dislike_crid = carer_to_crid.get(dislike_carer)
    if not dislike_prids or not dislike_crid:
        continue
    feasibility_pairs = [fp for fp in feasibility_pairs
                          if not (fp['prid'] in dislike_prids and fp['crid'] == dislike_crid)]
    seen_pairs = {k for k in seen_pairs if not (k[0] in dislike_prids and k[1] == dislike_crid)}
    for prid in dislike_prids:
        add_pair(prid, dislike_crid, 0.0)
    print(f"Applied dislike: '{dislike_client}' x '{dislike_carer}' ({len(dislike_prids)} request(s))")

# ---------------------------------------------------------------------------
# Apply today's at-risk rotation (see AT_RISK_ROTATION above): for each departed carer with
# an automatically-selected replacement for today's weekday, that replacement takes
# weight=1.0 for every one of THAT departed carer's at-risk clients with a request today
# (demoting any other guaranteed claim on those same requests to 0.99), and her weight for
# every OTHER patient today is removed -- she's now committed to covering that rotation
# today, so she shouldn't also sit at full, undiminished availability elsewhere.
#
# Done in TWO passes: the same carer can be the selected rotation carer for more than one
# departed carer on the same day (e.g. she covers both Eimear Daly's and Ijeoma Iwuala's
# Monday clients). Assigning and removing one departed carer at a time would let a LATER
# departed carer's removal step wipe out an assignment an EARLIER one just made, since each
# only protects its own clients. Instead: assign everyone first (pass 1), then remove each
# rotation carer's other weight exactly once (pass 2), protecting the FULL cumulative set of
# at-risk requests she's now covering across every departed carer assigned to her today.
# ---------------------------------------------------------------------------
_rotation_carer_protected_prids = defaultdict(set)  # crid -> set of protected prids today
_rotation_carer_can_cover_both = {}  # crid -> True only if ALL her rotation assignments today show she historically also did her own routine work on covering days


def _carer_covers_both_historically(carer, dep_carer, wd):
    """Look at the real historical dates this carer covered dep_carer's at-risk clients on
    this weekday, and check whether the MAJORITY of those dates also show her doing at least
    one of her own routine (set-roster) visits that same day. This is empirical evidence of
    whether covering has historically meant she also kept up her own routine, or historically
    meant dropping it -- used to decide whether her other feasibility should be preserved or
    stripped when she's assigned this rotation today."""
    dep_clients = {c for c, dc in _at_risk_clients.items() if dc == dep_carer}
    covering_dates = set()
    for client in dep_clients:
        for cluster in client_slot_history.get((client, wd), []):
            for c, dt, a_s in cluster:
                if c == carer:
                    covering_dates.add(dt.date())
    if not covering_dates:
        return False  # no historical basis -- default to the safe, conservative strip
    her_routine_clients = set_roster.get((carer, wd), set())
    both_count = 0
    for d in covering_dates:
        that_day_visits = [v for v in roster.get(carer, {}).get(wd, []) if v['start_dt'].date() == d]
        if any(v['client'] in her_routine_clients for v in that_day_visits):
            both_count += 1
    return both_count / len(covering_dates) > 0.5


for _dep_carer in _departed_carers:
    if _dep_carer in carer_to_crid:
        # She IS actually working today, despite having a termination date within the
        # window (e.g. it's upcoming, not yet arrived) -- she's normally covering her own
        # clients today, so no rotation is needed for her specifically today.
        continue
    _rotation_carer = AT_RISK_ROTATION.get((_dep_carer, TARGET_WEEKDAY))
    if not _rotation_carer:
        continue
    _rotation_crid = carer_to_crid.get(_rotation_carer)
    if not _rotation_crid:
        continue
    _dep_clients_today = {c for c, dc in _at_risk_clients.items() if dc == _dep_carer}
    # Only assign the rotation carer to the SPECIFIC at-risk clients she genuinely has real
    # history with -- pool eligibility only required 2+ distinct clients overall, so she may
    # not have ever seen every one of the departed carer's clients. Assigning her weight=1.0
    # for someone she's literally never met would violate the "no feasibility without real
    # history" rule that holds everywhere else in this pipeline.
    _rotation_carer_real_clients = carer_all_visited_clients.get(_rotation_carer, set())
    _at_risk_prids_today = {r['prid'] for r in day_requirements
                             if r['client'] in _dep_clients_today
                             and r['client'] in _rotation_carer_real_clients}
    _skipped_clients = _dep_clients_today - _rotation_carer_real_clients
    if _skipped_clients:
        print(f"  NOTE: '{_rotation_carer}' has no real history with {sorted(_skipped_clients)} "
              f"(also at-risk for {_dep_carer}) -- not assigned to those specifically.")
    if not _at_risk_prids_today:
        continue
    _covered = 0
    for _prid in _at_risk_prids_today:
        _found = False
        for _r in feasibility_pairs:
            if _r['prid'] == _prid and _r['crid'] == _rotation_crid:
                _r['weight'] = 1.0
                _found = True
            elif _r['prid'] == _prid and _r['weight'] >= 1.0:
                _r['weight'] = 0.99
        if not _found:
            add_pair(_prid, _rotation_crid, 1.0)
        _covered += 1
    _rotation_carer_protected_prids[_rotation_crid].update(_at_risk_prids_today)

    _covers_both = _carer_covers_both_historically(_rotation_carer, _dep_carer, TARGET_WEEKDAY)
    if _rotation_crid not in _rotation_carer_can_cover_both:
        _rotation_carer_can_cover_both[_rotation_crid] = _covers_both
    else:
        _rotation_carer_can_cover_both[_rotation_crid] = _rotation_carer_can_cover_both[_rotation_crid] and _covers_both

    print(f"Assigned at-risk rotation for {_dep_carer} on {TARGET_WEEKDAY}: '{_rotation_carer}' takes "
          f"weight=1.0 for {_covered} at-risk request(s) today "
          f"({'historically also covers her own routine on these days' if _covers_both else 'historically drops her own routine on these days'}).")

for _rotation_crid, _protected_prids in _rotation_carer_protected_prids.items():
    # A genuine weight=2.0 she independently holds elsewhere (her own real exclusive
    # relationship with a different patient, unrelated to the rotation) is NOT wiped out
    # like her other, weaker pairs -- the rotation assignment is never itself a genuine
    # exclusive claim (always weight=1.0), so if she's now covering a call via rotation
    # while also genuinely exclusive to someone else, neither can honestly claim full
    # exclusivity anymore. Downgrade her own weight=2.0 pair(s) to match at 1.0, and mark
    # her as no longer available for extra/new work today, since she's now genuinely
    # committed across two different patients rather than exclusively bound to just one.
    _own_two_downgraded = 0
    for r in feasibility_pairs:
        if r['crid'] == _rotation_crid and r['prid'] not in _protected_prids and r['weight'] == 2.0:
            r['weight'] = 1.0
            _own_two_downgraded += 1
    if _own_two_downgraded:
        print(f"Downgraded {_own_two_downgraded} independent weight=2.0 pair(s) for crid "
              f"{_rotation_crid} to 1.0 -- she's now also covering an at-risk rotation "
              f"call today, so she can't be treated as exclusively bound to just one patient.")
        _out_c = next((c for c in caregivers if c['crid'] == _rotation_crid), None)
        if _out_c is not None and _out_c['extend_feasibility']['extend']:
            _out_c['extend_feasibility']['extend'] = False
            print(f"  Forced extend=False for crid {_rotation_crid} (now genuinely committed "
                  f"across two different patients today).")

    if _rotation_carer_can_cover_both.get(_rotation_crid):
        print(f"crid {_rotation_crid} historically covers rotation calls while also keeping up her "
              f"own routine -- her other feasibility today is left untouched.")
        continue

    _to_remove_prids = {r['prid'] for r in feasibility_pairs
                         if r['crid'] == _rotation_crid and r['prid'] not in _protected_prids and r['weight'] != 1.0}
    _removed = len(_to_remove_prids)
    feasibility_pairs = [r for r in feasibility_pairs
                          if not (r['crid'] == _rotation_crid and r['prid'] in _to_remove_prids)]
    seen_pairs = {k for k in seen_pairs if not (k[1] == _rotation_crid and k[0] in _to_remove_prids)}
    if _removed:
        print(f"Removed weight for {_removed} other patient(s) today for crid {_rotation_crid} "
              f"(now committed to {len(_protected_prids)} at-risk request(s) total across all "
              f"departed carers she's covering today).")

# Re-run the double-up conflict check: the at-risk rotation above has no awareness of
# double-up grouping, so it could independently assign the same rotation carer to both legs
# of a simultaneous double-up. Guarantee the invariant holds regardless.
_double_up_demoted_2 = resolve_double_up_conflicts()
if _double_up_demoted_2:
    print(f"Demoted {_double_up_demoted_2} pair(s) after at-risk rotation where the same carer "
          f"held a guaranteed weight on multiple legs of the same simultaneous double-up.")

print(f"\nBuilt {len(feasibility_pairs)} feasibility pairs across {len(patients)} patient requests")
print(f"Average feasible caregivers per request: {len(feasibility_pairs) / len(patients):.1f}")

all_prids = set(p['prid'] for p in patients)
prids_with_options = set(r['prid'] for r in feasibility_pairs)
zero = all_prids - prids_with_options
print(f"Patient requests with ZERO feasible carers: {len(zero)}" + (f"  {list(zero)[:10]}" if zero else ""))

double_up_prids = set()
for r in day_requirements:
    if r['match_request_list']:
        double_up_prids.add(r['prid'])
        double_up_prids.update(r['match_request_list'])
missing_double_up = double_up_prids - prids_with_options
print(f"Double-up members with ZERO feasible carers: {len(missing_double_up)}")

weight_counts = Counter(round(p['weight'], 2) for p in feasibility_pairs)
print("Weight distribution (sample):", dict(sorted(weight_counts.items())[-10:]))

# strip the internal-only tracking field before validation/output -- not part of the schema
for p in patients:
    p.pop('_slot_key', None)

print("\n" + "=" * 70)
print("Validate against hhs.py schema and write output")
print("=" * 70)

sys.path.insert(0, HHS_SCHEMA_PATH)
from hhs import Patient as PatientModel, Caregiver as CaregiverModel, FeasibilityPair as FeasibilityPairModel

validation_errors = []
for p in patients:
    try:
        PatientModel(**p)
    except Exception as e:
        validation_errors.append(('patient', p.get('prid'), str(e)[:200]))
for c in caregivers:
    try:
        CaregiverModel(**c)
    except Exception as e:
        validation_errors.append(('caregiver', c.get('crid'), str(e)[:200]))
for fp in feasibility_pairs:
    try:
        FeasibilityPairModel(**fp)
    except Exception as e:
        validation_errors.append(('feasibility', (fp.get('prid'), fp.get('crid')), str(e)[:200]))

print(f"Schema validation: {len(validation_errors)} errors out of "
      f"{len(patients) + len(caregivers) + len(feasibility_pairs)} records")
for e in validation_errors[:10]:
    print(' ', e)

with open(f'{OUTPUT_DIR}/patient.json', 'w') as f:
    json.dump(patients, f, indent=2)
with open(f'{OUTPUT_DIR}/caregivers.json', 'w') as f:
    json.dump(caregivers, f, indent=2)
with open(f'{OUTPUT_DIR}/crid_prid_feasible.json', 'w') as f:
    json.dump(feasibility_pairs, f, indent=2)

print(f"\nSaved patient.json, caregivers.json, crid_prid_feasible.json to {OUTPUT_DIR}/ "
      f"(for {TARGET_DATE.isoformat()}) -- all recomputed from history, using only "
      f"identity and need/availability time from your export.")

print("\n" + "=" * 70)
print("Build per-carer DECISION ANALYSIS workbook")
print("=" * 70)
print("Same foundation and same in-memory results as above -- this is the visible")
print("reasoning behind every number already written to patient.json/caregivers.json/")
print("crid_prid_feasible.json, not a separate analysis.")

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Rebuild the slot-weight BREAKDOWN (same formula as the feasibility step above, but keeping
# every intermediate value -- consistency %, status, recency decay, concentration factor --
# instead of just the final raw score) so the workbook can show the full "why" per pair.
# ---------------------------------------------------------------------------
def identify_carer_status_display(overall_pct, days_since_last_visit):
    if days_since_last_visit > 50:
        return "Former / Relief"
    if overall_pct >= 40:
        return "Current Primary"
    return "Support / Relief"

slot_breakdown = {}  # key -> {carer: {breakdown dict}}
slot_clusters_by_client_wd_display = defaultdict(list)
for (client, wd), clusters in client_slot_history.items():
    total_cust_visits = customer_totals.get(client, 0)
    if total_cust_visits <= 0:
        continue
    calls_per_day = total_cust_visits / float(WINDOW_DAYS)
    freq_factor = 1 + math.log1p(calls_per_day)
    for idx, cluster in enumerate(clusters):
        slot_total = len(cluster)
        by_carer = defaultdict(list)
        for carer, dt, a_s in cluster:
            by_carer[carer].append(a_s.date() if a_s is not None else dt.date())
        minutes = [dt.hour * 60 + dt.minute for _, dt, a_s in cluster]
        median_minute = sorted(minutes)[len(minutes) // 2]
        key = (client, wd, idx)
        slot_clusters_by_client_wd_display[(client, wd)].append((median_minute, key))
        entry = {}
        for carer, dates in by_carer.items():
            last_visit = max(dates)
            days_since_last_visit = max((TARGET_DATE - last_visit).days, 0)
            overall_pct = round((len(dates) / slot_total) * 100, 1)
            status = identify_carer_status_display(overall_pct, days_since_last_visit)
            consistency = overall_pct / 100.0
            recency_decay = math.exp(-days_since_last_visit / 21.0)
            status_factor = STATUS_FACTORS.get(status, 0.3)
            cf = concentration_factor(carer)
            raw = consistency * freq_factor * recency_decay * status_factor * cf
            entry[carer] = {
                'consistency_pct': overall_pct, 'status': status, 'days_since': days_since_last_visit,
                'freq_factor': freq_factor, 'recency_decay': recency_decay,
                'status_factor': status_factor, 'concentration_factor': cf, 'raw': raw,
            }
        slot_breakdown[key] = entry

def find_slot_breakdown(client, wd, start_minute):
    best_key, best_diff = None, None
    for median_minute, key in slot_clusters_by_client_wd_display.get((client, wd), []):
        diff = abs(median_minute - start_minute)
        if diff <= TIME_GAP_MINUTES and (best_diff is None or diff < best_diff):
            best_key, best_diff = key, diff
    return slot_breakdown.get(best_key, {})

# Use the SAME in-memory results already written to output/ above -- no re-reading from disk.
pid_to_client = {r['pid']: r['client'] for r in day_requirements}
caregiver_by_crid = {c['crid']: c for c in caregivers}
feas_by_crid = defaultdict(list)
for row in feasibility_pairs:
    feas_by_crid[row['crid']].append(row)
prid_to_pid_display = {r['prid']: r['pid'] for r in day_requirements}

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

FONT = 'Arial'
header_fill = PatternFill('solid', fgColor='2F5597')
header_font = Font(name=FONT, bold=True, color='FFFFFF')
title_font = Font(name=FONT, bold=True, size=14)
section_font = Font(name=FONT, bold=True, size=12)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color='808080')
good_fill = PatternFill('solid', fgColor='E2EFDA')
warn_fill = PatternFill('solid', fgColor='FCE4D6')

notes_ws = wb2.create_sheet('Read Me')
notes = [
    f"Carer decision analysis -- {TARGET_DATE} ({TARGET_WEEKDAY})",
    "",
    "One sheet per carer working today. This is the visible reasoning behind every number "
    "in caregivers.json and crid_prid_feasible.json for this same day -- not a separate "
    "analysis, the same one, exposed.",
    "",
    "Each sheet has 3 sections:",
    "1. EXTEND_FEASIBILITY DECISION -- the classifier's verdict (open/closed to unfamiliar "
    "clients) and the actual evidence behind it: how many routine (Weekly) vs off-routine "
    "clients she's ever had, her off-routine rate compared to the peer median, whether "
    "she's too new to judge yet, or whether she's geographically isolated (extends "
    "regardless, since there's no real alternative around her).",
    "2. DATA-DRIVEN TRAVEL BOUNDS -- her max_distance_km/max_time_minutes (recency-weighted "
    "90th percentile of her own real trips) and border-crossing allowance (95th percentile "
    "gap beyond that), with the number of real observations behind each.",
    "3. TODAY'S FEASIBILITY ASSIGNMENTS -- every patient she's a candidate for today, her "
    "final weight, and the full breakdown that produced it (consistency %, status, "
    "recency decay, concentration factor).",
    "",
    f"Day situation: {today_carers_n} carers worked today vs a {avg_carers_recent:.1f} "
    f"recent-{TARGET_WEEKDAY} average (ratio {staffing_ratio:.2f}) -- "
    f"{'SHORT-STAFFED' if day_is_short_staffed else 'a normal/adequate day'}.",
]
for i, line in enumerate(notes, start=1):
    c = notes_ws.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, bold=(i == 1), size=13 if i == 1 else 10)
notes_ws.column_dimensions['A'].width = 135

summary_ws = wb2.create_sheet('Summary')
summary_headers = ['Carer', 'Extend?', 'Why', 'Max Distance (km)', 'Max Time (min)',
                    'Border Distance (km)', 'Border Time (min)', 'Today\'s Assignments']
for col, h in enumerate(summary_headers, start=1):
    cell = summary_ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

def sheet_name(name, used):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    base = name[:28].strip() or 'Unnamed'
    cand = base
    i = 2
    while cand in used:
        cand = f"{base[:25]}_{i}"
        i += 1
    used.add(cand)
    return cand

used_names = set()
row_i = 2

for carer in carers_today:
    shift = carer_shift_input[carer]
    crid = shift['crid']
    out_c = caregiver_by_crid.get(crid)
    if not out_c:
        continue
    ef = out_c['extend_feasibility']

    total_nearby = carer_nearby_total_count.get(carer, 0)
    coverage = carer_nearby_coverage_ratio.get(carer, 0)
    is_isolated = total_nearby <= ISOLATED_TOTAL_THRESHOLD and coverage >= ISOLATED_COVERAGE_THRESHOLD
    rate = carer_off_routine_rate.get(carer)
    n_routine = len(carer_routine_clients.get(carer, set()))
    n_off_routine = len(carer_all_visited_clients.get(carer, set())) - n_routine

    if is_isolated:
        classifier_reason = (
            f"ISOLATED: only {total_nearby} client(s) reachable to her at all "
            f"(bottom quartile, threshold={ISOLATED_TOTAL_THRESHOLD}), and she already covers "
            f"{coverage:.0%} of them -- extends regardless of her own rate, since there's "
            f"realistically nowhere left for her to expand to."
        )
    elif rate is None:
        classifier_reason = "NEW CARER: under 90 days' tenure -- not enough history to judge yet, defaults open."
    else:
        verdict = "AT/ABOVE median -- extends" if rate >= MEDIAN_OFF_ROUTINE_RATE else "BELOW median -- does not extend"
        classifier_reason = (
            f"{n_routine} routine (Weekly) client(s), {n_off_routine} off-routine client(s) ever -- "
            f"rate {rate:.2f} per {CASELOAD_WINDOW_DAYS}-day period vs peer median "
            f"{MEDIAN_OFF_ROUTINE_RATE:.2f} -- {verdict}."
        )

    obs = []
    for wd2, visits2 in roster.get(carer, {}).items():
        for v2 in visits2:
            d2 = carer_client_km.get((carer, v2['client']))
            t2 = carer_client_min.get((carer, v2['client']))
            if d2 is not None and t2 is not None:
                obs.append((d2, t2))
    n_obs = len(obs)

    summary_ws.cell(row=row_i, column=1, value=carer).font = normal_font
    summary_ws.cell(row=row_i, column=2, value=ef['extend']).font = normal_font
    summary_ws.cell(row=row_i, column=3, value=classifier_reason).font = note_font
    summary_ws.cell(row=row_i, column=4, value=ef['max_distance_km']).font = normal_font
    summary_ws.cell(row=row_i, column=5, value=ef['max_time_minutes']).font = normal_font
    summary_ws.cell(row=row_i, column=6, value=ef['max_distance_border_crossings_km']).font = normal_font
    summary_ws.cell(row=row_i, column=7, value=ef['max_time_border_crossings_minutes']).font = normal_font
    summary_ws.cell(row=row_i, column=8, value=len(feas_by_crid.get(crid, []))).font = normal_font
    row_i += 1

    ws = wb2.create_sheet(sheet_name(carer, used_names))
    ws['A1'] = carer
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    r = 3

    ws.cell(row=r, column=1, value='1. EXTEND_FEASIBILITY DECISION').font = section_font
    r += 1
    ws.cell(row=r, column=1, value=f"Verdict: extend = {ef['extend']}").font = normal_font
    r += 1
    ws.cell(row=r, column=1, value=classifier_reason).font = note_font
    r += 2

    ws.cell(row=r, column=1, value='2. DATA-DRIVEN TRAVEL BOUNDS').font = section_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"max_distance_km={ef['max_distance_km']}, max_time_minutes={ef['max_time_minutes']} "
                   f"(recency-weighted 90th percentile of her {n_obs} real historical trips)")).font = normal_font
    r += 1
    ws.cell(row=r, column=1,
            value=(f"border_crossing: {ef['max_distance_border_crossings_km']}km / "
                   f"{ef['max_time_border_crossings_minutes']}min (95th percentile gap beyond her normal range)")).font = normal_font
    r += 2

    ws.cell(row=r, column=1, value="3. TODAY'S FEASIBILITY ASSIGNMENTS").font = section_font
    r += 1
    headers = ['Client', 'Weight', 'Consistency %', 'Status', 'Days Since Last Visit',
               'Recency Decay', 'Concentration Factor']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    r += 1
    rows_here = sorted(feas_by_crid.get(crid, []), key=lambda x: -x['weight'])
    for row in rows_here:
        pid = prid_to_pid_display.get(row['prid'])
        client = pid_to_client.get(pid, '?')
        req = next((rq for rq in day_requirements if rq['prid'] == row['prid']), None)
        start_min = req['start_min'] if req else 0
        bd = find_slot_breakdown(client, TARGET_WEEKDAY, start_min).get(carer)
        vals = [client, row['weight']]
        if row['weight'] == 2.0:
            vals += ['ROUTINE (Weekly) carer for this exact slot -- fixed at 2.0', '', '', '', '']
        elif bd:
            vals += [f"{bd['consistency_pct']}%", bd['status'], bd['days_since'],
                      round(bd['recency_decay'], 3), round(bd['concentration_factor'], 3)]
        else:
            vals += ['(no direct history found for breakdown)', '', '', '', '']
        fill = PatternFill('solid', fgColor='FFE699') if row['weight'] == 2.0 else (
            good_fill if row['weight'] >= 0.8 else (warn_fill if row['weight'] <= 0.3 else None))
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = normal_font
            if fill:
                cell.fill = fill
        r += 1

    widths = [30, 10, 14, 16, 20, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

summary_ws.column_dimensions['A'].width = 28
summary_ws.column_dimensions['C'].width = 70
for col in 'DEFGH':
    summary_ws.column_dimensions[col].width = 16
summary_ws.freeze_panes = 'A2'

decision_wb_path = f'{OUTPUT_DIR}/Carer_Decision_Analysis_{TARGET_DATE.isoformat()}.xlsx'
wb2.save(decision_wb_path)
print(f"Saved {decision_wb_path} ({len(wb2.sheetnames)} sheets)")

# ---------------------------------------------------------------------------
# Today's distance files (driving/walking/cycling, pid_cid keyed) -- only if your day
# export included distances.json. Skipped with a note if it's not present.
# ---------------------------------------------------------------------------
print("\n" + "=" * 70)
print("Today's distance files")
print("=" * 70)

TODAY_OUTPUT_DIR = f'{PROJECT_ROOT}/today'
_dist_path = f'{DAY_EXPORT_DIR}/distances.json'  # <- edit filename here if yours differs
if os.path.isfile(_dist_path):
    os.makedirs(TODAY_OUTPUT_DIR, exist_ok=True)
    with open(_dist_path) as f:
        _dist_raw = json.load(f)
    for mode in ['driving', 'walking', 'cycling']:
        if mode not in _dist_raw.get('distances', {}):
            print(f"NOTE: '{mode}' not present in distances.json -- skipped.")
            continue
        _out = _dist_raw['distances'][mode]
        _out_path = f'{TODAY_OUTPUT_DIR}/{mode}_data.json'
        with open(_out_path, 'w') as f:
            json.dump(_out, f)
        print(f"Saved {_out_path}: {len(_out['distances'])} entries")
else:
    print(f"NOTE: '{_dist_path}' not found -- skipping today's distance files "
          f"(only needed if your day export includes distances.json).")
